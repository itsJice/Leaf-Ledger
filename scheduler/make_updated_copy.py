#!/usr/bin/env python3
"""
Make a copy of the ORIGINAL client spreadsheet with:
  - every original detail preserved (all sheets, all columns),
  - the Area (col 6) + Zone (col 7) columns standardized to the recomputed
    uniform values,
  - three new columns appended: 2026 Install Date, Crew, Order.
"""
import datetime
import json
import os
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
SRC = os.path.join(HERE, "2025 CHRISTMAS CLIENTS (ORIGINAL).xlsx")
OUT = os.path.join(HERE, "2025 CHRISTMAS CLIENTS (2026 zoning + assignments).xlsx")

clients = json.load(open(os.path.join(CACHE, "clients.json")))["clients"]
sched = json.load(open(os.path.join(CACHE, "schedule.json")))


def merge_crew(prev, new):
    """Joint stops appear on two crews' cards -> merged label."""
    names = set()
    for c in (prev, new):
        names.update(x.strip() for x in c.replace(" (joint)", "").split(" + "))
    return " + ".join(sorted(names)) + " (joint)"

# row -> uniform (area, zone)
row_zone = {c["row"]: (c["area"], c["zone"]) for c in clients}
# row -> (2026 date, crew, order-in-day)
row_assign = {}
for d in sched["days"]:
    for i, s in enumerate(d["stops"], 1):
        if s["row"] in row_assign:
            dt, cr, o = row_assign[s["row"]]
            row_assign[s["row"]] = (dt, merge_crew(cr, d["crew"]), o)
        else:
            row_assign[s["row"]] = (d["date"], d["crew"], i)
# no-address clients + dropped (no 2026 install)
noaddr_rows = {c["row"] for c in clients if c["no_address"]}
dropped_names = {x["name"] for x in sched.get("dropped", [])}
dropped_rows = {c["row"] for c in clients if c["name"] in dropped_names}

shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)      # keep formulas / all sheets / formatting
ws = wb["2025 Christmas"]

AREA_COL, ZONE_COL = 6, 7
BOX_COL = 13
NEW = {"2026 Install Date": 44, "Crew": 45, "Order": 46,
       "Storage Boxes (verified)": 47}

hdr_fill = PatternFill("solid", fgColor="0B5D3B")
hdr_font = Font(bold=True, color="FFFFFF")
for title, col in NEW.items():
    cell = ws.cell(1, col, title)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.column_dimensions["AR"].width = 16   # 2026 Install Date
ws.column_dimensions["AS"].width = 22   # Crew
ws.column_dimensions["AT"].width = 7    # Order
ws.column_dimensions["AU"].width = 14   # Storage Boxes (verified)

row_boxes = {c["row"]: c for c in clients if c.get("box_verified")}
MISMATCH = PatternFill("solid", fgColor="FDEBD0")
updated_zone = updated_assign = updated_boxes = 0
for r in range(2, ws.max_row + 1):
    if r in row_zone:
        area, zone = row_zone[r]
        if area != "UNKNOWN":          # leave the 2 no-address rows' cells alone
            ws.cell(r, AREA_COL, area)
            ws.cell(r, ZONE_COL, zone)
            updated_zone += 1
    if r in row_assign:
        date_s, crew, order = row_assign[r]
        y, m, day = (int(x) for x in date_s.split("-"))
        c = ws.cell(r, NEW["2026 Install Date"], datetime.date(y, m, day))
        c.number_format = "ddd mm/dd/yyyy"
        ws.cell(r, NEW["Crew"], crew)
        ws.cell(r, NEW["Order"], order)
        updated_assign += 1
    if r not in row_assign:
        if r in dropped_rows:
            ws.cell(r, NEW["2026 Install Date"], "NO INSTALL 2026")
        elif r in noaddr_rows:
            ws.cell(r, NEW["2026 Install Date"], "NEEDS ADDRESS")
    if r in row_boxes:
        c = row_boxes[r]
        cell = ws.cell(r, NEW["Storage Boxes (verified)"], c["box_count"])
        sheet_v = c.get("box_count_sheet")
        try:
            sheet_n = int(float(sheet_v)) if sheet_v not in (None, "", "?") else None
        except (ValueError, TypeError):
            sheet_n = None
        if sheet_n != c["box_count"]:
            cell.fill = MISMATCH          # differs from (or missing in) BOX COUNT col
        updated_boxes += 1

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  standardized zoning on {updated_zone} client rows")
print(f"  wrote 2026 date/crew/order on {updated_assign} client rows")
print(f"  flagged NEEDS ADDRESS on {len(noaddr_rows)} rows")
print(f"  storage-verified box counts on {updated_boxes} rows")
