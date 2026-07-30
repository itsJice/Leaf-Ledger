#!/usr/bin/env python3
"""
TBDG 2026 — Team Review spreadsheet for the planning meeting.

One row per client with everything the team needs side-by-side:
history (2024/2025 dates, 2025 real hours, 2025 crew), logistics (storage,
boxes, phone, email), and the proposed 2026 assignment (date, crew, route
order, people needed). Sorted by 2026 date -> crew -> route order so it
reads like the run of show. Second tab summarizes each crew-day.
"""
import datetime
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
OUT = os.path.join(HERE, "2026 Team Review.xlsx")

HDR_FILL = PatternFill("solid", fgColor="0B5D3B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D5D8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CREW_FILL = {
    "Alberto": "C0392B",            # red
    "Lesly": "1F618D",              # blue
    "Niurka": "1E8449",             # green
    "Alberto + Lesly (stacked)": "7D3C98",
    "Alberto + Lesly + Niurka (stacked)": "B9770E",
}
DATE_BANDS = ["FFFFFF", "F2F4F4"]  # alternate per date
OVER_FILL = PatternFill("solid", fgColor="F5B7B1")
WEEKEND_FILL = PatternFill("solid", fgColor="FDEBD0")


def merge_crew(prev, new):
    """Joint stops appear on two crews' cards -> merged label."""
    names = set()
    for c in (prev, new):
        names.update(x.strip() for x in c.replace(" (joint)", "").split(" + "))
    return " + ".join(sorted(names)) + " (joint)"


def load():
    sched = json.load(open(os.path.join(CACHE, "schedule.json")))
    return sched


def fmt_date(iso):
    if not iso or "-" not in str(iso)[:8]:
        return iso or ""
    try:
        y, m, d = (int(x) for x in str(iso)[:10].split("-"))
        return datetime.date(y, m, d)
    except ValueError:
        return iso


def main():
    sched = load()
    days = sched["days"]
    clients = {c["row"]: c for c in sched["all_clients"]}

    # row -> assignment
    assign = {}
    for d in days:
        for i, s in enumerate(d["stops"], 1):
            if s["row"] in assign:
                d0, i0 = assign[s["row"]]
                d0 = dict(d0)
                d0["crew"] = merge_crew(d0["crew"], d["crew"])
                assign[s["row"]] = (d0, i0)
            else:
                assign[s["row"]] = (d, i)

    wb = openpyxl.Workbook()

    # ---------------- Client Review (main) ----------------
    ws = wb.active
    ws.title = "Client Review"
    hdr = ["2026 Install", "Day", "Crew", "Stop #", "Client",
           "Address", "City", "ZIP", "Phone", "Email",
           "TBDG Storage", "Boxes", "2024 Install", "2025 Install",
           "2025 Real Hrs", "2025 Crew", "2026 Plan Hrs", "Hrs Basis",
           "People Needed", "Zone", "Notes / Flags"]
    ws.append(hdr)
    for c in ws[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ordered client rows: assigned first (by date/crew/order), then unassigned
    ordered = sorted(assign.items(), key=lambda kv: (kv[1][0]["date"], kv[1][0]["crew"], kv[1][1]))
    band_idx, last_date = 0, None
    for row_id, (d, order) in ordered:
        c = clients[row_id]
        if d["date"] != last_date:
            band_idx = 1 - band_idx if last_date else 0
            last_date = d["date"]
        # actual 2025 crew size beats the sheet's staffing estimate
        people = c.get("crew_size_2025") or c.get("people_needed")
        people_s = str(people) if people else "5 (lead+4)"
        note = d["note"] if order == 1 else ""
        flags = "; ".join(d["flags"]) if order == 1 and d["flags"] else ""
        ws.append([
            fmt_date(d["date"]), d["dow"], d["crew"], order, c["name"],
            c["street"], c["city"], c["zip"], c.get("phone", ""), c.get("email", ""),
            c.get("storage", ""), c.get("box_count") or "",
            fmt_date(c.get("date_2024", "")), fmt_date(c.get("prior_install_date", "")),
            c.get("real_hours") if c.get("real_hours") else "",
            c.get("crew_2025", ""), c.get("cal_hours"), c.get("hours_basis", ""),
            people_s, c["zone"], (note + (" | " if note and flags else "") + flags)])
        r = ws.max_row
        ws.cell(r, 1).number_format = "ddd mm/dd"
        ws.cell(r, 13).number_format = "mm/dd/yy"
        ws.cell(r, 14).number_format = "mm/dd/yy"
        if isinstance(ws.cell(r, 15).value, float):
            ws.cell(r, 15).number_format = "0.0"
        band = PatternFill("solid", fgColor=DATE_BANDS[band_idx])
        for cc in range(1, len(hdr) + 1):
            cell = ws.cell(r, cc)
            cell.border = BORDER
            if d["dow"] in ("Sat", "Sun"):
                cell.fill = WEEKEND_FILL
            else:
                cell.fill = band
        fill = CREW_FILL.get(d["crew"])
        if fill:
            ws.cell(r, 3).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, 3).font = Font(color="FFFFFF", bold=True)

    # unassigned at bottom
    for c in sched["all_clients"]:
        if c["row"] in assign:
            continue
        status = ("NEEDS ADDRESS" if c["no_address"] else
                  "NO INSTALL 2026" if any(x["name"] == c["name"] for x in sched.get("dropped", []))
                  else "UNPLACED")
        ws.append(["", "", status, "", c["name"], c["street"], c["city"], c["zip"],
                   c.get("phone", ""), c.get("email", ""), c.get("storage", ""),
                   c.get("box_count") or "", fmt_date(c.get("date_2024", "")),
                   fmt_date(c.get("prior_install_date", "")),
                   c.get("real_hours") or "", c.get("crew_2025", ""),
                   c.get("cal_hours"), c.get("hours_basis", ""), "", c["zone"], status])
        for cc in range(1, len(hdr) + 1):
            ws.cell(ws.max_row, cc).fill = OVER_FILL
            ws.cell(ws.max_row, cc).border = BORDER

    ws.freeze_panes = "F2"
    widths = [12, 6, 22, 6, 34, 28, 14, 8, 15, 26, 12, 7, 11, 11, 10, 14, 10, 11, 12, 24, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{ws.max_row}"

    # ---------------- Crew Days (approval sheet) ----------------
    ws2 = wb.create_sheet("Crew Days")
    hdr2 = ["Date", "Day", "Crew", "# Stops", "Stops (in route order)",
            "Install Hrs", "Drive Hrs", "Total Hrs", "Window OK?",
            "2025 Real Hrs (these stops)", "Storage Stops", "Total Boxes",
            "Max People Needed", "Approved?", "Notes"]
    ws2.append(hdr2)
    for c in ws2[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for d in days:
        stops = d["stops"]
        cl = [clients[s["row"]] for s in stops]
        real25 = sum(x.get("real_hours") or 0 for x in cl)
        storage_n = sum(1 for x in cl if str(x.get("storage", "")).lower().startswith("y"))

        def num(v):
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return 0
        boxes = sum(num(x.get("box_count")) for x in cl)
        peoples = [x.get("crew_size_2025") or x.get("people_needed") or 5 for x in cl]
        ws2.append([fmt_date(d["date"]), d["dow"], d["crew"], len(stops),
                    " → ".join(s["name"] for s in stops),
                    d["install_h"], round(d["drive_min"] / 60, 2),
                    round(d["total_min"] / 60, 2),
                    "YES" if d["total_min"] <= d.get("window_min", 600) else "OVER",
                    round(real25, 1) if real25 else "",
                    storage_n, boxes or "", max(peoples), "", d["note"]])
        r = ws2.max_row
        ws2.cell(r, 1).number_format = "ddd mm/dd"
        fill = CREW_FILL.get(d["crew"])
        if fill:
            ws2.cell(r, 3).fill = PatternFill("solid", fgColor=fill)
            ws2.cell(r, 3).font = Font(color="FFFFFF", bold=True)
        for cc in range(1, len(hdr2) + 1):
            ws2.cell(r, cc).border = BORDER
            ws2.cell(r, cc).alignment = Alignment(vertical="top", wrap_text=(cc == 5))
        if d["total_min"] > d.get("window_min", 600):
            ws2.cell(r, 9).fill = OVER_FILL
    ws2.freeze_panes = "A2"
    for i, w in enumerate([12, 6, 24, 7, 70, 9, 9, 9, 10, 12, 9, 9, 10, 10, 40], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
