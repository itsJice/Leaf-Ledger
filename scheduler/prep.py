#!/usr/bin/env python3
"""
TBDG 2026 Christmas install schedule -- PREP stage.

Parses the raw 2025 spreadsheet, excludes billing line-items, flags no-address
clients, recomputes Area/Zone from ZIP (the file's Area/Zone columns are
unreliable), calibrates 2026 install hours, classifies Business vs Residence,
categorizes the special clients that drive the hard scheduling rules, then
geocodes every address (Nominatim, cached) and builds the OSRM drive-time
matrix (cached).

Outputs (all in cache/):
  clients.json     -- fully enriched client records
  geocoded.csv     -- address -> lat/lon cache (Nominatim)
  matrix.json      -- N+1 x N+1 driving-duration matrix in seconds (OSRM)
"""
import csv
import json
import os
import re
import statistics
import time

import openpyxl
import requests

import client_config_loader

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
os.makedirs(CACHE, exist_ok=True)
SRC = os.path.join(HERE, "CHRISTMAS CLIENTS - Storage - Delivery - Install +Takedown.xlsx")
SHEET = "2026 Christmas"
DEPOT = "2860 Antoine Dr, Houston, TX 77092"
UA = {"User-Agent": "TBDG-christmas-scheduler/1.0 (justice@wenzdays.com)"}

EXCLUDE = {
    "GENERAL INSTALL LABOR", "PICKUP & DELIVERY - INSTALL",
    "PICKUP & DELIVERY - TAKEDOWN", "SPECIALTY INSTALL LABOR",
    "STORAGE FEE PER BOX", "CREW LEAD", "DESIGNER ART DIRECTOR LEAD",
}

# Every per-client override below (ZIP, no-address, storage box counts,
# storage status, manual coordinates, the single-crew-priority category)
# lives in client_config.json, never in this file (user, 2026-08-02: no
# client names in the codebase). See client_config_loader.py.
CLIENT_CONFIG = client_config_loader.load()

# ---------------------------------------------------------------------------
# ZIP -> (Area, Zone).  Reconstructed from the brief's taxonomy + Houston/DFW
# geography (the shipped zone_map.txt was not present).  ZIP wins over the
# file's old label.
# ---------------------------------------------------------------------------
ZIP_ZONE = {
    # ---- Houston: Central / Inner Loop ----
    "77002": ("Central", "Downtown/Midtown"),
    "77006": ("Central", "Downtown/Midtown"),
    "77007": ("Central", "Heights/Near Northside"),
    "77008": ("Central", "Heights/Near Northside"),
    "77018": ("Central", "Heights/Near Northside"),
    "77030": ("Central", "Medical Center"),
    "77054": ("Central", "Medical Center"),
    "77027": ("Central", "River Oaks/Upper Kirby"),
    "77005": ("Central", "West University/Rice"),
    "77025": ("Central", "West University/Rice"),
    # ---- Houston: West ----
    "77056": ("West", "Uptown/Galleria"),
    "77057": ("West", "Uptown/Galleria"),
    "77063": ("West", "Uptown/Galleria"),
    "77024": ("West", "Memorial/Spring Branch"),
    "77055": ("West", "Memorial/Spring Branch"),
    "77079": ("West", "Energy Corridor/Westchase"),
    "77042": ("West", "Energy Corridor/Westchase"),
    "77082": ("West", "Energy Corridor/Westchase"),
    "77441": ("West", "Katy/Fulshear"),
    "77450": ("West", "Katy/Fulshear"),
    "77449": ("West", "Katy/Fulshear"),
    "77493": ("West", "Katy/Fulshear"),
    "77474": ("West", "Katy/Fulshear"),
    "77406": ("West", "Katy/Fulshear"),
    "77423": ("West", "Katy/Fulshear"),
    # ---- Houston: Northwest ----
    "77095": ("Northwest", "Cypress/Copperfield"),
    "77433": ("Northwest", "Cypress/Copperfield"),
    # ---- Houston: North ----
    "77380": ("North", "The Woodlands"),
    "77381": ("North", "The Woodlands"),
    "77382": ("North", "The Woodlands"),
    "77385": ("North", "The Woodlands"),
    "77389": ("North", "The Woodlands"),
    "77355": ("North", "Magnolia/Tomball"),
    "77373": ("North", "Spring/Klein/North Belt"),
    "77379": ("North", "Spring/Klein/North Belt"),
    "77069": ("North", "Spring/Klein/North Belt"),
    "77073": ("North", "Spring/Klein/North Belt"),
    "77356": ("North", "Montgomery/Lake Conroe"),
    # ---- Houston: Northeast ----
    "77345": ("Northeast", "Kingwood/Humble/Porter"),
    "77346": ("Northeast", "Kingwood/Humble/Porter"),
    "77365": ("Northeast", "Kingwood/Humble/Porter"),
    # ---- Houston: Southwest ----
    "77479": ("Southwest", "Sugar Land/Stafford"),
    "77459": ("Southwest", "Missouri City/Sienna"),
    "77489": ("Southwest", "Missouri City/Sienna"),
    # ---- Houston: South ----
    "77581": ("South", "Pearland"),
    "77584": ("South", "Pearland"),
    # ---- Houston: Southeast ----
    "77058": ("Southeast", "Clear Lake/Bay Area"),
    "77059": ("Southeast", "Clear Lake/Bay Area"),
    "77573": ("Southeast", "Clear Lake/Bay Area"),
    "77505": ("Southeast", "Pasadena/Deer Park/Ship Channel"),
    "77536": ("Southeast", "Pasadena/Deer Park/Ship Channel"),
    "77029": ("Southeast", "Pasadena/Deer Park/Ship Channel"),
    "77521": ("Southeast", "Baytown"),
    # ---- Houston: Far West ----
    "77418": ("Far West", "Bellville/Sealy"),
    "77426": ("Far West", "Brenham/Chappell Hill"),
    "77833": ("Far West", "Brenham/Chappell Hill"),
    # ---- Dallas-Fort Worth (M Crowd) ----
    "75201": ("Dallas - Central", "Uptown/Park Cities"),
    "75204": ("Dallas - Central", "Uptown/Park Cities"),
    "75205": ("Dallas - Central", "Uptown/Park Cities"),
    "75207": ("Dallas - Central", "Uptown/Park Cities"),
    "75230": ("Dallas - North", "Preston Hollow/NorthPark"),
    "75240": ("Dallas - North", "Preston Hollow/NorthPark"),
    "75214": ("Dallas - East", "Lakewood/Lake Highlands"),
    "75231": ("Dallas - East", "Lakewood/Lake Highlands"),
    "75024": ("Dallas - North", "Plano/Frisco"),
    "75093": ("Dallas - North", "Plano/Frisco"),
    "75034": ("Dallas - North", "Plano/Frisco"),
    "75013": ("Dallas - North", "Allen/McKinney"),
    "75070": ("Dallas - North", "Allen/McKinney"),
    "75006": ("Dallas - North", "Addison/Carrollton/Richardson"),
    "75080": ("Dallas - North", "Addison/Carrollton/Richardson"),
    "75287": ("Dallas - North", "Addison/Carrollton/Richardson"),
    "75056": ("Dallas - North", "The Colony/Grandscape"),
    "75063": ("Dallas - West", "Las Colinas/Irving"),
    "75028": ("Dallas - West", "Flower Mound/Southlake/Grapevine"),
    "76092": ("Dallas - West", "Flower Mound/Southlake/Grapevine"),
    "76102": ("Dallas - West", "Fort Worth"),
    "76107": ("Dallas - West", "Fort Worth"),
    "76177": ("Dallas - West", "Fort Worth"),
    "75087": ("Dallas - East", "Rockwall/East"),
}

# Per-client ZIP overrides keyed by exact client name (brief's corrections
# and the messy rows).  ZIP then flows through ZIP_ZONE.
ZIP_OVERRIDE = CLIENT_CONFIG["zip_override"]

# Clients with no usable address at all -> flag, do not route/map.
NO_ADDRESS = set(CLIENT_CONFIG["no_address"])

# VERIFIED box counts from the storage-room binder photos (2026-07-30).
# The binder reflects what is physically racked and OVERRIDES the sheet's
# BOX COUNT column (which was wildly off for the clubs).
STORAGE_BOX_COUNTS = CLIENT_CONFIG["storage_box_counts"]

# Manual storage-status corrections (client, 2026-07-31): overrides the
# sheet's "TBDG STORAGE YES/NO" column when a client's situation changed.
STORAGE_OVERRIDE = CLIENT_CONFIG["storage_override"]

# Manual coordinate fixes for rows whose street cell mis-geocodes (verified
# against Nominatim by ZIP/known location). One entry (a rural road with no
# exact match in Nominatim or the Census geocoder) is a ZIP-centroid
# fallback rather than house-precise -- see client_config.json's comment
# convention / RULES.md for the full story.
MANUAL_COORDS = CLIENT_CONFIG["manual_coords"]

BUSINESS_KEYWORDS = [
    "m crowd", "bank", "club", "cc", "hotel", "suites", "inn", "church",
    "daycare", "academy", "office", "salon", "market", "grill", "cafe",
    "center", "rotary", "llc", "inc", "company", "school", "corporate",
    "restaurant", "cocina", "tavern", "group", "medical", "clinic",
    "mercury", "district",
]


def clean_zip(v):
    if v is None:
        return ""
    try:
        return str(int(float(v))).zfill(5)
    except (ValueError, TypeError):
        m = re.search(r"\d{5}", str(v))
        return m.group(0) if m else ""


def clean_street(addr, city, zc):
    """Isolate the street from a messy ADDRESS cell that may have the full
    'street, city, ST zip' jammed in.  Strip a trailing city/ST/zip tail.

    Bug (found 2026-08-01, Keffer/Pam): the old generic tail regex
    (`[A-Za-z .]+,?\\s*TX\\s*\\d{5}`) has no required comma before the
    letters, so on an ADDRESS cell with no comma between street and city
    ("5231 Wincroft Houston, TX 77069") it greedily eats "Wincroft Houston"
    too, leaving just "5231". Strip using the KNOWN city (from the sheet's
    own CITY column) FIRST, comma optional -- that's anchored to a real
    place name instead of "however many words happen to precede TX zip" --
    and only fall back to the generic pattern (now comma-REQUIRED, so it
    can no longer eat an unknown word) if the city isn't in the string.
    """
    if not addr:
        return ""
    s = str(addr).strip()
    if city:
        c = re.escape(str(city).strip())
        s = re.sub(r",?\s*" + c + r"\s*,?\s*TX\s*\d{5}.*$", "", s, flags=re.I)
        s = re.sub(r",?\s*" + c + r"\s*,\s*TX.*$", "", s, flags=re.I)
        s = re.sub(r",\s*" + c + r"\s*$", "", s, flags=re.I)
    # fallback: cut a trailing ", City, TX 77xxx" tail -- comma required
    # before the city words, so this can't swallow an unrecognized street
    # name the way the old unanchored version did.
    s = re.sub(r",\s*[A-Za-z .]+,?\s*TX\s*\d{5}.*$", "", s, flags=re.I)
    s = re.sub(r",?\s*TX\s*\d{5}.*$", "", s, flags=re.I)
    s = re.sub(r",?\s*\d{5}(-\d{4})?\s*$", "", s)
    if city:
        s = re.sub(r",\s*" + re.escape(str(city).strip()) + r"\s*$", "", s, flags=re.I)
    return s.strip().rstrip(",").strip()


def classify(name):
    n = name.lower()
    if re.match(r"^[A-Za-z'`.\- ]+,\s*[A-Za-z]", name) and "crowd" not in n:
        # "Lastname, Firstname" -> residence (unless a keyword overrides)
        if not any(k in n for k in BUSINESS_KEYWORDS if k not in ("cc", "club")):
            return "Residence"
    if "residence" in n:
        return "Residence"
    if any(re.search(r"\b" + re.escape(k) + r"\b", n) for k in BUSINESS_KEYWORDS):
        return "Business"
    if re.match(r"^[A-Za-z'`.\- ]+,\s*[A-Za-z]", name):
        return "Residence"
    return "Business"


def categorize(name):
    n = name.lower()
    if n.startswith("m crowd"):
        return "M Crowd"
    if "capital bank" in n:
        return "Capital Bank"
    if any(k in n for k in ["carlton woods", "country club"]) or \
       re.search(r"\bcc\b", n) or re.search(r"\bclub\b", n):
        return "Country Club"
    if "rotary house" in n:
        return "Rotary House"
    scp = CLIENT_CONFIG["single_crew_priority"]
    if name.strip() == scp["client_name"]:
        return scp["category"]
    return "Standard"


def parse():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    # Header-name column lookup (NOT fixed indices) -- the sheet's column
    # order has already shifted once between spreadsheet revisions, and a
    # hardcoded index list silently misreads data on the next reshuffle.
    col = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value:
            key = re.sub(r"\s+", " ", str(cell.value).strip())
            col.setdefault(key, cell.column)

    def h(r, header, *fallbacks):
        for name in (header,) + fallbacks:
            c = col.get(name)
            if c:
                return ws.cell(r, c).value
        return None

    clients = []
    for r in range(2, ws.max_row + 1):
        name = h(r, "TBDG CLIENT")
        if name is None or not str(name).strip():
            continue
        name = re.sub(r"\s+", " ", str(name).strip().splitlines()[0]).strip()
        if name.upper() in EXCLUDE:
            continue

        city = (str(h(r, "CITY")).strip() if h(r, "CITY") else "")
        st = (str(h(r, "ST")).strip() if h(r, "ST") else "TX")
        zc = ZIP_OVERRIDE.get(name, clean_zip(h(r, "ZIP")))
        street = clean_street(h(r, "ADDRESS"), city, zc)

        # Messy city cell that actually holds the full address (e.g. Musser)
        if not street and "," in city:
            street = clean_street(city, "", zc)
            city = ""

        est = h(r, "ESTIMATED TOTAL HOURS FOR INSTALL")
        real = h(r, "2025 Real Hours For Install")
        try:
            est = float(est) if est is not None else None
        except (ValueError, TypeError):
            est = None
        try:
            real = float(real) if real is not None else None
        except (ValueError, TypeError):
            real = None

        def as_date(v):
            return (v.date().isoformat() if hasattr(v, "date")
                    else (str(v).strip() if v else ""))

        def as_int(v):
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return None

        def as_money(v):
            # Several fee columns are broken formulas ("#VALUE!") on most
            # rows -- a formula error reads back as that literal string, not
            # a raise, so isinstance(v, str) alone isn't enough to catch it.
            try:
                return round(float(v), 2)
            except (ValueError, TypeError):
                return None

        prior_date = as_date(h(r, "Install Date 2025"))
        date_2024 = as_date(h(r, "2024 INSTALL DATE"))

        # 2025 crew name — first line; size from "(6)" or roster lines
        crew_raw_v = h(r, "2025 Crew Name", "Crew Name")
        crew_raw = str(crew_raw_v).strip() if crew_raw_v else ""
        crew_2025 = crew_raw.splitlines()[0].strip() if crew_raw else ""
        m_sz = re.search(r"\((\d+)\)", crew_2025)
        if m_sz:
            crew_size_2025 = int(m_sz.group(1))
        elif "\n" in crew_raw:
            crew_size_2025 = len([l for l in crew_raw.splitlines() if l.strip()])
        else:
            crew_size_2025 = None

        # staffing asks
        staff = [x for x in (as_int(h(r, "# CREW LEADS NEEDED")),
                             as_int(h(r, "# SPECIALTY LABOR (SCAFFOLDING EXTRA TALL LADDER)")),
                             as_int(h(r, "# DESIGNER / ART DIRECTOR")),
                             as_int(h(r, "# GENERAL INSTALLERS NEEDED"))) if x]
        people_needed = sum(staff) if staff else None

        box = h(r, "BOX COUNT")
        area, zone = ZIP_ZONE.get(zc, ("UNKNOWN", "UNKNOWN"))

        # Billing-adjacent columns for the review tool's billing export.
        # INSTALL/TAKEDOWN LABOR FEE and the TOTAL columns are broken
        # formulas ("#VALUE!") on effectively every row as of 2026-08-10 --
        # as_money() returns None for those rather than a garbage string.
        # STORAGE FEE is a real, working column. Production Notes carries
        # repair/relight/replacement flags ("needs new tree", "had to
        # relight") -- distinct from Confirmation Notes, which is
        # scheduling/logistics chatter, not billing- or repair-relevant.
        storage_fee = as_money(h(r, "STORAGE FEE (BASED ON # OF BOXES)"))
        install_fee = as_money(h(r, "INSTALL LABOR FEE"))
        takedown_fee = as_money(h(r, "TAKEDOWN LABOR FEE"))
        invoice_2025_total = as_money(h(r, "2025 Invoice Total Actual Created"))
        production_notes_v = h(r, "2025 Production Notes", "Production Notes")
        production_notes = (str(production_notes_v).strip()
                            if production_notes_v else "")

        box_sheet = box
        if name in STORAGE_BOX_COUNTS:
            box = STORAGE_BOX_COUNTS[name]

        # 2026 Install Date column: a real date = client already deposited &
        # reserved that date (hard pin); free text may flag "same day as X"
        # groupings, "no install" (drop from 2026 entirely), or other notes.
        raw_2026 = h(r, "2026 Install Date")
        install_2026_confirmed = (raw_2026.date().isoformat()
                                  if hasattr(raw_2026, "date") else "")
        install_2026_note = ("" if raw_2026 is None or hasattr(raw_2026, "date")
                             else str(raw_2026).strip())
        install_2026_no_install = bool(re.search(
            r"no\s*(?:2026\s*)?install\b", install_2026_note, re.I))

        rec = {
            "row": r,
            "name": name,
            "street": street,
            "city": city,
            "st": st or "TX",
            "zip": zc,
            "area": area,
            "zone": zone,
            "box_count": box,
            "box_count_sheet": box_sheet,
            "box_verified": name in STORAGE_BOX_COUNTS,
            "phone": str(h(r, "PHONE")).strip() if h(r, "PHONE") else "",
            "email": str(h(r, "EMAIL")).strip() if h(r, "EMAIL") else "",
            "storage": STORAGE_OVERRIDE.get(
                name, str(h(r, "TBDG STORAGE YES/NO")).strip() if h(r, "TBDG STORAGE YES/NO") else ""),
            "date_2024": date_2024,
            "crew_2025": crew_2025,
            "crew_size_2025": crew_size_2025,
            "people_needed": people_needed,
            "est_hours": est,
            "real_hours": real,
            "prior_install_date": prior_date,
            "business": classify(name),
            "category": categorize(name),
            "no_address": name in NO_ADDRESS or (not street and not zc),
            "install_2026_confirmed": install_2026_confirmed,
            "install_2026_note": install_2026_note,
            "install_2026_no_install": install_2026_no_install,
            "storage_fee": storage_fee,
            "install_fee_2026": install_fee,
            "takedown_fee_2026": takedown_fee,
            "invoice_2025_total": invoice_2025_total,
            "production_notes": production_notes,
        }
        clients.append(rec)
    return clients


def calibrate_hours(clients):
    """1) real hours if present; 2) est * 0.81; 3) zone median real; else 2.8."""
    zone_reals = {}
    for c in clients:
        if c["real_hours"]:
            zone_reals.setdefault(c["zone"], []).append(c["real_hours"])
    all_reals = [c["real_hours"] for c in clients if c["real_hours"]]
    overall_median = statistics.median(all_reals) if all_reals else 2.8

    for c in clients:
        if c["real_hours"]:
            c["cal_hours"] = round(c["real_hours"], 2)
            c["hours_basis"] = "2025 real"
        elif c["est_hours"]:
            c["cal_hours"] = round(c["est_hours"] * 0.81, 2)
            c["hours_basis"] = "est x0.81"
        else:
            zr = zone_reals.get(c["zone"])
            med = statistics.median(zr) if zr else overall_median
            c["cal_hours"] = round(med, 2)
            c["hours_basis"] = "zone median"
    return overall_median


def census_geocode(cache, street, city, st, zc):
    """US Census one-line geocoder — free, rooftop-accurate for US streets.
    Used when Nominatim can't resolve the house number."""
    addr = ", ".join(p for p in [street, city, f"{st} {zc}".strip()] if p)
    key = "CENSUS|" + addr
    if key in cache:
        return cache[key]
    lat = lon = disp = ""
    try:
        r = requests.get(
            "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
            params={"address": addr, "benchmark": "Public_AR_Current",
                    "format": "json"}, timeout=25)
        matches = r.json().get("result", {}).get("addressMatches", [])
        if matches:
            co = matches[0]["coordinates"]
            lat, lon = str(co["y"]), str(co["x"])
            disp = "census: " + matches[0].get("matchedAddress", "")
    except Exception as e:
        print("  census retry skipped:", e)
    time.sleep(0.5)
    cache[key] = (lat, lon, disp)
    return cache[key]


def geocode_all(clients):
    cache_path = os.path.join(CACHE, "geocoded.csv")
    cache = {}
    if os.path.exists(cache_path):
        with open(cache_path, newline="") as f:
            for row in csv.DictReader(f):
                cache[row["query"]] = (row["lat"], row["lon"], row["display"])

    def geocode(query, structured=None):
        # Build a UNIQUE cache key: structured queries would otherwise all
        # share the empty-string key and clobber each other.
        key = query if not structured else "STRUCT|" + "|".join(
            f"{k}={v}" for k, v in sorted(structured.items()))
        if key in cache:
            return cache[key]
        params = {"format": "json", "countrycodes": "us", "limit": 1}
        if structured:
            params.update(structured)
        else:
            params["q"] = query
        lat = lon = disp = ""
        for attempt in range(3):
            try:
                r = requests.get("https://nominatim.openstreetmap.org/search",
                                 params=params, headers=UA, timeout=25)
                if r.status_code == 200 and r.json():
                    j = r.json()[0]
                    lat, lon, disp = j["lat"], j["lon"], j.get("display_name", "")
                break
            except Exception as e:
                print("  geocode retry:", e)
                time.sleep(2)
        time.sleep(1.1)  # respect 1 req/sec
        cache[key] = (lat, lon, disp)
        return cache[key]

    # depot first
    dlat, dlon, ddisp = geocode(DEPOT)
    depot = {"name": "DEPOT (2860 Antoine Dr)", "lat": float(dlat), "lon": float(dlon)}

    for c in clients:
        if c["no_address"]:
            c["lat"], c["lon"], c["geo_source"] = None, None, "NO ADDRESS"
            continue
        if c["name"] in MANUAL_COORDS:
            c["lat"], c["lon"] = MANUAL_COORDS[c["name"]]
            c["geo_display"] = "manual override"
            c["geo_source"] = "manual"
            continue
        # Fallback chain (precise -> coarse). The structured city+postalcode
        # query is LAST because Nominatim ignores the postalcode when a city
        # is given and returns the city centroid (collapsing many clients);
        # the free-form "{zip}, TX" query correctly returns a per-ZIP centroid.
        st = c["st"] or "TX"
        source = "street"
        lat = lon = disp = ""
        if c["street"] and c["zip"]:
            lat, lon, disp = geocode(f"{c['street']}, {st} {c['zip']}")
        if not lat and c["street"] and c["city"]:  # street + city + zip
            parts = [p for p in [c["street"], c["city"], st, c["zip"]] if p]
            lat, lon, disp = geocode(", ".join(parts))
        if not lat and c["street"] and c["city"]:  # street + city, no zip
            lat, lon, disp = geocode(f"{c['street']}, {c['city']}, {st}")
        if not lat and c["street"]:  # US Census geocoder (rooftop-accurate)
            lat, lon, disp = census_geocode(cache, c["street"], c["city"], st, c["zip"])
            source = "census"
        if not lat and c["zip"]:  # per-ZIP centroid (free-form)
            lat, lon, disp = geocode(f"{c['zip']}, {st}")
            source = "zip"
        if not lat and c["city"]:  # structured city (coarsest)
            lat, lon, disp = geocode("", {"city": c["city"], "state": "TX"})
            source = "city"
        c["lat"] = float(lat) if lat else None
        c["lon"] = float(lon) if lon else None
        c["geo_display"] = disp
        c["geo_source"] = source if lat else "FAILED"

    with open(cache_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["query", "lat", "lon", "display"])
        for q, (la, lo, d) in cache.items():
            w.writerow([q, la, lo, d])
    return depot


def build_matrix(depot, clients):
    """OSRM table service; node 0 = depot, then routable clients in order.

    The public server caps coordinates per request (~100), so we tile the
    matrix in source x destination blocks and only send the union of each
    block pair's coordinates.
    """
    nodes = [depot] + [c for c in clients if c.get("lat") is not None]
    coords = [(n["lon"], n["lat"]) for n in nodes]
    N = len(coords)

    # Skip the (slow) rebuild when nothing moved since the cached matrix.
    mat_path = os.path.join(CACHE, "matrix.json")
    if os.path.exists(mat_path):
        try:
            old = json.load(open(mat_path))
            oc = [tuple(x) for x in old.get("coords", [])]
            if len(oc) == N and all(
                    abs(a[0] - b[0]) < 1e-6 and abs(a[1] - b[1]) < 1e-6
                    for a, b in zip(oc, coords)):
                print("  matrix unchanged — reusing cache")
                return old
        except Exception:
            pass

    mat = [[None] * N for _ in range(N)]

    BLK = 45  # union of two blocks <= 90 coords, safely under the cap
    blocks = [list(range(i, min(i + BLK, N))) for i in range(0, N, BLK)]

    def fmt(idx):
        return ";".join(f"{coords[i][0]},{coords[i][1]}" for i in idx)

    for sb in blocks:
        for db in blocks:
            combined = sb + [i for i in db if i not in sb]
            pos = {node: k for k, node in enumerate(combined)}
            src_pos = ";".join(str(pos[i]) for i in sb)
            dst_pos = ";".join(str(pos[i]) for i in db)
            url = f"https://router.project-osrm.org/table/v1/driving/{fmt(combined)}"
            params = {"annotations": "duration",
                      "sources": src_pos, "destinations": dst_pos}
            ok = False
            for attempt in range(5):
                try:
                    r = requests.get(url, params=params, timeout=60)
                    j = r.json()
                    if j.get("code") == "Ok":
                        for ri, s in enumerate(sb):
                            for ci, d in enumerate(db):
                                mat[s][d] = j["durations"][ri][ci]
                        ok = True
                        break
                    print("  OSRM code:", j.get("code"), j.get("message"))
                except Exception as e:
                    print("  OSRM retry:", e)
                time.sleep(3)
            if not ok:
                print(f"  !! block ({sb[0]}..{sb[-1]})x({db[0]}..{db[-1]}) FAILED")
            time.sleep(1)
        print(f"  matrix source block {sb[0]}..{sb[-1]} done")

    out = {
        "node_names": [n["name"] for n in nodes],
        "node_ids": [None] + [c["row"] for c in clients if c.get("lat") is not None],
        "coords": coords,
        "durations": mat,
    }
    with open(os.path.join(CACHE, "matrix.json"), "w") as f:
        json.dump(out, f)
    return out


def main():
    print("Parsing spreadsheet...")
    clients = parse()
    overall_median = calibrate_hours(clients)
    print(f"  {len(clients)} clients; overall median real hours = {overall_median:.2f}")

    from collections import Counter
    print("  categories:", dict(Counter(c["category"] for c in clients)))
    print("  business/res:", dict(Counter(c["business"] for c in clients)))
    unk = [c["name"] for c in clients if c["area"] == "UNKNOWN" and not c["no_address"]]
    if unk:
        print("  !! UNKNOWN zone:", unk)
    noaddr = [c["name"] for c in clients if c["no_address"]]
    print("  no-address flagged:", noaddr)

    print("Geocoding (cached)...")
    depot = geocode_all(clients)
    failed = [c["name"] for c in clients if c.get("geo_source") == "FAILED"]
    citylevel = [c["name"] for c in clients if c.get("geo_source") in ("city/zip", "zip")]
    print(f"  depot @ {depot['lat']:.5f},{depot['lon']:.5f}")
    print(f"  geocode failed: {failed}")
    print(f"  city/zip-level (eyeball): {citylevel}")

    print("Building OSRM matrix (cached)...")
    mat_path = os.path.join(CACHE, "matrix.json")
    build_matrix(depot, clients)

    with open(os.path.join(CACHE, "clients.json"), "w") as f:
        json.dump({"depot": depot, "clients": clients}, f, indent=2)
    print("Wrote cache/clients.json, cache/geocoded.csv, cache/matrix.json")


if __name__ == "__main__":
    main()
