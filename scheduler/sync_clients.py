#!/usr/bin/env python3
"""
Push the Christmas-install client list into the main app's `clients` table
(and a per-season `client_activity` row per client), so a client's greenery
work and their Christmas install history live in one record in the Clients
tab instead of two disconnected systems.

Run after prep.py (so cache/schedule.json is current for this season):

    .venv/bin/python3 sync_clients.py
    .venv/bin/python3 sync_clients.py --file "/path/to/some/other/workbook.xlsx"
    TBDG_SEASON=2027 .venv/bin/python3 sync_clients.py   # file under 2027

The CURRENT season comes from cache/schedule.json (already parsed, already
has everything, including the actual scheduled date from the crew-day
placement) and is labelled from season.py, not from a hardcoded year: the
label is the `season` column on client_activity, and a 2027 run that wrote
"2026" would land on top of the real 2026 history rather than beside it.
TBDG_SEASON overrides it (the same variable the rest of the pipeline reads).

2025/2024/2023/2022 come straight from the workbook's other season sheets,
since the pipeline cache only ever holds the current season -- each of those
sheets has its own header layout (or, for 2022, no headers at all), handled
by the per-season extractors below. Those are finished history and stay
pinned to their literal years.

Mergeable-field conflict rule (phone/email/street/city/state/zip): a value
edited in the app must survive a re-sync UNLESS the spreadsheet itself has
since changed that field -- see `merge_field` for the three-way compare
against `christmas_synced_snapshot`, which is exactly what the pipeline
pushed last time.
"""
import argparse
import asyncio
import datetime
import json
import os
import re
from typing import Any, Optional

import sys

import asyncpg
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)  # season.py is a sibling; this may be run from anywhere
from season import season_for  # noqa: E402

ENV_FILE = os.path.join(HERE, "..", "backend", ".env.supabase")
DEFAULT_XLSX = os.path.join(
    HERE, "CHRISTMAS CLIENTS - Storage - Delivery - Install +Takedown.xlsx"
)
CACHE = os.path.join(HERE, "cache", "schedule.json")

MERGE_FIELDS = ["phone", "email", "street", "city", "state", "zip"]

#: Same variable the rest of the pipeline reads (schedule.py, publish_pages.py).
SEASON_ENV = "TBDG_SEASON"

#: Seasons whose sheets are finished history. These stay literal -- the 2023
#: sheet is the 2023 season no matter what year it is read in.
HISTORICAL_SEASONS = ("2022", "2023", "2024", "2025")


def current_season() -> str:
    """The season this run is filing the cached schedule under.

    From season.py so it rolls over on 1 February without anyone editing a
    literal; TBDG_SEASON overrides it for a re-run of a past season or an
    early build of the next one. Validated, because a typo'd export would
    otherwise create a whole phantom season of client_activity rows.
    """
    override = (os.environ.get(SEASON_ENV) or "").strip()
    if override:
        if len(override) != 4 or not override.isdigit():
            raise SystemExit(
                f"{SEASON_ENV}={override!r} is not a four-digit season year "
                "(a season is named for the year its October falls in)."
            )
        return override
    return str(season_for())


def load_env(path):
    if not os.path.exists(path):
        return
    with open(path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k, v.strip().strip('"').strip("'"))


def clean_name(name) -> str:
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).strip().splitlines()[0]).strip()


def clean_str(v) -> Optional[str]:
    if v is None:
        return None
    # openpyxl hands back a whole-number-looking cell (zip codes, phone
    # numbers typed as digits) as a Python float -- "77418" reads back as
    # 77418.0 and str()'s straight to "77418.0" if not caught here. Applies
    # to any field, not just zip: whichever column got read as General/
    # Number format in the source sheet.
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = re.sub(r"\s+", " ", str(v).strip())
    return s or None


def to_date(v) -> Optional[datetime.date]:
    """asyncpg's `date` codec needs a real datetime.date to bind -- passing
    the ISO strings the extractors produce straight through raised
    `'str' object has no attribute 'toordinal'`. Every extractor hands
    install_date around as a plain ISO string (or None) right up to this
    one conversion point, since that's also the shape json.dumps(detail)
    wants for the activity row's JSONB snapshot."""
    if not v:
        return None
    try:
        return datetime.date.fromisoformat(str(v)[:10])
    except ValueError:
        return None


def money(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def header_lookup(ws, header_row=1):
    """{normalised header text: column index} -- same convention prep.py
    uses (col name, not a fixed letter, since the sheet has reshuffled
    columns between revisions before)."""
    col = {}
    for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
        if cell.value:
            col[re.sub(r"\s+", " ", str(cell.value).strip())] = cell.column
    return col


def make_h(ws, col):
    def h(r, *headers):
        for name in headers:
            c = col.get(name)
            if c:
                return ws.cell(r, c).value
        return None
    return h


# ---------------------------------------------------------------------------
# Per-season extraction. Each returns a list of dicts in the common shape:
# name, street, city, state, zip, phone, email, install_fee, takedown_fee,
# storage_fee, total, notes, install_date (ISO str or None), cancelled.
# ---------------------------------------------------------------------------

#: Season-templated field names, remembered so each is only complained about
#: once per run rather than once per client.
_FIELD_WARNED: set[str] = set()


def season_field(rec: dict, season: str, *templates, default=None):
    """Read a field whose name has the season year baked into it.

    prep.py names these `install_fee_2026`, `takedown_fee_2026`,
    `install_2026_no_install` -- the year is part of the key. prep.py is owned
    elsewhere in the pipeline, so this reads them by template instead of
    depending on one spelling: `"install_fee_{}"` is tried with this season's
    year first, then any literal alternatives given (the obvious de-year-baked
    form, `"install_fee"`).

    The failure this exists to prevent is the quiet one. If prep.py renames a
    key and this just did `rec.get("install_fee_2026")`, every client would
    sync with a blank fee and a summary reading "No date recorded" with no
    dollar figure, and nothing would say why -- a whole season filed empty. So
    a key that resolves to nothing shouts once, naming the keys the record
    actually has, and the operator can see it in the run output.

    It does NOT quietly fall back to a different year's key (`install_fee_2025`
    is right there in every record): last season's fees filed under this
    season's label is worse than a blank, because it looks correct.
    """
    keys = [t.format(season) if "{}" in t else t for t in templates]
    for key in keys:
        if key in rec:
            return rec[key]
    canonical = keys[0]
    if canonical not in _FIELD_WARNED:
        _FIELD_WARNED.add(canonical)
        prefix = templates[0].split("{}")[0]
        near = sorted(k for k in rec if k.startswith(prefix))
        print(f"  *** WARNING: no {canonical!r} on the cached record "
              f"(tried {keys!r}) -- prep.py may have renamed it. "
              f"Falling back to {default!r} for every client this run.")
        if near:
            print(f"      similar keys present: {near}")
    return default


def extract_current_season(season: str):
    """This season's clients, from the pipeline cache prep.py/schedule.py left.

    `season` is the label, not a filter: the cache only ever holds one season
    (the one prep.py was last run for), so this reads whatever is in it and
    files it under the season the caller derived.
    """
    with open(CACHE) as f:
        sched = json.load(f)
    row_dates = {}
    for day in sched["days"]:
        for s in day["stops"]:
            row_dates.setdefault(s["row"], day["date"])
    out = []
    for c in sched["all_clients"]:
        # Default False, not None: an unresolvable key must not silently drop
        # (or silently keep) clients -- season_field has already shouted.
        if season_field(c, season, "install_{}_no_install", "no_install",
                        default=False):
            continue
        name = clean_name(c.get("name"))
        if not name:
            continue
        out.append({
            "name": name,
            "street": clean_str(c.get("street")), "city": clean_str(c.get("city")),
            "state": clean_str(c.get("st")), "zip": clean_str(c.get("zip")),
            "phone": clean_str(c.get("phone")), "email": clean_str(c.get("email")),
            "install_fee": money(
                season_field(c, season, "install_fee_{}", "install_fee")),
            "takedown_fee": money(
                season_field(c, season, "takedown_fee_{}", "takedown_fee")),
            "storage_fee": money(c.get("storage_fee")),
            "total": None,
            "notes": clean_str(c.get("production_notes")),
            "install_date": row_dates.get(c["row"]),
            "cancelled": False,
        })
    return out


def extract_headered(ws, name_headers, field_map, header_row=1, data_start=2):
    """field_map: {out_key: (header_name, ...) } -- first matching header wins."""
    col = header_lookup(ws, header_row)
    h = make_h(ws, col)
    out = []
    for r in range(data_start, ws.max_row + 1):
        name = clean_name(h(r, *name_headers))
        if not name:
            continue
        rec = {"name": name}
        for out_key, headers in field_map.items():
            rec[out_key] = h(r, *headers)
        out.append(rec)
    return out


def extract_2025(ws):
    raw = extract_headered(ws, ("TBDG CLIENT",), {
        "street": ("ADDRESS",), "city": ("CITY",), "state": ("ST",), "zip": ("ZIP",),
        "phone": ("PHONE",), "email": ("EMAIL",),
        "install_fee": ("INSTALL LABOR FEE",), "takedown_fee": ("TAKEDOWN LABOR FEE",),
        "storage_fee": ("STORAGE FEE (BASED ON # OF BOXES)",),
        "total": ("TOTAL PICK UP & DELIVERY INSTALL + TAKEDOWN",),
        "notes": ("Production Notes",),
        "install_date": ("Install Date 2025",),
    })
    for rec in raw:
        for k in ("street", "city", "state", "zip", "phone", "email", "notes"):
            rec[k] = clean_str(rec.get(k))
        for k in ("install_fee", "takedown_fee", "storage_fee", "total"):
            rec[k] = money(rec.get(k))
        d = rec.get("install_date")
        rec["install_date"] = d.date().isoformat() if hasattr(d, "date") else (str(d) if d else None)
        rec["cancelled"] = False
    return raw


def extract_2024(ws):
    raw = extract_headered(ws, ("CUSTOMER NAME",), {
        "street": ("ADDRESS",), "city": ("CITY",), "state": ("ST",), "zip": ("ZIP",),
        "phone": ("PHONE",), "email": ("EMAIL",),
        "install_fee": ("TOTAL INSTALL LABOR FEE",), "takedown_fee": ("TOTAL TAKEDOWN LABOR FEE",),
        "storage_fee": ("TOTAL TBDG STORAGE FEE (BASED ON # OF BOXES)",),
        "total": ("TOTAL PICK UP & DELIVERY INSTALL + TAKEDOWN",),
    })
    for rec in raw:
        for k in ("street", "city", "state", "zip", "phone", "email"):
            rec[k] = clean_str(rec.get(k))
        for k in ("install_fee", "takedown_fee", "storage_fee", "total"):
            rec[k] = money(rec.get(k))
        rec["notes"] = None
        rec["install_date"] = None  # not tracked on this sheet
        rec["cancelled"] = False
    return raw


def extract_2023(ws):
    raw = extract_headered(ws, ("Customer Name",), {
        "street": ("Address",), "zip": ("Zip Code",),
        "phone": ("Phone #",), "email": ("E-mail",),
        "install_fee": ("Install Fee",), "takedown_fee": ("Take Dn Fee",),
        "storage_fee": ("Storage",), "total": ("Totals",),
        "notes": ("Special Notes for Installers",),
    })
    for rec in raw:
        rec["city"] = None
        rec["state"] = None
        rec["zip"] = clean_str(rec.get("zip"))
        rec["street"] = clean_str(rec.get("street"))
        rec["phone"] = clean_str(rec.get("phone"))
        rec["email"] = clean_str(rec.get("email"))
        rec["notes"] = clean_str(rec.get("notes"))
        for k in ("install_fee", "takedown_fee", "storage_fee", "total"):
            rec[k] = money(rec.get(k))
        rec["install_date"] = None
        rec["cancelled"] = False
    return raw


def extract_2022_cancelled(ws):
    """No header row on this sheet -- position is the only signal:
    0 name, 1 address, 2 zip, 3 phone, 4 email, 5 install fee, 6 takedown
    fee, 9 residential/commercial code, 11 total. These are cancellations,
    not completed installs -- flagged as such, never read as a real job."""
    out = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        name = clean_name(row[0] if len(row) > 0 else None)
        if not name:
            continue
        out.append({
            "name": name,
            "street": clean_str(row[1] if len(row) > 1 else None),
            "city": None, "state": None,
            "zip": clean_str(row[2] if len(row) > 2 else None),
            "phone": clean_str(row[3] if len(row) > 3 else None),
            "email": clean_str(row[4] if len(row) > 4 else None),
            "install_fee": money(row[5] if len(row) > 5 else None),
            "takedown_fee": money(row[6] if len(row) > 6 else None),
            "storage_fee": None,
            "total": money(row[11] if len(row) > 11 else None),
            "notes": None,
            "install_date": None,
            "cancelled": True,
        })
    return out


#
# WORDING: these two go into client_activity.summary and are rendered verbatim
# by the Clients tab, next to a badge that ALREADY shows the season year. They
# must match backend/app/apis/install_schedule/__init__.py exactly -- both
# files write these same rows, and a disagreement means the line changes
# wording depending on which path last touched the client. Change them in both,
# and run backfill_summary_wording.py for the rows already stored.
#
# "this year" and "2026 season" were both wrong here: redundant next to the
# badge, and actively misleading once the season turns over -- a 2026 row read
# "Not installing this year" while the app was planning 2027.
NOT_INSTALLING_SUMMARY = "Not installing"
#: Fallback when no install date is on record. Not "not scheduled" -- for the
#: older seasons the date simply was never tracked, which is a different thing.
NO_DATE_SUMMARY = "No date recorded"


def summarize(season: str, rec: dict) -> str:
    if rec.get("cancelled"):
        return "Cancelled before install"
    bits = []
    if rec.get("install_date"):
        bits.append(f"Scheduled {rec['install_date']}")
    else:
        bits.append(NO_DATE_SUMMARY)
    total = rec.get("total")
    if total is None and rec.get("install_fee") is not None:
        total = round((rec.get("install_fee") or 0) + (rec.get("takedown_fee") or 0)
                       + (rec.get("storage_fee") or 0), 2)
    if total:
        bits.append(f"${total:,.0f}")
    return " · ".join(bits)


def merge_field(live: Optional[str], last_synced: Optional[str], new_value: Optional[str]) -> str:
    """Three-way compare. `live` is what's in the app now, `last_synced` is
    what the previous sync pushed (None if never synced), `new_value` is
    what this sync run found in the spreadsheet.

    - No new value at all -> keep whatever's live (nothing to say here).
    - Live unchanged since last sync (or first sync ever) -> take the new
      spreadsheet value, that's the normal path.
    - Live WAS hand-edited since last sync, and the spreadsheet's value for
      this field hasn't moved -> keep the hand-edit; the spreadsheet isn't
      the one asking for a change.
    - Live was hand-edited AND the spreadsheet also changed -> the
      spreadsheet's new value wins; that's a real, intentional correction
      at the source, not a coincidence.
    """
    if new_value is None:
        return live
    if live == last_synced:
        return new_value
    if new_value == last_synced:
        return live
    return new_value


async def upsert_client(conn, rec: dict, season: str, counts: dict):
    name = rec["name"]
    # "Beaver, Austin" (this season's sheet, comma convention) and "Beaver
    # Austin" (an older row -- pre-Christmas-sync manual entry, or an
    # earlier import that didn't use commas) are the same person, but a
    # plain LOWER(TRIM(name)) match treats them as different clients and
    # inserts a duplicate every time. Stripping commas/periods before
    # comparing catches that; the stored `name` is left untouched either
    # way, so this doesn't disturb arrangements.client_name's free-text
    # matching against whichever row already exists.
    row = await conn.fetchrow(
        "SELECT id, phone, email, street, city, state, zip, christmas_synced_snapshot "
        "FROM clients WHERE regexp_replace(LOWER(TRIM(name)), '[,.]', '', 'g') "
        "= regexp_replace(LOWER(TRIM($1)), '[,.]', '', 'g')",
        name,
    )
    if row is None:
        row = await conn.fetchrow(
            "INSERT INTO clients (name, created_by) VALUES ($1, $2) "
            "ON CONFLICT (LOWER(TRIM(name))) DO NOTHING "
            "RETURNING id, phone, email, street, city, state, zip, christmas_synced_snapshot",
            name, "sync_clients.py",
        )
        if row is None:  # lost a create race against another process -- reselect
            row = await conn.fetchrow(
                "SELECT id, phone, email, street, city, state, zip, christmas_synced_snapshot "
                "FROM clients WHERE LOWER(TRIM(name)) = LOWER(TRIM($1))", name,
            )
        counts["clients_created"] += 1
    else:
        counts["clients_seen"] += 1

    last_synced = row["christmas_synced_snapshot"] or {}
    if isinstance(last_synced, str):
        last_synced = json.loads(last_synced)

    merged = {}
    changed = False
    for field in MERGE_FIELDS:
        new_val = rec.get(field)
        live_val = row[field]
        merged_val = merge_field(live_val, last_synced.get(field), new_val)
        merged[field] = merged_val
        if merged_val != live_val:
            changed = True

    new_snapshot = {f: rec.get(f) if rec.get(f) is not None else last_synced.get(f)
                     for f in MERGE_FIELDS}

    await conn.execute(
        "UPDATE clients SET phone=$2, email=$3, street=$4, city=$5, state=$6, zip=$7, "
        "christmas_synced_snapshot=$8::jsonb, christmas_synced_at=now(), updated_at=now() "
        "WHERE id=$1",
        row["id"], merged["phone"], merged["email"], merged["street"], merged["city"],
        merged["state"], merged["zip"], json.dumps(new_snapshot),
    )
    if changed:
        counts["fields_updated"] += 1

    summary = summarize(season, rec)
    detail = {k: v for k, v in rec.items() if k != "name"}

    # Staff taking a client out of the season is a decision made in the app, and
    # the app is the source of truth -- so it survives a re-sync. This script
    # builds `rec` from the schedule spreadsheet, which still lists them, and
    # would otherwise overwrite the "not installing" line with "Scheduled ..."
    # on every run, silently undoing the removal.
    #
    # The flag lives in detail.not_installing (set by the install_schedule API
    # when the scheduler saves). Carry it forward and keep the summary, while
    # still refreshing everything else the sheet is authoritative for.
    prior = await conn.fetchval(
        "SELECT detail FROM client_activity "
        " WHERE client_id=$1 AND kind='christmas_install' AND season=$2",
        row["id"], season,
    )
    if prior:
        prior = prior if isinstance(prior, dict) else json.loads(prior)
        if prior.get("not_installing"):
            detail["not_installing"] = True
            summary = NOT_INSTALLING_SUMMARY

    result = await conn.execute(
        "INSERT INTO client_activity (client_id, kind, season, summary, detail, occurred_at) "
        "VALUES ($1, 'christmas_install', $2, $3, $4::jsonb, $5::date) "
        "ON CONFLICT (client_id, kind, season) DO UPDATE SET "
        "summary=EXCLUDED.summary, detail=EXCLUDED.detail, occurred_at=EXCLUDED.occurred_at, "
        "updated_at=now()",
        row["id"], season, summary, json.dumps(detail, default=str), to_date(rec.get("install_date")),
    )
    counts["activity_rows"] += 1


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=DEFAULT_XLSX,
                     help="Workbook holding the 2025/2024/2023/2022 season sheets")
    args = ap.parse_args()

    load_env(ENV_FILE)
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise SystemExit(f"DATABASE_URL not set (checked env and {ENV_FILE})")

    seasons: list[tuple[str, list[dict]]] = []

    season = current_season()
    source = SEASON_ENV if os.environ.get(SEASON_ENV) else "season.py"
    print(f"Current season: {season} (from {source})")
    print(f"{season}: reading {CACHE}")
    current_recs = extract_current_season(season)
    print(f"  {len(current_recs)} clients")

    print(f"2025/2024/2023/2022: reading {args.file}")
    wb = openpyxl.load_workbook(args.file, data_only=True)

    def sheet_or_none(name):
        if name not in wb.sheetnames:
            print(f"  WARNING: sheet {name!r} not found, skipping that season")
            return None
        return wb[name]

    ws = sheet_or_none("2025 Christmas")
    rec_2025 = extract_2025(ws) if ws else []
    ws = sheet_or_none("2024 Christmas")
    rec_2024 = extract_2024(ws) if ws else []
    ws = sheet_or_none("2023.Christmas Analysis")
    rec_2023 = extract_2023(ws) if ws else []
    ws = sheet_or_none("Cancelled 2022")
    rec_2022 = extract_2022_cancelled(ws) if ws else []

    # The historical years are literals -- those sheets are finished seasons.
    # The last entry is the CURRENT one and its label is derived, so a 2027 run
    # writes 2027 rows beside the 2026 history instead of on top of it. (Point
    # TBDG_SEASON at a historical year and it deliberately re-files the cache
    # over that season -- the same override that lets you rebuild one.)
    historical = dict(zip(HISTORICAL_SEASONS,
                          (rec_2022, rec_2023, rec_2024, rec_2025)))
    for name, recs in [*historical.items(), (season, current_recs)]:
        print(f"  {name}: {len(recs)} rows")
        seasons.append((name, recs))

    conn = await asyncpg.connect(db_url, statement_cache_size=0)
    counts = {"clients_created": 0, "clients_seen": 0, "fields_updated": 0, "activity_rows": 0}
    try:
        # Ascending order: oldest history first, the current season last, so a
        # client's newest row is the one written most recently.
        for season_label, recs in seasons:
            for rec in recs:
                await upsert_client(conn, rec, season_label, counts)
    finally:
        await conn.close()

    print()
    print(f"Clients created: {counts['clients_created']}")
    print(f"Existing clients matched: {counts['clients_seen']}")
    print(f"Clients with a contact-field change this run: {counts['fields_updated']}")
    print(f"Activity rows written/updated: {counts['activity_rows']}")


if __name__ == "__main__":
    asyncio.run(main())
