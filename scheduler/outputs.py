#!/usr/bin/env python3
"""
TBDG 2026 Christmas -- build the Excel workbook + interactive Leaflet map
from cache/schedule.json.
"""
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "cache")
XLSX = os.path.join(HERE, "2026 Christmas Schedule.xlsx")
MAP = os.path.join(HERE, "map.html")

HDR_FILL = PatternFill("solid", fgColor="0B5D3B")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILL = {
    "M Crowd": "C0392B", "Country Club": "8E44AD", "Capital Bank": "2C7873",
    "Rotary House": "D35400", "Brenda Ryan": "B7950B", "Standard": "1F618D",
}
FLAG_FILL = PatternFill("solid", fgColor="F9E79F")
OVER_FILL = PatternFill("solid", fgColor="F5B7B1")
THIN = Side(style="thin", color="D5D8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

AREA_COLORS = {
    "Central": "#e6194B", "West": "#3cb44b", "Northwest": "#ffe119",
    "North": "#4363d8", "Northeast": "#f58231", "Southwest": "#911eb4",
    "South": "#42d4f4", "Southeast": "#f032e6", "Far West": "#9A6324",
    "Dallas - Central": "#800000", "Dallas - North": "#808000",
    "Dallas - West": "#000075", "Dallas - East": "#e6beff",
    "UNKNOWN": "#a9a9a9",
}


def merge_crew(prev, new):
    """Joint stops appear on two crews' cards -> merged label."""
    names = set()
    for c in (prev, new):
        names.update(x.strip() for x in c.replace(" (joint)", "").split(" + "))
    return " + ".join(sorted(names)) + " (joint)"


def gmaps_link(depot, stops, depot_anchored):
    def pt(o):
        return f"{o['lat']},{o['lon']}"
    if depot_anchored:
        pts = [pt(depot)] + [pt(s) for s in stops] + [pt(depot)]
    else:
        pts = [pt(s) for s in stops]
    if len(pts) < 2:
        return ""
    origin, dest = pts[0], pts[-1]
    way = "|".join(pts[1:-1])
    url = f"https://www.google.com/maps/dir/?api=1&origin={origin}&destination={dest}"
    if way:
        url += f"&waypoints={way}"
    return url


def style_header(ws, row=1):
    for c in ws[row]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_xlsx(data):
    depot = data["depot"]
    days = data["days"]
    clients = data["all_clients"]
    wb = openpyxl.Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    install_dates = sorted({d["date"] for d in days})
    total_drive = sum(d["drive_min"] for d in days) / 60.0
    total_install = sum(d["install_h"] * d["stacked_crews"] for d in days)
    over = [d for d in days if any("OVER" in f for f in d["flags"])]
    singles = [d for d in days if any("single-stop" in f for f in d["flags"])]
    placed = len({s["row"] for d in days for s in d["stops"]})
    noaddr = data["flagged_noaddr"]

    rows = [
        ["TBDG 2026 Christmas Install Schedule — Summary", ""],
        ["", ""],
        ["Clients placed (routed)", placed],
        ["Clients flagged NEED ADDRESS", len(noaddr)],
        ["Total clients", placed + len(noaddr)],
        ["Distinct install dates", len(install_dates)],
        ["Date range", f"{install_dates[0]} → {install_dates[-1]}"],
        ["Crew-days", len(days)],
        ["Total install-hours (all crews)", round(total_install, 1)],
        ["Total real drive-hours (OSRM)", round(total_drive, 1)],
        ["Avg drive per crew-day (min)", round(sum(d['drive_min'] for d in days) / len(days), 1)],
        ["Houston day shape", "arrive 8:00 · roll out 8:30 · back ~4:00-6:30pm (7.5-10h)"],
        ["Crew-days OVER 10h cap", len(over)],
        ["Single-stop days (flagged)", len(singles)],
        ["", ""],
        ["FLAGGED — need attention", ""],
    ]
    for c in noaddr:
        rows.append([f"  NEED ADDRESS: {c['name']}", "cannot map/route"])
    for c in data.get("dropped", []):
        rows.append([f"  DROPPED: {c['name']}", c["reason"]])
    for d in over:
        rows.append([f"  OVER 10h: {d['date']} {d['crew']}",
                     f"{d['total_min']} min — {'; '.join(d['flags'])}"])
    for d in singles:
        rows.append([f"  Single-stop: {d['date']} {d['crew']}",
                     f"{d['stops'][0]['name']} ({d['note']})"])
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14, color="0B5D3B")
    ws["A15"].font = Font(bold=True, color="C0392B")
    autofit(ws, [46, 40])

    # ---------------- Calendar ----------------
    ws = wb.create_sheet("Calendar")
    hdr = ["Date", "Day", "Crew", "Category", "# Stops", "Ordered Stops",
           "Install h", "Drive h", "Lunch h", "Total h", "Window", "Zones", "Flags"]
    ws.append(hdr)
    style_header(ws)
    for d in days:
        stops = " → ".join(s["name"] for s in d["stops"])
        row = [d["date"], d["dow"], d["crew"], d["category"], len(d["stops"]),
               stops, d["install_h"], round(d["drive_min"] / 60, 2),
               round(d["lunch"] / 60, 2), round(d["total_min"] / 60, 2),
               round(d.get("window_min", 600) / 60, 1), ", ".join(d["zones"]), "; ".join(d["flags"])]
        ws.append(row)
        r = ws.max_row
        fill = CAT_FILL.get(d["category"], "1F618D")
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=fill)
        ws.cell(r, 4).font = Font(color="FFFFFF", bold=True)
        if any("OVER" in f for f in d["flags"]):
            for cc in range(1, len(hdr) + 1):
                ws.cell(r, cc).fill = OVER_FILL
        elif d["flags"]:
            ws.cell(r, 13).fill = FLAG_FILL
        for cc in range(1, len(hdr) + 1):
            ws.cell(r, cc).border = BORDER
            ws.cell(r, cc).alignment = Alignment(vertical="top", wrap_text=(cc == 6))
    ws.freeze_panes = "A2"
    autofit(ws, [11, 6, 22, 14, 7, 60, 9, 8, 8, 8, 8, 26, 26])

    # ---------------- Routes ----------------
    ws = wb.create_sheet("Routes")
    hdr = ["Date", "Crew", "Seq", "Stop", "Street", "City", "Zone",
           "Drive from prev (min)", "Install h", "Google Maps"]
    ws.append(hdr)
    style_header(ws)
    for d in days:
        link = gmaps_link(depot, d["stops"], d["depot_anchored"])
        seq_names = (["DEPOT"] + [s["name"] for s in d["stops"]] + ["DEPOT"]
                     if d["depot_anchored"] else [s["name"] for s in d["stops"]])
        legs = d["legs"]
        # depot start row
        start_r = ws.max_row + 1
        if d["depot_anchored"]:
            ws.append([d["date"], d["crew"], 0, "DEPOT — 2860 Antoine Dr",
                       "", "", "", "", "", link])
            ws.cell(ws.max_row, 10).hyperlink = link
            ws.cell(ws.max_row, 10).value = "Open route ↗"
            ws.cell(ws.max_row, 10).font = Font(color="1F618D", underline="single")
        for i, s in enumerate(d["stops"]):
            leg_min = legs[i] if d["depot_anchored"] else (legs[i - 1] if i > 0 else 0)
            ws.append([d["date"], d["crew"], i + 1, s["name"], s["street"],
                       s["city"], s["zone"], leg_min, s["cal_hours"],
                       "" if d["depot_anchored"] else (link if i == 0 else "")])
            if not d["depot_anchored"] and i == 0 and link:
                ws.cell(ws.max_row, 10).hyperlink = link
                ws.cell(ws.max_row, 10).value = "Open route ↗"
                ws.cell(ws.max_row, 10).font = Font(color="1F618D", underline="single")
        if d["depot_anchored"] and d["stops"]:
            ws.append([d["date"], d["crew"], len(d["stops"]) + 1,
                       "DEPOT — return", "", "", "", legs[-1], "", ""])
        # light separator styling per day
        for rr in range(start_r, ws.max_row + 1):
            for cc in range(1, len(hdr) + 1):
                ws.cell(rr, cc).border = BORDER
    ws.freeze_panes = "A2"
    autofit(ws, [11, 22, 5, 34, 30, 14, 24, 20, 9, 14])

    # ---------------- 2026 Plan ----------------
    ws = wb.create_sheet("2026 Plan")
    hdr = ["Client", "Category", "Bus/Res", "Area", "Zone", "Street", "City",
           "ZIP", "Cal Hours", "Hours Basis", "Assigned Date", "Assigned Crew",
           "2025 Install Date", "Geo Source", "Confirm"]
    ws.append(hdr)
    style_header(ws)
    # map row -> (date, crew)
    assign = {}
    for d in days:
        for s in d["stops"]:
            if s["row"] in assign:
                assign[s["row"]] = (assign[s["row"]][0],
                                    merge_crew(assign[s["row"]][1], d["crew"]))
            else:
                assign[s["row"]] = (d["date"], d["crew"])
    dropped_names = {x["name"] for x in data.get("dropped", [])}
    for c in sorted(clients, key=lambda c: c["name"].lower()):
        date, crew = assign.get(c["row"], ("", ""))
        if c["no_address"]:
            date, crew = "NEEDS ADDRESS", ""
        elif c["name"] in dropped_names:
            date, crew = "NO INSTALL 2026", ""
        ws.append([c["name"], c["category"], c["business"], c["area"], c["zone"],
                   c["street"], c["city"], c["zip"], c.get("cal_hours"),
                   c.get("hours_basis"), date, crew, c["prior_install_date"],
                   c.get("geo_source", ""), ""])
        r = ws.max_row
        if c["no_address"]:
            for cc in range(1, len(hdr) + 1):
                ws.cell(r, cc).fill = OVER_FILL
        fill = CAT_FILL.get(c["category"], None)
        if c["category"] != "Standard":
            ws.cell(r, 2).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, 2).font = Font(color="FFFFFF", bold=True)
        for cc in range(1, len(hdr) + 1):
            ws.cell(r, cc).border = BORDER
    ws.freeze_panes = "A2"
    autofit(ws, [34, 14, 10, 16, 26, 28, 14, 8, 9, 12, 14, 22, 15, 11, 12])

    # ---------------- Capacity ----------------
    ws = wb.create_sheet("Capacity")
    hdr = ["Date", "Day", "Crew", "Category", "# Stops", "Install h", "Drive h",
           "Lunch h", "Total h", "Window h", "Slack h", "Utilization %", "Flags"]
    ws.append(hdr)
    style_header(ws)
    for d in days:
        total_h = d["total_min"] / 60
        row = [d["date"], d["dow"], d["crew"], d["category"], len(d["stops"]),
               d["install_h"], round(d["drive_min"] / 60, 2),
               round(d["lunch"] / 60, 2), round(total_h, 2),
               round(d.get("window_min", 600) / 60, 1),
               round(d.get("window_min", 600) / 60 - total_h, 2),
               round(total_h / (d.get("window_min", 600) / 60) * 100),
               "; ".join(d["flags"])]
        ws.append(row)
        r = ws.max_row
        if any("OVER" in f for f in d["flags"]):
            for cc in range(1, len(hdr) + 1):
                ws.cell(r, cc).fill = OVER_FILL
        elif d["flags"]:
            ws.cell(r, 13).fill = FLAG_FILL
        for cc in range(1, len(hdr) + 1):
            ws.cell(r, cc).border = BORDER
    ws.freeze_panes = "A2"
    autofit(ws, [11, 6, 22, 14, 7, 9, 8, 8, 8, 9, 8, 12, 26])

    wb.save(XLSX)
    print("Wrote", XLSX)


def build_map(data):
    depot = data["depot"]
    days = data["days"]
    clients = data["all_clients"]

    markers = []
    for c in clients:
        if c.get("lat") is None:
            continue
        markers.append({
            "name": c["name"], "lat": c["lat"], "lon": c["lon"],
            "area": c["area"], "zone": c["zone"], "cat": c["category"],
            "hours": c.get("cal_hours"), "bus": c["business"],
        })
    routes = []
    for d in days:
        pts = [[s["lat"], s["lon"]] for s in d["stops"] if s.get("lat")]
        if d["depot_anchored"]:
            pts = [[depot["lat"], depot["lon"]]] + pts + [[depot["lat"], depot["lon"]]]
        if len(pts) >= 2:
            # `geometry` is the real road-following path from route_geometry.py
            # (OSRM /route, distinct from the /table durations used to plan
            # the stop order); fall back to straight segments if it's missing.
            routes.append({"crew": d["crew"], "date": d["date"],
                           "cat": d["category"], "pts": pts,
                           "geom": d.get("geometry"), "mi": d.get("distance_mi")})

    # Compute a starting center+zoom from the real data bounds (self-adjusts
    # if the client roster changes next year). We deliberately do NOT rely on
    # Leaflet's fitBounds()/container-size auto-fit: some embedding contexts
    # (e.g. an iframe mid-attach) report the map container's pixel size as
    # 0x0 at load time, which makes fitBounds compute a nonsensical zoom
    # that nothing ever corrects. setView(center, zoom) needs no container
    # size at all, so it's immune to that class of bug; users can still
    # freely zoom/pan afterward.
    import math
    all_lat = [m["lat"] for m in markers] + [depot["lat"]]
    all_lon = [m["lon"] for m in markers] + [depot["lon"]]
    lat0, lat1 = min(all_lat), max(all_lat)
    lon0, lon1 = min(all_lon), max(all_lon)
    center = {"lat": (lat0 + lat1) / 2, "lon": (lon0 + lon1) / 2}
    # standard Web Mercator "zoom to fit" formula, assuming a ~1000px-wide
    # map viewport (a reasonable default for a full-page desktop map)
    lon_span = max(lon1 - lon0, 0.01)
    zoom = max(4, min(11, math.floor(math.log2(360 * 1000 / (256 * lon_span))) - 2))

    payload = json.dumps({
        "depot": {"lat": depot["lat"], "lon": depot["lon"]},
        "markers": markers, "routes": routes, "areaColors": AREA_COLORS,
        "startCenter": center, "startZoom": zoom,
    })
    html = MAP_TEMPLATE.replace("__DATA__", payload)
    with open(MAP, "w") as f:
        f.write(html)
    print("Wrote", MAP)


MAP_TEMPLATE = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>TBDG 2026 Christmas Install Map</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  /* Leaf & Ledger design tokens */
  html,body{margin:0;height:100%;font-family:'Montserrat',sans-serif;color:#292524;background:#f7f4ef}
  #map{height:100%}
  .panel{position:absolute;top:12px;right:12px;z-index:1000;background:#fff;
    padding:14px 16px;border-radius:8px;border:1px solid #e7e5e4;
    box-shadow:0 1px 2px rgba(41,37,36,.05),0 4px 16px rgba(41,37,36,.10);
    max-height:88%;overflow:auto;font-size:12.5px;min-width:220px}
  .panel h3{margin:.1em 0 .35em;font-size:15px;font-family:Georgia,serif;
    letter-spacing:.04em;font-weight:600;color:#1f3d2b}
  .legend span{display:inline-block;width:11px;height:11px;border-radius:50%;
    margin-right:7px;vertical-align:middle;box-shadow:inset 0 0 0 1px rgba(41,37,36,.2)}
  .row{margin:3px 0;cursor:pointer;color:#57534e;font-weight:500}
  .muted{color:#78716c;font-size:11px}
  label{cursor:pointer;font-weight:500;color:#292524}
  hr{border:none;border-top:1px solid #e7e5e4;margin:10px 0}
  input[type=checkbox]{accent-color:#2d5a33}
</style></head>
<body><div id="map"></div>
<div class="panel">
  <h3><svg style="width:15px;height:15px;stroke:#1f3d2b;fill:none;stroke-width:2;stroke-linecap:round;stroke-linejoin:round;vertical-align:-2px" viewBox="0 0 24 24"><path d="M12 2 7 9h2.5L5 15h3l-3.5 5h15L16 15h3l-4.5-6H17L12 2Z"/><path d="M12 22v-2"/></svg> TBDG · 2026 Christmas</h3>
  <div class="muted">Depot: 2860 Antoine Dr, Houston</div>
  <div style="margin:9px 0"><label><input type="checkbox" id="showRoutes" checked>
    show crew-day routes</label></div>
  <div id="legend" class="legend"></div>
  <hr><div id="crewlist"></div>
</div>
<script>
const DATA = __DATA__;
const map = L.map('map', {zoomAnimation:false}).setView(
  [DATA.startCenter.lat, DATA.startCenter.lon], DATA.startZoom);
L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png',
  {maxZoom:19, attribution:'© OpenStreetMap'}).addTo(map);

// depot
const depotIcon = L.divIcon({className:'',html:
  '<div style="background:#1f3d2b;color:#fff;border-radius:6px;padding:2px 7px;'+
  'font-size:11px;font-weight:700;white-space:nowrap;font-family:Montserrat,sans-serif;'+
  'letter-spacing:.03em;border:1px solid #162d20">'+
  '<svg style="width:11px;height:11px;stroke:#fff;fill:none;stroke-width:2;stroke-linecap:round;stroke-linejoin:round;vertical-align:-1px" viewBox="0 0 24 24">'+
  '<path d="m3 9 9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg> DEPOT</div>'});
L.marker([DATA.depot.lat,DATA.depot.lon],{icon:depotIcon,zIndexOffset:1000})
  .addTo(map).bindPopup('<b>DEPOT</b><br>2860 Antoine Dr, Houston TX 77092');

// markers by area color
const bounds = [];
DATA.markers.forEach(m=>{
  const col = DATA.areaColors[m.area]||'#888';
  bounds.push([m.lat,m.lon]);
  L.circleMarker([m.lat,m.lon],{radius:6,color:'#44403c',weight:1,
    fillColor:col,fillOpacity:.9}).addTo(map)
    .bindPopup(`<b>${m.name}</b><br>${m.cat} · ${m.bus}<br>${m.area} — ${m.zone}`+
      `<br>Install: ${m.hours??'?'} h`);
});
bounds.push([DATA.depot.lat,DATA.depot.lon]);
// (view is already set from DATA.startCenter/startZoom above; fitBounds is
// intentionally not used here, see the comment on that computation)

// legend
const areas = [...new Set(DATA.markers.map(m=>m.area))].sort();
document.getElementById('legend').innerHTML = areas.map(a=>
  `<div class="row"><span style="background:${DATA.areaColors[a]||'#888'}"></span>${a}</div>`).join('');

// routes
const CREW_COLORS = {"Alberto":"#c2410c","Lesly":"#0369a1","Niurka":"#2d5a33"};
function crewCol(crew){
  for(const k in CREW_COLORS) if(crew.includes(k)) return CREW_COLORS[k];
  return "#7d3c98";
}
let routeLayers = [];
function drawRoutes(){
  routeLayers.forEach(l=>map.removeLayer(l)); routeLayers=[];
  if(!document.getElementById('showRoutes').checked) return;
  DATA.routes.forEach((r,i)=>{
    // real road-following path when available; straight segments as a
    // fallback (e.g. OSRM couldn't be reached when this map was built)
    const pts = r.geom || r.pts;
    const miTxt = r.mi!=null ? `${r.mi.toFixed(1)} mi` : 'mileage unavailable';
    const pl = L.polyline(pts,{color:crewCol(r.crew),weight:r.geom?3:2.5,
      opacity:.6,dashArray:r.geom?null:'5 5'}).addTo(map)
      .bindPopup(`${r.date} · ${r.crew}<br>${r.cat}<br><b>${miTxt}</b>`);
    routeLayers.push(pl);
  });
}
drawRoutes();
document.getElementById('showRoutes').onchange=drawRoutes;

// crew-day list
const byDate = {};
DATA.routes.forEach(r=>{(byDate[r.date]=byDate[r.date]||[]).push(r)});
document.getElementById('crewlist').innerHTML =
  '<b>Crew-days: '+DATA.routes.length+'</b><br>'+
  Object.keys(byDate).sort().map(d=>
    `<div class="muted">${d}: ${byDate[d].map(r=>
      r.crew+(r.mi!=null?` (${r.mi.toFixed(0)}mi)`:'')).join(', ')}</div>`).join('');
</script></body></html>"""


def main():
    with open(os.path.join(CACHE, "schedule.json")) as f:
        data = json.load(f)
    build_xlsx(data)
    build_map(data)


if __name__ == "__main__":
    main()
