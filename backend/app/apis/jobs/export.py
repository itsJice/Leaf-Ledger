"""Job exports.

tracking_xlsx     The buyer's binder sheet, in the layout the team already
                  uses: Client · Project · Item · Vendor · SKU · Description ·
                  Need Qty · O/O Qty · Unit Cost · Adj. Unit Cost · Freight ·
                  Arrival · Checked in · Notes, grouped under Placed /
                  Placed - vendor follow up required / Need to check
                  allocations from existing orders.

manufacturing_order_pdf
                  FUTURE, not exposed by any endpoint while the designers keep
                  the paper MO (scope decision, Sept 2026).
                  The TBDG Xmas Manufacturing Order: header from intake, the
                  piece specs, the PRODUCT list (manufacturer · product number
                  · description) filled from what was actually sourced, a
                  blank TIME log for the builders, and notes.
"""
from __future__ import annotations

import io
from typing import Optional


def _money(n) -> str:
    return "" if n is None else f"${float(n):,.2f}"


def _qty(n) -> str:
    if n is None:
        return ""
    f = float(n)
    return f"{int(f)}" if f.is_integer() else f"{f:g}"


def _d(v) -> str:
    if not v:
        return ""
    try:
        return v.strftime("%b %d, %Y")
    except Exception:
        return str(v)


def _esc(s) -> str:
    return str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ── Tracking sheet (xlsx) ───────────────────────────────────────────────────
def _section_for(line: dict) -> str:
    if line["status"] == "allocated":
        return "allocations"
    if line["status"] == "follow_up" or (line.get("order_status") == "follow_up"):
        return "follow_up"
    if line["order_item_id"]:
        return "placed"
    if line["status"] in ("sold_out", "on_hold"):
        return "placed"
    return "to_order"


SECTION_TITLES = [
    ("placed", "Placed"),
    ("follow_up", "Placed - vendor follow up required"),
    ("allocations", "Need to check allocations from existing orders"),
    ("to_order", "Not yet ordered"),
]


def tracking_xlsx(job: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Order"
    from openpyxl.drawing.image import Image as XLImage
    from app.apis.orders.export import _thumb

    cols = ["Client", "Project", "Item", "Vendor", "Sku", "Description", "Need Qty", "O/O QTY",
            "Unit Cost", "Adj. Unit Cost", "Freight", "Arrival", "Checked in", "Notes", "Picture"]
    widths = [14, 14, 26, 16, 16, 40, 9, 9, 10, 12, 9, 12, 11, 48, 12]
    head_fill = PatternFill("solid", fgColor="A6A6A6")
    section_fill = PatternFill("solid", fgColor="FFF2A8")
    strike = Font(strike=True, color="666666")
    ws.append(cols)
    for i in range(1, len(cols) + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True); c.fill = head_fill
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    client = job.get("client_name") or ""
    project = job.get("collection") or job.get("name") or ""
    sections: dict = {k: [] for k, _ in SECTION_TITLES}
    for need in job["needs"]:
        if not need["lines"]:
            sections["to_order"].append((need, None))
        for line in need["lines"]:
            sections[_section_for(line)].append((need, line))

    for key, title in SECTION_TITLES:
        rows = sections[key]
        if not rows:
            continue
        ws.append([title])
        r = ws.max_row
        for i in range(1, len(cols) + 1):
            ws.cell(row=r, column=i).fill = section_fill
        ws.cell(row=r, column=1).font = Font(bold=True)
        for need, line in rows:
            if line is None:
                ws.append([client, project, need["label"], "", "", need.get("spec") or "",
                           float(need["need_qty"]), "", "", "", "", "", "", need.get("notes") or ""])
                continue
            notes = []
            if line["status"] == "sold_out":
                notes.append("sold out")
            if line["status"] == "allocated":
                notes.append(f"allocated {_qty(line['allocated_qty'])} from {line.get('allocated_order_name') or 'existing order'}")
            if line.get("overage_qty"):
                notes.append(f"{_qty(line['overage_qty'])} extra to stock")
            if line.get("notes"):
                notes.append(line["notes"])
            if line["status"] in ("sold_out", "on_hold", "allocated"):
                oo = None  # struck through on the binder sheet: nothing is on order for this line
            else:
                oo = line.get("po_quantity") if line.get("order_item_id") else (line["order_qty"] or None)
            ws.append([
                client, project, need["label"], line.get("vendor_name") or "", line.get("sku") or "",
                line.get("description") or "", float(need["need_qty"]),
                float(oo) if oo else "",
                float(line["unit_cost"]) if line.get("unit_cost") is not None else "",
                float(line["adj_unit_cost"]) if line.get("adj_unit_cost") is not None and line.get("price_per") == "pack" else "",
                "",
                "sold out" if line["status"] == "sold_out" else _d(line.get("expected_arrival")),
                _qty(line.get("received_qty")) if line.get("received_qty") else "",
                "; ".join(notes),
            ])
            r = ws.max_row
            for c in (9, 10):
                ws.cell(row=r, column=c).number_format = '"$"#,##0.00'
            for c in (7, 8):
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
            if line["status"] == "sold_out":
                for c in (4, 5, 6):
                    ws.cell(row=r, column=c).font = strike
            # The buyer pastes product pictures onto the binder sheet by hand;
            # put them on the row instead.
            png = _thumb(line.get("image_url")) if line.get("image_url") else None
            if png:
                img = XLImage(io.BytesIO(png))
                img.width, img.height = 64, 64
                ws.add_image(img, f"O{r}")
                ws.row_dimensions[r].height = 52
        ws.append([])

    # Request sheet: the intake as the client gave it.
    rq = wb.create_sheet("Request")
    rq.column_dimensions["A"].width = 6
    rq.column_dimensions["B"].width = 70
    rq.append(["", f"Color Scheme : {job.get('collection') or job.get('color_palette') or ''}"])
    for i, p in enumerate(job["pieces"], start=1):
        spec = p.get("spec") or {}
        bits = [f"{k}: {v}" for k, v in spec.items() if v not in (None, "", [])]
        rq.append([i, f"{p['piece_type']} : {_qty(p['qty'])}" + (f"  ({', '.join(bits)})" if bits else "")])
    for k, v in (job.get("intake") or {}).items():
        if v not in (None, "", False):
            rq.append(["", f"{k.replace('_', ' ').title()} : {v}"])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Manufacturing Order (pdf) ───────────────────────────────────────────────
def manufacturing_order_pdf(job: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
                                    PageBreak, KeepTogether)

    olive = colors.HexColor("#6E7A3A")
    grey = colors.HexColor("#8A8A8A")
    styles = getSampleStyleSheet()
    base = ParagraphStyle("base", parent=styles["Normal"], fontSize=9, leading=12)
    label = ParagraphStyle("label", parent=base, fontName="Helvetica-Bold")
    small = ParagraphStyle("small", parent=base, fontSize=8, leading=10, textColor=grey)
    title = ParagraphStyle("title", parent=base, fontName="Helvetica-Bold", fontSize=14, leading=18)
    brand = ParagraphStyle("brand", parent=base, fontName="Helvetica-Bold", fontSize=26, leading=28, textColor=olive)

    intake = job.get("intake") or {}

    def fld(name, val, width=None):
        v = _esc(val) if val not in (None, "") else ""
        return Paragraph(f"<b>{_esc(name)}</b>&nbsp; {v or '_______________'}", base)

    def grid(rows, col_widths, style_extra=()):
        t = Table(rows, colWidths=col_widths)
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                               ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                               ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                               *style_extra]))
        return t

    W = 7.5 * inch
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    story = []

    # Header
    kind = intake.get("client_kind") or ""
    story.append(grid([[Paragraph("tbdg", brand), Paragraph("<b>XMAS</b> MANUFACTURING ORDER", title),
                        Paragraph(f"Order # {_esc(job.get('order_no') or '')}", label)]],
                      [1.3 * inch, 4.2 * inch, 2.0 * inch],
                      [("ALIGN", (2, 0), (2, 0), "RIGHT")]))
    story.append(Spacer(1, 6))
    story.append(grid([
        [fld("Design Firm / End User", kind), fld("Name", job.get("client_name")), fld("Install Date", _d(job.get("install_date")))],
        [fld("Sidemark", job.get("sidemark")), fld("Designer", job.get("designer")), fld("Order Date", _d(job.get("order_date")))],
        [fld("Phone", intake.get("phone")), fld("Email", intake.get("email")), fld("Client Due Date", _d(job.get("due_date")))],
        [fld("Delivery / Pickup / Shipping", job.get("delivery_method")), fld("Sales", intake.get("sales")), fld("TBDG SO", intake.get("tbdg_so"))],
    ], [2.5 * inch, 3.0 * inch, 2.0 * inch]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>Color Palette:</b> {_esc(job.get('color_palette') or job.get('collection') or '')}", base))
    story.append(Spacer(1, 6))
    story.append(grid([[""]], [W], [("LINEABOVE", (0, 0), (-1, 0), 2, olive)]))

    # Piece specs, one row per piece in the order the form lists them.
    piece_rows = []
    for p in job["pieces"]:
        spec = p.get("spec") or {}
        bits = [f"{k.replace('_', ' ').title()} {v}" for k, v in spec.items() if v not in (None, "", [])]
        piece_rows.append([Paragraph(f"<b>{_esc(p['piece_type']).upper()}</b>", base),
                           Paragraph(f"Qty {_qty(p['qty'])}", base),
                           Paragraph(_esc(" · ".join(bits)), base)])
    if not piece_rows:
        piece_rows.append([Paragraph("<i>No pieces recorded</i>", small), "", ""])
    story.append(grid(piece_rows, [1.6 * inch, 0.7 * inch, 5.2 * inch],
                      [("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD"))]))
    story.append(Spacer(1, 6))

    notes_on_product = intake.get("notes_on_product") or job.get("notes") or ""
    story.append(Paragraph(f"<b>NOTES ON PRODUCT</b> {_esc(notes_on_product)}", base))
    flags = [f for f in ("Mirrored", "Matching") if intake.get(f.lower())]
    story.append(Paragraph(f"<b>Mirrored / Matching:</b> {', '.join(flags) or '—'}", base))
    story.append(Paragraph(f"<b>LOCATION OF ELECTRICAL PLUG</b> {_esc(intake.get('plug_location') or '')}", base))
    story.append(Paragraph(f"<b>LOCATION TAG</b> {_esc(intake.get('location_tag') or '')}", base))
    story.append(Spacer(1, 6))
    story.append(grid([[""]], [W], [("LINEABOVE", (0, 0), (-1, 0), 2, olive)]))
    story.append(grid([[fld("QUOTED PRICE (W / R)", intake.get("quoted_price")), fld("Build To", intake.get("build_to"))]],
                      [3.75 * inch, 3.75 * inch]))

    # Page 2: time log, product list, notes
    story.append(PageBreak())
    story.append(Paragraph(f"<para align=right>Order # {_esc(job.get('order_no') or '')}</para>", label))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>TIME</b>", label))
    time_rows = [[Paragraph("Date", small), Paragraph("Hours", small), Paragraph("Creating (garland, wreath, pulling product, etc.)", small)]]
    for _ in range(6):
        time_rows.append(["", "", ""])
    story.append(grid(time_rows, [1.0 * inch, 0.9 * inch, 5.6 * inch],
                      [("LINEBELOW", (0, 1), (-1, -1), 0.4, colors.HexColor("#999999")), ("ROWHEIGHT", (0, 1), (-1, -1), 16)]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("<b>PRODUCT</b>", label))
    prod_rows = [[Paragraph("Manufacturer", small), Paragraph("Product Number", small),
                  Paragraph("Description (size, color, texture, etc.)", small), Paragraph("Qty", small),
                  Paragraph("Status", small)]]
    for need in job["needs"]:
        lines = [l for l in need["lines"] if l["status"] != "sold_out"]
        if not lines:
            prod_rows.append(["", "", Paragraph(f"{_esc(need['label'])} {_esc(need.get('spec') or '')}", base),
                              Paragraph(_qty(need["need_qty"]), base), Paragraph("shelf" if need["ready"] else "unsourced", small)])
            continue
        for l in lines:
            if l["status"] == "allocated":
                status = f"allocated from {l.get('allocated_order_name') or 'open order'}"
            elif l.get("order_item_id"):
                got = l.get("received_qty") or 0
                status = "received" if got >= (l.get("po_quantity") or 0) else ("on order" if not got else f"partial {_qty(got)}")
            else:
                status = l["status"].replace("_", " ")
            prod_rows.append([
                Paragraph(_esc(l.get("vendor_name") or ""), base),
                Paragraph(_esc(l.get("sku") or ""), base),
                Paragraph(f"{_esc(need['label'])} — {_esc(l.get('description') or '')}", base),
                Paragraph(_qty(l.get("order_qty") or l.get("allocated_qty") or need["need_qty"]), base),
                Paragraph(_esc(status), small),
            ])
    while len(prod_rows) < 8:
        prod_rows.append(["", "", "", "", ""])
    story.append(grid(prod_rows, [1.3 * inch, 1.3 * inch, 3.4 * inch, 0.5 * inch, 1.0 * inch],
                      [("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
                       ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2"))]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("<b>NOTES</b>", label))
    note_lines = []
    for t in job.get("tasks") or []:
        if not t.get("done_at"):
            note_lines.append(f"Follow-up: {t['title']}" + (f" ({t['assignee']})" if t.get("assignee") else ""))
    if job.get("notes"):
        note_lines.insert(0, job["notes"])
    note_rows = [[Paragraph(_esc(n), base)] for n in note_lines] or []
    while len(note_rows) < 5:
        note_rows.append([""])
    story.append(grid(note_rows, [W], [("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
                                       ("ROWHEIGHT", (0, 0), (-1, -1), 16)]))

    doc.build(story)
    return buf.getvalue()
