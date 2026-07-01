from __future__ import annotations

import hashlib
import json
import mimetypes
import os
import re
import subprocess
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from xml.etree import ElementTree

import asyncpg
import databutton as db
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel

router = APIRouter(tags=["recipe_intelligence"])

DATABASE_URL = os.environ.get("DATABASE_URL")
SOURCE_ROOT = Path(os.environ.get("RECIPE_INTELLIGENCE_ROOT", "/Users/justice/Documents/TBDG Pricing Recipies "))
PARSER_VERSION = "2026-06-03.v1"
PREVIEW_CACHE_ROOT = Path(os.environ.get("RECIPE_PREVIEW_CACHE_ROOT", "/tmp/leaf-ledger-visual-previews"))

KNOWN_PREFIX_LABELS = {
    "TT": "Tree",
    "OR": "Orchid Arrangement",
    "WG": "Greenery Arrangement",
    "SG": "Succulent Arrangement",
    "CG": "Container Garden",
    "FP": "Foliage Arrangement",
    "TL": "Tree / Plant",
    "SM": "Moss Arrangement",
    "DR": "Drop-in Arrangement",
    "CT": "Container Arrangement",
    "DI": "Drop-in Arrangement",
    "GT": "Greenery Tree",
    "PV": "Plant / Vase",
    "PM": "Premade",
}
KNOWN_EXAMPLE_FILES = {"FP3-5322-2023.xlsx", "OR7-73820-2023.xlsx"}
LOCAL_STORE_KEY = "leaf-ledger-recipe-intelligence-v1"
CANONICAL_BUILD_ALIASES = {
    "Arrangement": ["Arrangement", "Orchid Arrangement", "Succulent Arrangement", "Greenery Arrangement", "Foliage Arrangement"],
    "Planter": ["Planter", "Container Garden", "Plant / Vase", "Container Arrangement"],
    "Drop-in Arrangement": ["Drop-in Arrangement", "Drop in", "Drop-in", "Dropin Arrangement"],
    "Tree": ["Tree", "Tree / Plant", "Greenery Tree"],
    "Christmas Tree": ["Christmas Tree"],
    "Garland": ["Garland"],
    "Wreath": ["Wreath"],
}
DEFAULT_PRICING_RULES = {
    "landed_cost_multiplier": 1.2,
    "retail_multiplier": 6.0,
    "wholesale_multiplier": 3.0,
    "arrangement_markup_multiplier": 1.25,
    "completed_history_policy": "approved_paid_purchased_only",
}


def build_type_aliases(build_type: str) -> list[str]:
    requested = clean_text(build_type) or "Custom Arrangement"
    for canonical, aliases in CANONICAL_BUILD_ALIASES.items():
        if requested.lower() == canonical.lower() or requested.lower() in {alias.lower() for alias in aliases}:
            return aliases
    return [requested]


async def get_conn():
    return await asyncpg.connect(DATABASE_URL, statement_cache_size=0)


class ImportRequest(BaseModel):
    root_path: Optional[str] = None
    limit: int = 40
    include_assets: bool = True
    reset_failed: bool = False


class SuggestRequest(BaseModel):
    build_type: str
    height: Optional[str] = None
    width: Optional[str] = None
    depth: Optional[str] = None
    length: Optional[str] = None
    quantity: int = 1
    notes: Optional[str] = None


class PricingRuleUpdate(BaseModel):
    rules: dict[str, Any]
    project_id: Optional[int] = None


class SkuStandardUpdate(BaseModel):
    prefix: str
    label: str
    description: Optional[str] = None
    active: bool = True


class HistoricalCompleteRequest(BaseModel):
    arrangement_id: int
    container_id: Optional[int] = None
    finished_sku: str
    completion_status: str = "approved_paid_purchased"
    notes: Optional[str] = None


def clean_text(value: Any) -> str:
    return str(value or "").replace("\n", " ").strip()


def normalize_code(value: str) -> str:
    text = clean_text(value).upper()
    text = re.sub(r"\s+", "", text)
    text = text.replace("_", "-")
    text = re.sub(r"-{2,}", "-", text)
    return text.strip("-")


def infer_item_code(text: str) -> Optional[str]:
    raw = clean_text(text)
    pattern = re.compile(
        r"\b(TT|OR|WG|SG|CG|FP|TL|SM|DR|CT|PV|DI|GT|PM)\s*[-_]?\s*[A-Z0-9.]+(?:\s*[-_]\s*[A-Z0-9]+)*\b",
        re.IGNORECASE,
    )
    match = pattern.search(raw)
    if not match:
        return None
    return normalize_code(match.group(0))


def item_prefix(item_code: Optional[str]) -> Optional[str]:
    if not item_code:
        return None
    match = re.match(r"([A-Z]+)", item_code)
    return match.group(1) if match else None


def infer_build_type(item_code: Optional[str], description: str, path: str = "") -> str:
    text = f"{description} {item_code or ''} {path}".lower()
    prefix = item_prefix(item_code)
    if "orchid" in text or prefix == "OR":
        return "Orchid Arrangement"
    if any(word in text for word in ["tree", "fiddle", "palm", "yucca", "schefflera"]) or prefix == "TT":
        return "Tree"
    if "garland" in text:
        return "Garland"
    if "wreath" in text:
        return "Wreath"
    if any(word in text for word in ["succulent", "cactus", "echeveria", "agave"]) or prefix in {"SG", "CG"}:
        return "Succulent Arrangement" if prefix == "SG" else "Container Garden"
    if any(word in text for word in ["wall", "green wall"]):
        return "Green Wall"
    if any(word in text for word in ["foliage", "grass", "plant", "greenery"]) or prefix in {"WG", "FP"}:
        return KNOWN_PREFIX_LABELS.get(prefix or "", "Foliage Arrangement")
    return KNOWN_PREFIX_LABELS.get(prefix or "", "Custom Arrangement")


def infer_component_label(row: dict[str, Any], build_type: str = "") -> str:
    text = " ".join(
        clean_text(row.get(key))
        for key in ["product_sku", "vendor", "description", "product_name"]
    ).lower()
    if any(word in text for word in ["container", "pot", "vase", "cylinder", "bowl", "urn", "plate", "zinc", "concrete", "glass"]):
        return "Container"
    if "foam" in text or "styro" in text:
        return "Foam"
    if "moss" in text or "lichen" in text or "palm fiber" in text:
        return "Moss / Fiber"
    if "rock" in text or "stone" in text or "gravel" in text or "top dressing" in text:
        return "Top Dressing"
    if any(word in text for word in ["branch", "trunk", "dragonwood", "wood", "birch"]):
        return "Trunks & Branches"
    if "orchid" in text or "phalaenopsis" in text:
        return "Orchid Stem"
    if any(word in text for word in ["succulent", "cactus", "echeveria", "agave", "sedum"]):
        return "Succulent / Cactus"
    if any(word in text for word in ["leaf", "leaves", "foliage", "grass", "palm", "bush", "spray", "stem", "fern"]):
        return "Leaves / Greenery" if "tree" in build_type.lower() else "Foliage / Greenery"
    if any(word in text for word in ["ribbon", "ornament", "decor", "ball", "berry", "pinecone"]):
        return "Decor"
    if "mechanic" in text:
        return "Mechanics"
    return "Product"


def safe_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace("$", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def stable_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def json_dumps(value: Any) -> str:
    return json.dumps(value, default=str)


def load_item_meta_sidecar(item_id: int) -> dict[str, Any]:
    path = Path(__file__).resolve().parents[1] / "arrangements" / "container_item_meta.local.json"
    try:
        data = json.loads(path.read_text())
        meta = data.get(str(item_id), {}) if isinstance(data, dict) else {}
        return meta if isinstance(meta, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def dimensions_from_scope_notes(notes: Optional[str]) -> dict[str, str]:
    dimensions: dict[str, str] = {}
    for line in (notes or "").splitlines():
        if line.startswith("LL_BUILD_INTELLIGENCE:"):
            continue
        if ":" not in line:
            continue
        label, value = line.split(":", 1)
        key = label.strip().lower().replace(" / ", "_").replace(" ", "_")
        if key in {"height", "width_canopy", "depth_density"} and value.strip():
            dimensions[key] = value.strip()
    return dimensions


def empty_local_store() -> dict[str, Any]:
    return {
        "sources": {},
        "recipes": {},
        "components": [],
        "visual_refs": {},
        "pricing_rules": {"global": DEFAULT_PRICING_RULES, "projects": {}, "sources": {}},
        "sku_standards": {},
        "completed_builds": [],
    }


def load_local_store() -> dict[str, Any]:
    store = db.storage.json.get(LOCAL_STORE_KEY, default=empty_local_store())
    base = empty_local_store()
    if not isinstance(store, dict):
        return base
    for key, value in base.items():
        store.setdefault(key, value)
    store.setdefault("pricing_rules", {}).setdefault("global", DEFAULT_PRICING_RULES)
    store.setdefault("pricing_rules", {}).setdefault("projects", {})
    store.setdefault("pricing_rules", {}).setdefault("sources", {})
    return store


def save_local_store(store: dict[str, Any]) -> None:
    db.storage.json.put(LOCAL_STORE_KEY, store)


def local_source_id(path: Path) -> str:
    return str(path)


def local_upsert_source(store: dict[str, Any], path: Path, root: Path, status: str = "pending", metadata: Optional[dict[str, Any]] = None) -> str:
    source_id = local_source_id(path)
    rel = str(path.relative_to(root)) if path.is_relative_to(root) else path.name
    existing = store["sources"].get(source_id, {})
    item_code = infer_item_code(path.stem)
    next_status = existing.get("status") if existing.get("status") in {"parsed", "asset_linked", "reference_only", "unsupported_deferred"} else status
    store["sources"][source_id] = {
        **existing,
        "id": source_id,
        "source_path": str(path),
        "relative_path": rel,
        "file_name": path.name,
        "extension": path.suffix.lower() or "[none]",
        "file_kind": file_kind(path),
        "sha256": existing.get("sha256") or stable_hash(path),
        "size_bytes": path.stat().st_size,
        "status": next_status,
        "item_code": existing.get("item_code") or item_code,
        "linked_item_code": existing.get("linked_item_code") or item_code,
        "parser_version": PARSER_VERSION,
        "error_message": existing.get("error_message"),
        "metadata": {**(existing.get("metadata") or {}), **(metadata or {})},
        "updated_at": datetime.utcnow().isoformat(),
    }
    return source_id


def local_refresh_derived_standards(store: dict[str, Any]) -> None:
    counts: dict[str, int] = {}
    examples: dict[str, list[str]] = {}
    for recipe in store["recipes"].values():
        prefix = recipe.get("product_family")
        code = recipe.get("item_code")
        if not prefix:
            continue
        counts[prefix] = counts.get(prefix, 0) + 1
        if code:
            examples.setdefault(prefix, [])
            if code not in examples[prefix]:
                examples[prefix].append(code)
    for prefix, count in counts.items():
        existing = store["sku_standards"].get(prefix, {})
        description = existing.get("description")
        if not description or str(description).startswith("Inferred from "):
            description = f"Inferred from {count} historical recipe file(s)."
        store["sku_standards"][prefix] = {
            "prefix": prefix,
            "label": existing.get("label") or KNOWN_PREFIX_LABELS.get(prefix, prefix),
            "description": description,
            "inferred_count": count,
            "examples": examples.get(prefix, [])[:8],
            "active": existing.get("active", True),
            "updated_at": datetime.utcnow().isoformat(),
        }


def local_classify_or_parse_file(store: dict[str, Any], path: Path, root: Path) -> dict[str, Any]:
    source_id = local_upsert_source(store, path, root)
    source = store["sources"][source_id]
    if source.get("status") not in {"pending", "failed_needs_review"}:
        return {"path": str(path), "status": source.get("status")}
    ext = path.suffix.lower()
    try:
        if ext == ".xlsx":
            parsed = parse_xlsx_recipe(path)
            store["recipes"][source_id] = {
                "source_file_id": source_id,
                "item_code": parsed.get("item_code"),
                "customer_item_code": parsed.get("customer_item_code"),
                "product_family": parsed.get("product_family"),
                "build_type": parsed.get("build_type"),
                "description": parsed.get("description"),
                "source_collection": str(path.relative_to(root).parts[0]) if path.is_relative_to(root) else None,
                "recipe_year": parsed.get("recipe_year"),
                "dimensions": parsed.get("dimensions") or {},
                "container_details": parsed.get("container_details") or {},
                "pricing_summary": parsed.get("pricing_summary") or {},
                "raw_header": parsed.get("raw_header") or {},
                "updated_at": datetime.utcnow().isoformat(),
            }
            store["components"] = [row for row in store["components"] if row.get("recipe_id") != source_id]
            for component in parsed.get("components") or []:
                store["components"].append({"recipe_id": source_id, **component})
            source.update({
                "status": "parsed",
                "item_code": parsed.get("item_code"),
                "linked_item_code": parsed.get("item_code"),
                "error_message": None,
                "metadata": {**(source.get("metadata") or {}), "components": len(parsed.get("components") or [])},
            })
            return {"path": str(path), "status": "parsed", "item_code": parsed.get("item_code")}
        if ext == ".docx":
            parsed = parse_pricing_doc(path)
            if parsed["rules"]:
                store["pricing_rules"]["sources"][f"source:{path.name}"] = parsed["rules"]
            source.update({"status": "reference_only", "metadata": {**(source.get("metadata") or {}), **parsed}, "error_message": None})
            return {"path": str(path), "status": "reference_only"}
        if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".psd"}:
            item_code = infer_item_code(path.stem)
            store["visual_refs"][source_id] = {
                "id": len(store["visual_refs"]) + 1,
                "source_file_id": source_id,
                "item_code": item_code,
                "file_path": str(path),
                "file_name": path.name,
                "extension": ext,
                "asset_type": "psd_source" if ext == ".psd" else "image",
                "status": "indexed",
                "metadata": {"folder": str(path.parent)},
                "updated_at": datetime.utcnow().isoformat(),
            }
            source.update({"status": "asset_linked", "linked_item_code": item_code or source.get("linked_item_code"), "error_message": None})
            return {"path": str(path), "status": "asset_linked"}
        source.update({"status": "unsupported_deferred", "error_message": f"{ext} registered for later parser support"})
        return {"path": str(path), "status": "unsupported_deferred"}
    except Exception as exc:
        source.update({"status": "failed_needs_review", "error_message": str(exc)[:1000]})
        return {"path": str(path), "status": "failed_needs_review", "error": str(exc)}


def local_import_pending_files(root: Path, limit: int, include_assets: bool, reset_failed: bool) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    store = load_local_store()
    for path in files_to_scan(root):
        if not include_assets and file_kind(path) == "image_asset":
            continue
        local_upsert_source(store, path, root)
    if reset_failed:
        for source in store["sources"].values():
            if source.get("status") == "failed_needs_review":
                source["status"] = "pending"
                source["error_message"] = None
    extension_rank = {".xlsx": 1, ".docx": 2, ".png": 3, ".jpg": 4, ".jpeg": 5, ".psd": 6}
    pending = sorted(
        [source for source in store["sources"].values() if source.get("status") == "pending"],
        key=lambda source: (
            0 if source.get("file_name") in KNOWN_EXAMPLE_FILES else 1,
            extension_rank.get(source.get("extension"), 9),
            source.get("source_path", ""),
        ),
    )[: max(1, min(limit, 250))]
    results = [local_classify_or_parse_file(store, Path(source["source_path"]), root) for source in pending]
    local_refresh_derived_standards(store)
    save_local_store(store)
    return results, store


def local_summary_payload(store: Optional[dict[str, Any]] = None) -> dict[str, Any]:
    store = store or load_local_store()
    statuses: dict[str, int] = {}
    extensions: dict[str, int] = {}
    for source in store["sources"].values():
        statuses[source.get("status", "pending")] = statuses.get(source.get("status", "pending"), 0) + 1
        ext = source.get("extension", "[none]")
        extensions[ext] = extensions.get(ext, 0) + 1
    failures = [
        {"relative_path": source.get("relative_path"), "status": source.get("status"), "error_message": source.get("error_message")}
        for source in store["sources"].values()
        if source.get("status") in {"failed_needs_review", "unsupported_deferred"}
    ][:25]
    return {
        "source_root": str(SOURCE_ROOT),
        "statuses": [{"status": key, "count": value} for key, value in sorted(statuses.items())],
        "extensions": [{"extension": key, "count": value} for key, value in sorted(extensions.items(), key=lambda item: item[1], reverse=True)],
        "recipe_count": len(store["recipes"]),
        "component_count": len(store["components"]),
        "asset_count": len(store["visual_refs"]),
        "failures": failures,
        "parser_version": f"{PARSER_VERSION}.local",
        "storage": "local",
    }


def local_build_types_payload() -> list[dict[str, Any]]:
    store = load_local_store()
    counts: dict[str, dict[str, Any]] = {}
    for recipe in store["recipes"].values():
        build_type = recipe.get("build_type")
        if not build_type:
            continue
        row = counts.setdefault(build_type, {"label": build_type, "evidence_count": 0, "prefixes": set()})
        row["evidence_count"] += 1
        if recipe.get("product_family"):
            row["prefixes"].add(recipe["product_family"])
    result = []
    for row in counts.values():
        result.append({**row, "prefixes": sorted(row["prefixes"])})
    return sorted(result, key=lambda row: (-row["evidence_count"], row["label"]))


def local_suggest_payload(body: SuggestRequest) -> dict[str, Any]:
    store = load_local_store()
    build_type = body.build_type.strip() or "Custom Arrangement"
    aliases = {alias.lower() for alias in build_type_aliases(build_type)}
    recipe_ids = [key for key, recipe in store["recipes"].items() if clean_text(recipe.get("build_type")).lower() in aliases]
    grouped: dict[str, dict[str, Any]] = {}
    for component in store["components"]:
        if component.get("recipe_id") not in recipe_ids:
            continue
        label = component.get("component_label") or "Product"
        row = grouped.setdefault(label, {"quantities": [], "totals": [], "vendors": set(), "descriptions": set()})
        if safe_float(component.get("quantity")):
            row["quantities"].append(float(component["quantity"]))
        if safe_float(component.get("extended_total")):
            row["totals"].append(float(component["extended_total"]))
        if clean_text(component.get("vendor")):
            row["vendors"].add(clean_text(component.get("vendor")))
        if clean_text(component.get("description")):
            row["descriptions"].add(clean_text(component.get("description")))
    if not grouped:
        grouped = {
            label: {"quantities": [1], "totals": [], "vendors": set(), "descriptions": set()}
            for label in ["Container", "Foliage / Greenery", "Foam", "Top Dressing"]
        }
    components = []
    for label, data in sorted(grouped.items(), key=lambda item: (-len(item[1]["quantities"]), item[0]))[:8]:
        quantities = sorted(data["quantities"] or [1])
        median = quantities[len(quantities) // 2]
        avg_quantity = sum(quantities) / len(quantities)
        totals = data["totals"]
        vendors = sorted(data["vendors"])[:6]
        descriptions = sorted(data["descriptions"])[:6]
        terms = [label, *vendors[:3], *descriptions[:3]]
        components.append({
            "label": label,
            "suggested_quantity": round(float(median or avg_quantity or 1), 2),
            "average_quantity": round(float(avg_quantity or 1), 2),
            "evidence_count": len(quantities) if recipe_ids else 0,
            "average_extended_total": (sum(totals) / len(totals)) if totals else None,
            "vendors": vendors,
            "examples": descriptions,
            "search_terms": [clean_text(term) for term in terms if clean_text(term)][:8],
        })
    totals = []
    for recipe_id in recipe_ids:
        pricing = store["recipes"].get(recipe_id, {}).get("pricing_summary") or {}
        total = pricing.get("total", {}).get("value") if isinstance(pricing.get("total"), dict) else None
        if safe_float(total):
            totals.append(float(total))
    return {
        "build_type": build_type,
        "input": body.model_dump(),
        "evidence_count": len(recipe_ids),
        "confidence": "high" if len(recipe_ids) >= 10 else "medium" if len(recipe_ids) >= 3 else "starter",
        "components": components,
        "cost_range": {
            "avg_total": (sum(totals) / len(totals)) if totals else None,
            "min_total": min(totals) if totals else None,
            "max_total": max(totals) if totals else None,
        },
    }


def local_pricing_rules_payload(project_id: Optional[int] = None) -> dict[str, Any]:
    store = load_local_store()
    rules = store["pricing_rules"]
    project_key = str(project_id) if project_id else None
    return {
        "global_rules": rules.get("global") or DEFAULT_PRICING_RULES,
        "project_rules": rules.get("projects", {}).get(project_key) if project_key else None,
        "source_rules": [
            {"scope": scope, "rules": rule, "updated_at": ""}
            for scope, rule in sorted((rules.get("sources") or {}).items())
        ],
    }


def local_update_pricing_rules(body: PricingRuleUpdate) -> dict[str, bool]:
    store = load_local_store()
    if body.project_id:
        store["pricing_rules"].setdefault("projects", {})[str(body.project_id)] = body.rules
    else:
        store["pricing_rules"]["global"] = body.rules
    save_local_store(store)
    return {"ok": True}


def local_sku_standards_payload() -> list[dict[str, Any]]:
    store = load_local_store()
    local_refresh_derived_standards(store)
    save_local_store(store)
    return sorted(store["sku_standards"].values(), key=lambda row: (-int(row.get("inferred_count") or 0), row.get("prefix") or ""))


def local_update_sku_standard(body: SkuStandardUpdate) -> dict[str, bool]:
    store = load_local_store()
    prefix = normalize_code(body.prefix)
    existing = store["sku_standards"].get(prefix, {})
    store["sku_standards"][prefix] = {
        "prefix": prefix,
        "label": body.label,
        "description": body.description,
        "inferred_count": int(existing.get("inferred_count") or 0),
        "examples": existing.get("examples") or [],
        "active": body.active,
        "updated_at": datetime.utcnow().isoformat(),
    }
    save_local_store(store)
    return {"ok": True}


def local_visual_references_payload(item_code: Optional[str] = None, limit: int = 50) -> list[dict[str, Any]]:
    store = load_local_store()
    refs = list(store["visual_refs"].values())
    if item_code:
        code = normalize_code(item_code)
        refs = [row for row in refs if row.get("item_code") == code]
    refs.sort(key=lambda row: row.get("updated_at") or "", reverse=True)
    return refs[: min(limit, 200)]


def local_visual_reference_by_id(asset_id: int) -> Optional[dict[str, Any]]:
    store = load_local_store()
    for row in store["visual_refs"].values():
        if int(row.get("id") or 0) == asset_id:
            return row
    return None


def validate_preview_path(file_path: str) -> Path:
    path = Path(file_path).expanduser().resolve()
    try:
        path.relative_to(SOURCE_ROOT.expanduser().resolve())
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="Preview file is outside the recipe source folder") from exc
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Preview file not found")
    return path


def generated_psd_preview(path: Path) -> Path:
    cache_key = hashlib.sha256(f"{path}:{path.stat().st_mtime_ns}:{path.stat().st_size}".encode("utf-8")).hexdigest()
    output_path = PREVIEW_CACHE_ROOT / f"{cache_key}.png"
    if output_path.exists() and output_path.stat().st_size > 0:
        return output_path
    PREVIEW_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        ["/usr/bin/sips", "-Z", "720", "-s", "format", "png", str(path), "--out", str(output_path)],
        capture_output=True,
        text=True,
        timeout=30,
        check=False,
    )
    if result.returncode != 0 or not output_path.exists() or output_path.stat().st_size == 0:
        try:
            output_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise HTTPException(status_code=415, detail="PSD preview conversion failed")
    return output_path


def preview_file_response(row: dict[str, Any]) -> FileResponse:
    path = validate_preview_path(str(row.get("file_path") or ""))
    ext = path.suffix.lower()
    if ext == ".psd":
        preview_path = generated_psd_preview(path)
        return FileResponse(preview_path, media_type="image/png", filename=f"{path.stem}.png")
    if ext not in {".png", ".jpg", ".jpeg", ".gif", ".webp"}:
        raise HTTPException(status_code=415, detail="Preview is only available for browser-readable image files")
    media_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return FileResponse(path, media_type=media_type, filename=path.name)


def file_kind(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return "spreadsheet"
    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".psd"}:
        return "image_asset"
    if ext in {".docx", ".pdf", ".pptx", ".csv"}:
        return "reference"
    if ext in {".zip"}:
        return "archive"
    return "unsupported"


async def ensure_schema(conn):
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS recipe_source_files (
            id SERIAL PRIMARY KEY,
            source_path TEXT UNIQUE NOT NULL,
            relative_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            extension TEXT NOT NULL,
            file_kind TEXT NOT NULL,
            sha256 TEXT,
            size_bytes BIGINT,
            status TEXT NOT NULL DEFAULT 'pending',
            item_code TEXT,
            linked_item_code TEXT,
            parser_version TEXT,
            error_message TEXT,
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS historical_recipes (
            id SERIAL PRIMARY KEY,
            source_file_id INTEGER UNIQUE REFERENCES recipe_source_files(id) ON DELETE CASCADE,
            item_code TEXT,
            customer_item_code TEXT,
            product_family TEXT,
            build_type TEXT,
            description TEXT,
            source_collection TEXT,
            recipe_year INTEGER,
            dimensions JSONB DEFAULT '{}'::jsonb,
            container_details JSONB DEFAULT '{}'::jsonb,
            pricing_summary JSONB DEFAULT '{}'::jsonb,
            raw_header JSONB DEFAULT '{}'::jsonb,
            visual_reference_count INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS historical_recipe_components (
            id SERIAL PRIMARY KEY,
            recipe_id INTEGER REFERENCES historical_recipes(id) ON DELETE CASCADE,
            line_order INTEGER NOT NULL DEFAULT 0,
            component_label TEXT,
            vendor TEXT,
            supplier_sku TEXT,
            description TEXT,
            quantity NUMERIC,
            first_cost NUMERIC,
            landed_cost NUMERIC,
            retail NUMERIC,
            extended_total NUMERIC,
            formulas JSONB DEFAULT '{}'::jsonb,
            raw_row JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS visual_reference_assets (
            id SERIAL PRIMARY KEY,
            source_file_id INTEGER UNIQUE REFERENCES recipe_source_files(id) ON DELETE CASCADE,
            item_code TEXT,
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            extension TEXT NOT NULL,
            asset_type TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'indexed',
            metadata JSONB DEFAULT '{}'::jsonb,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS recipe_pricing_rules (
            id SERIAL PRIMARY KEY,
            scope TEXT NOT NULL DEFAULT 'global',
            project_id INTEGER NOT NULL DEFAULT 0,
            rules JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_by TEXT,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (scope, project_id)
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS sku_standards (
            prefix TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            description TEXT,
            inferred_count INTEGER NOT NULL DEFAULT 0,
            examples JSONB DEFAULT '[]'::jsonb,
            active BOOLEAN NOT NULL DEFAULT TRUE,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS completed_historical_builds (
            id SERIAL PRIMARY KEY,
            arrangement_id INTEGER NOT NULL,
            container_id INTEGER,
            finished_sku TEXT NOT NULL,
            build_type TEXT,
            dimensions JSONB DEFAULT '{}'::jsonb,
            component_snapshot JSONB DEFAULT '[]'::jsonb,
            pricing_snapshot JSONB DEFAULT '{}'::jsonb,
            completion_status TEXT NOT NULL,
            notes TEXT,
            completed_at TIMESTAMP DEFAULT NOW()
        )
    """)


async def upsert_source_file(conn, path: Path, root: Path, status: str = "pending", metadata: Optional[dict[str, Any]] = None) -> int:
    rel = str(path.relative_to(root)) if path.is_relative_to(root) else path.name
    ext = path.suffix.lower() or "[none]"
    item_code = infer_item_code(path.stem)
    file_hash = stable_hash(path)
    duplicate = await conn.fetchrow(
        "SELECT id, source_path FROM recipe_source_files WHERE sha256 = $1 AND source_path != $2 AND status IN ('parsed', 'asset_linked', 'reference_only') LIMIT 1",
        file_hash,
        str(path),
    )
    next_status = "duplicate" if duplicate else status
    next_metadata = metadata or {}
    if duplicate:
        next_metadata = {**next_metadata, "duplicate_of": duplicate["source_path"]}
    return await conn.fetchval("""
        INSERT INTO recipe_source_files (
            source_path, relative_path, file_name, extension, file_kind, sha256, size_bytes,
            status, item_code, linked_item_code, parser_version, error_message, metadata, updated_at
        )
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $9, $10, NULL, $11::jsonb, NOW())
        ON CONFLICT (source_path) DO UPDATE SET
            relative_path = EXCLUDED.relative_path,
            file_name = EXCLUDED.file_name,
            extension = EXCLUDED.extension,
            file_kind = EXCLUDED.file_kind,
            sha256 = EXCLUDED.sha256,
            size_bytes = EXCLUDED.size_bytes,
            status = CASE
                WHEN recipe_source_files.status = 'parsed' AND EXCLUDED.status = 'pending' THEN recipe_source_files.status
                ELSE EXCLUDED.status
            END,
            item_code = COALESCE(EXCLUDED.item_code, recipe_source_files.item_code),
            linked_item_code = COALESCE(EXCLUDED.linked_item_code, recipe_source_files.linked_item_code),
            parser_version = EXCLUDED.parser_version,
            error_message = EXCLUDED.error_message,
            metadata = COALESCE(recipe_source_files.metadata, '{}'::jsonb) || EXCLUDED.metadata,
            updated_at = NOW()
        RETURNING id
    """, str(path), rel, path.name, ext, file_kind(path), file_hash, path.stat().st_size, next_status, item_code, PARSER_VERSION, json_dumps(next_metadata))


def extract_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml")
    except Exception:
        return ""
    root = ElementTree.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    parts = []
    for node in root.findall(".//w:t", ns):
        if node.text:
            parts.append(node.text)
    return "\n".join(parts)


def parse_pricing_doc(path: Path) -> dict[str, Any]:
    text = extract_docx_text(path)
    rules: dict[str, Any] = {}
    if re.search(r"FC\s*\*\s*1\.2\s*=\s*LC", text, re.IGNORECASE):
        rules["landed_cost_multiplier"] = 1.2
    if re.search(r"LC\s*\*\s*6\s*=\s*R", text, re.IGNORECASE):
        rules["retail_multiplier"] = 6.0
    if re.search(r"LC\s*\*\s*3\s*=\s*W", text, re.IGNORECASE):
        rules["wholesale_multiplier"] = 3.0
    if re.search(r"\*\s*1\.25", text):
        rules["arrangement_markup_multiplier"] = 1.25
    return {"text_excerpt": text[:2000], "rules": rules}


def find_header_row(ws) -> Optional[int]:
    for r in range(1, min(ws.max_row, 80) + 1):
        values = [clean_text(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 16) + 1)]
        row_text = " ".join(values)
        if ("product" in row_text or "product sku" in row_text) and "vendor" in row_text and ("first cost" in row_text or " fc" in f" {row_text}"):
            return r
    return None


def adjacent_value(ws, row: int, col: int) -> Any:
    for offset in range(1, 4):
        value = ws.cell(row, col + offset).value
        if value not in (None, ""):
            return value
    return None


def parse_top_metadata(wb_formula, wb_values, path: Path, header_row: Optional[int]) -> dict[str, Any]:
    ws = wb_values.worksheets[0]
    max_row = min((header_row or 22) - 1, ws.max_row)
    description = ""
    item_code = infer_item_code(path.stem)
    customer_item_code = None
    dimensions: dict[str, Any] = {}
    container: dict[str, Any] = {}
    raw_header: dict[str, Any] = {}
    section = "overall"

    for r in range(1, max_row + 1):
        row_values = [clean_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 8) + 1)]
        row_text = " ".join(row_values)
        raw_header[str(r)] = row_values
        found_code = infer_item_code(row_text)
        if not item_code and found_code:
            item_code = found_code
        for c, value in enumerate(row_values, start=1):
            lower = value.lower()
            if "description" in lower:
                description = clean_text(adjacent_value(ws, r, c)) or clean_text(ws.cell(r, c + 1).value)
            if "cantoni item" in lower:
                customer_item_code = clean_text(adjacent_value(ws, r, c))
            if lower in {"container"}:
                section = "container"
            if "overall arrangement" in lower:
                section = "overall"
            if lower in {"height", "width", "depth", "length", "weight", "material"}:
                target = container if section == "container" or lower == "material" else dimensions
                target[lower] = clean_text(adjacent_value(ws, r, c))

    if not description:
        for r in range(1, max_row + 1):
            for c in range(1, min(ws.max_column, 8) + 1):
                value = clean_text(ws.cell(r, c).value)
                if value.lower().startswith("description"):
                    description = clean_text(ws.cell(r, c + 1).value)
                    break
            if description:
                break

    year_match = re.search(r"(20\d{2})", str(path))
    recipe_year = int(year_match.group(1)) if year_match else None
    build_type = infer_build_type(item_code, description, str(path))
    prefix = item_prefix(item_code)
    return {
        "item_code": item_code,
        "customer_item_code": customer_item_code,
        "product_family": prefix,
        "build_type": build_type,
        "description": description,
        "recipe_year": recipe_year,
        "dimensions": dimensions,
        "container_details": container,
        "raw_header": raw_header,
    }


def parse_pricing_summary(ws_values, ws_formula) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for row in range(1, ws_values.max_row + 1):
        for col in range(1, ws_values.max_column + 1):
            label = clean_text(ws_values.cell(row, col).value).lower()
            if label in {"total", "retail", "wholesale"}:
                value = None
                formula = None
                for offset in range(1, 4):
                    candidate = ws_values.cell(row, col + offset).value
                    candidate_formula = ws_formula.cell(row, col + offset).value
                    if candidate not in (None, ""):
                        value = safe_float(candidate)
                        formula = candidate_formula if isinstance(candidate_formula, str) and candidate_formula.startswith("=") else None
                        break
                summary[label] = {"value": value, "formula": formula, "cell": ws_values.cell(row, col).coordinate}
    return summary


def parse_components(wb_formula, wb_values, header_row: int, build_type: str) -> list[dict[str, Any]]:
    ws = wb_values.worksheets[0]
    wf = wb_formula.worksheets[0]
    header_values = [clean_text(ws.cell(header_row, c).value).lower() for c in range(1, ws.max_column + 1)]
    sub_values = [clean_text(ws.cell(header_row + 1, c).value).lower() for c in range(1, ws.max_column + 1)]

    def find_col(labels: list[str], fallback: Optional[int] = None) -> Optional[int]:
        for label in labels:
            for idx, value in enumerate(header_values, start=1):
                if value == label or label in value:
                    return idx
            for idx, value in enumerate(sub_values, start=1):
                if value == label or label in value:
                    return idx
        return fallback

    cols = {
        "product_sku": find_col(["product sku", "product"], 1),
        "vendor": find_col(["vendor"], 2),
        "description": find_col(["description"], 3),
        "first_cost": find_col(["first cost", "fc"], 4),
        "landed_cost": find_col(["landed cost", "lc"], 5),
        "quantity": find_col(["qty", "quantity"], 6),
        "retail": find_col(["retail", "prt"], 7),
        "extended_total": find_col(["ext", "pre-retail total"], 8),
    }
    rows: list[dict[str, Any]] = []
    for row in range(header_row + 2, ws.max_row + 1):
        label_candidates = [clean_text(ws.cell(row, c).value).lower() for c in range(1, min(ws.max_column, 10) + 1)]
        if any(value in {"total", "retail", "wholesale"} for value in label_candidates):
            continue
        data: dict[str, Any] = {}
        formulas: dict[str, str] = {}
        raw: dict[str, Any] = {}
        for key, col in cols.items():
            if not col:
                continue
            value = ws.cell(row, col).value
            formula_value = wf.cell(row, col).value
            raw[key] = clean_text(value)
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formulas[key] = formula_value
            if key in {"first_cost", "landed_cost", "quantity", "retail", "extended_total"}:
                data[key] = safe_float(value)
            else:
                data[key] = clean_text(value)
        if not any([data.get("product_sku"), data.get("vendor"), data.get("description")]):
            if not any(safe_float(data.get(key)) for key in ["first_cost", "landed_cost", "quantity", "retail", "extended_total"]):
                continue
        if (data.get("quantity") in (None, 0)) and not any([data.get("product_sku"), data.get("vendor"), data.get("description")]):
            continue
        data["line_order"] = len(rows)
        data["formulas"] = formulas
        data["raw_row"] = raw
        data["component_label"] = infer_component_label(data, build_type)
        rows.append(data)
    return rows


def parse_xlsx_recipe(path: Path) -> dict[str, Any]:
    import openpyxl

    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws = wb_values.worksheets[0]
    header_row = find_header_row(wb_formula.worksheets[0])
    if not header_row:
        raise ValueError("No recipe component header found")
    meta = parse_top_metadata(wb_formula, wb_values, path, header_row)
    components = parse_components(wb_formula, wb_values, header_row, meta["build_type"])
    if not components:
        raise ValueError("Recipe header found but no component rows extracted")
    pricing = parse_pricing_summary(ws, wb_formula.worksheets[0])
    meta["pricing_summary"] = pricing
    meta["components"] = components
    return meta


async def store_recipe(conn, source_file_id: int, parsed: dict[str, Any], root: Path, path: Path):
    recipe_id = await conn.fetchval("""
        INSERT INTO historical_recipes (
            source_file_id, item_code, customer_item_code, product_family, build_type,
            description, source_collection, recipe_year, dimensions, container_details,
            pricing_summary, raw_header, updated_at
        )
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9::jsonb, $10::jsonb, $11::jsonb, $12::jsonb, NOW())
        ON CONFLICT (source_file_id) DO UPDATE SET
            item_code = EXCLUDED.item_code,
            customer_item_code = EXCLUDED.customer_item_code,
            product_family = EXCLUDED.product_family,
            build_type = EXCLUDED.build_type,
            description = EXCLUDED.description,
            source_collection = EXCLUDED.source_collection,
            recipe_year = EXCLUDED.recipe_year,
            dimensions = EXCLUDED.dimensions,
            container_details = EXCLUDED.container_details,
            pricing_summary = EXCLUDED.pricing_summary,
            raw_header = EXCLUDED.raw_header,
            updated_at = NOW()
        RETURNING id
    """,
        source_file_id,
        parsed.get("item_code"),
        parsed.get("customer_item_code"),
        parsed.get("product_family"),
        parsed.get("build_type"),
        parsed.get("description"),
        str(path.relative_to(root).parts[0]) if path.is_relative_to(root) else None,
        parsed.get("recipe_year"),
        json_dumps(parsed.get("dimensions") or {}),
        json_dumps(parsed.get("container_details") or {}),
        json_dumps(parsed.get("pricing_summary") or {}),
        json_dumps(parsed.get("raw_header") or {}),
    )
    await conn.execute("DELETE FROM historical_recipe_components WHERE recipe_id = $1", recipe_id)
    for component in parsed.get("components") or []:
        await conn.execute("""
            INSERT INTO historical_recipe_components (
                recipe_id, line_order, component_label, vendor, supplier_sku, description,
                quantity, first_cost, landed_cost, retail, extended_total, formulas, raw_row
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12::jsonb, $13::jsonb)
        """,
            recipe_id,
            int(component.get("line_order") or 0),
            component.get("component_label"),
            component.get("vendor"),
            component.get("product_sku"),
            component.get("description"),
            component.get("quantity"),
            component.get("first_cost"),
            component.get("landed_cost"),
            component.get("retail"),
            component.get("extended_total"),
            json_dumps(component.get("formulas") or {}),
            json_dumps(component.get("raw_row") or {}),
        )
    await conn.execute("""
        UPDATE recipe_source_files
        SET status = 'parsed', item_code = $2, linked_item_code = $2, error_message = NULL,
            metadata = COALESCE(metadata, '{}'::jsonb) || $3::jsonb, updated_at = NOW()
        WHERE id = $1
    """, source_file_id, parsed.get("item_code"), json_dumps({"components": len(parsed.get("components") or [])}))


async def index_visual_asset(conn, source_file_id: int, path: Path):
    item_code = infer_item_code(path.stem)
    await conn.execute("""
        INSERT INTO visual_reference_assets (
            source_file_id, item_code, file_path, file_name, extension, asset_type, status, metadata, updated_at
        )
        VALUES ($1, $2, $3, $4, $5, $6, 'indexed', $7::jsonb, NOW())
        ON CONFLICT (source_file_id) DO UPDATE SET
            item_code = EXCLUDED.item_code,
            file_path = EXCLUDED.file_path,
            file_name = EXCLUDED.file_name,
            extension = EXCLUDED.extension,
            asset_type = EXCLUDED.asset_type,
            status = EXCLUDED.status,
            metadata = EXCLUDED.metadata,
            updated_at = NOW()
    """,
        source_file_id,
        item_code,
        str(path),
        path.name,
        path.suffix.lower(),
        "psd_source" if path.suffix.lower() == ".psd" else "image",
        json_dumps({"folder": str(path.parent)}),
    )
    await conn.execute(
        "UPDATE recipe_source_files SET status = 'asset_linked', linked_item_code = COALESCE($2, linked_item_code), updated_at = NOW() WHERE id = $1",
        source_file_id,
        item_code,
    )


async def refresh_derived_standards(conn):
    rows = await conn.fetch("""
        SELECT product_family AS prefix, COUNT(*)::int AS count,
               ARRAY_AGG(item_code ORDER BY item_code) FILTER (WHERE item_code IS NOT NULL) AS examples
        FROM historical_recipes
        WHERE product_family IS NOT NULL
        GROUP BY product_family
    """)
    for row in rows:
        prefix = row["prefix"]
        examples = list(row["examples"] or [])[:8]
        await conn.execute("""
            INSERT INTO sku_standards (prefix, label, description, inferred_count, examples, active, updated_at)
            VALUES ($1, $2, $3, $4, $5::jsonb, TRUE, NOW())
            ON CONFLICT (prefix) DO UPDATE SET
                inferred_count = EXCLUDED.inferred_count,
                examples = EXCLUDED.examples,
                label = COALESCE(NULLIF(sku_standards.label, ''), EXCLUDED.label),
                description = COALESCE(sku_standards.description, EXCLUDED.description),
                updated_at = NOW()
        """,
            prefix,
            KNOWN_PREFIX_LABELS.get(prefix, prefix),
            f"Inferred from {row['count']} historical recipe file(s).",
            row["count"],
            json_dumps(examples),
        )
    await conn.execute("""
        INSERT INTO recipe_pricing_rules (scope, project_id, rules, updated_at)
        VALUES ('global', 0, $1::jsonb, NOW())
        ON CONFLICT (scope, project_id) DO NOTHING
    """, json_dumps(DEFAULT_PRICING_RULES))


async def classify_or_parse_file(conn, path: Path, root: Path) -> dict[str, Any]:
    source_id = await upsert_source_file(conn, path, root)
    source_row = await conn.fetchrow("SELECT status FROM recipe_source_files WHERE id = $1", source_id)
    if source_row and source_row["status"] == "duplicate":
        return {"path": str(path), "status": "duplicate"}
    ext = path.suffix.lower()
    try:
        if ext == ".xlsx":
            parsed = parse_xlsx_recipe(path)
            await store_recipe(conn, source_id, parsed, root, path)
            return {"path": str(path), "status": "parsed", "item_code": parsed.get("item_code")}
        if ext == ".docx":
            parsed = parse_pricing_doc(path)
            if parsed["rules"]:
                await conn.execute("""
                    INSERT INTO recipe_pricing_rules (scope, project_id, rules, updated_at)
                    VALUES ($1, 0, $2::jsonb, NOW())
                    ON CONFLICT (scope, project_id) DO UPDATE SET rules = EXCLUDED.rules, updated_at = NOW()
                """, f"source:{path.name}", json_dumps(parsed["rules"]))
            await conn.execute(
                "UPDATE recipe_source_files SET status = 'reference_only', metadata = COALESCE(metadata, '{}'::jsonb) || $2::jsonb, updated_at = NOW() WHERE id = $1",
                source_id,
                json_dumps(parsed),
            )
            return {"path": str(path), "status": "reference_only"}
        if ext in {".png", ".jpg", ".jpeg", ".psd"}:
            await index_visual_asset(conn, source_id, path)
            return {"path": str(path), "status": "asset_linked"}
        if ext in {".xls", ".pdf", ".pptx", ".zip", ".csv"}:
            await conn.execute(
                "UPDATE recipe_source_files SET status = 'unsupported_deferred', error_message = $2, updated_at = NOW() WHERE id = $1",
                source_id,
                f"{ext} registered for later parser support",
            )
            return {"path": str(path), "status": "unsupported_deferred"}
        await conn.execute(
            "UPDATE recipe_source_files SET status = 'unsupported_deferred', error_message = $2, updated_at = NOW() WHERE id = $1",
            source_id,
            "Unsupported file type",
        )
        return {"path": str(path), "status": "unsupported_deferred"}
    except Exception as exc:
        await conn.execute(
            "UPDATE recipe_source_files SET status = 'failed_needs_review', error_message = $2, updated_at = NOW() WHERE id = $1",
            source_id,
            str(exc)[:1000],
        )
        return {"path": str(path), "status": "failed_needs_review", "error": str(exc)}


def files_to_scan(root: Path) -> list[Path]:
    if not root.exists():
        raise HTTPException(status_code=404, detail=f"Source folder not found: {root}")
    return sorted([path for path in root.rglob("*") if path.is_file() and not path.name.startswith("~$")])


async def import_pending_files(conn, root: Path, limit: int, include_assets: bool, reset_failed: bool):
    await ensure_schema(conn)
    paths = files_to_scan(root)
    for path in paths:
        kind = file_kind(path)
        if not include_assets and kind == "image_asset":
            continue
        await upsert_source_file(conn, path, root)
    if reset_failed:
        await conn.execute("UPDATE recipe_source_files SET status = 'pending', error_message = NULL WHERE status = 'failed_needs_review'")
    pending = await conn.fetch("""
        SELECT source_path
        FROM recipe_source_files
        WHERE status = 'pending'
        ORDER BY
            CASE WHEN file_name = ANY($2::text[]) THEN 0 ELSE 1 END,
            CASE extension
                WHEN '.xlsx' THEN 1
                WHEN '.docx' THEN 2
                WHEN '.png' THEN 3
                WHEN '.jpg' THEN 4
                WHEN '.jpeg' THEN 5
                WHEN '.psd' THEN 6
                ELSE 9
            END,
            source_path
        LIMIT $1
    """, max(1, min(limit, 250)), sorted(KNOWN_EXAMPLE_FILES))
    results = []
    for row in pending:
        results.append(await classify_or_parse_file(conn, Path(row["source_path"]), root))
    await refresh_derived_standards(conn)
    return results


async def summary_payload(conn) -> dict[str, Any]:
    await ensure_schema(conn)
    status_rows = await conn.fetch("SELECT status, COUNT(*)::int AS count FROM recipe_source_files GROUP BY status ORDER BY status")
    ext_rows = await conn.fetch("SELECT extension, COUNT(*)::int AS count FROM recipe_source_files GROUP BY extension ORDER BY count DESC")
    recipe_count = await conn.fetchval("SELECT COUNT(*) FROM historical_recipes")
    component_count = await conn.fetchval("SELECT COUNT(*) FROM historical_recipe_components")
    asset_count = await conn.fetchval("SELECT COUNT(*) FROM visual_reference_assets")
    failures = await conn.fetch("""
        SELECT relative_path, status, error_message
        FROM recipe_source_files
        WHERE status IN ('failed_needs_review', 'unsupported_deferred')
        ORDER BY updated_at DESC
        LIMIT 25
    """)
    return {
        "source_root": str(SOURCE_ROOT),
        "statuses": [dict(row) for row in status_rows],
        "extensions": [dict(row) for row in ext_rows],
        "recipe_count": int(recipe_count or 0),
        "component_count": int(component_count or 0),
        "asset_count": int(asset_count or 0),
        "failures": [dict(row) for row in failures],
        "parser_version": PARSER_VERSION,
    }


@router.post("/recipe-intelligence/import")
async def import_recipe_intelligence(body: ImportRequest):
    root = Path(body.root_path) if body.root_path else SOURCE_ROOT
    conn = await get_conn()
    try:
        try:
            results = await import_pending_files(conn, root, body.limit, body.include_assets, body.reset_failed)
            return {"processed": len(results), "results": results, "summary": await summary_payload(conn)}
        except asyncpg.InsufficientPrivilegeError:
            results, store = local_import_pending_files(root, body.limit, body.include_assets, body.reset_failed)
            return {"processed": len(results), "results": results, "summary": local_summary_payload(store)}
    finally:
        await conn.close()


@router.get("/recipe-intelligence/import-status")
async def get_recipe_import_status():
    conn = await get_conn()
    try:
        try:
            return await summary_payload(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_summary_payload()
    finally:
        await conn.close()


@router.get("/recipe-intelligence/build-types")
async def get_build_types():
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_build_types_payload()
        rows = await conn.fetch("""
            SELECT build_type, COUNT(*)::int AS evidence_count,
                   ARRAY_AGG(DISTINCT product_family) FILTER (WHERE product_family IS NOT NULL) AS prefixes
            FROM historical_recipes
            WHERE build_type IS NOT NULL
            GROUP BY build_type
            ORDER BY evidence_count DESC, build_type
        """)
        return [
            {
                "label": row["build_type"],
                "evidence_count": row["evidence_count"],
                "prefixes": list(row["prefixes"] or []),
            }
            for row in rows
        ]
    finally:
        await conn.close()


@router.post("/recipe-intelligence/suggest")
async def suggest_recipe_components(body: SuggestRequest):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_suggest_payload(body)
        build_type = body.build_type.strip() or "Custom Arrangement"
        aliases = [alias.lower() for alias in build_type_aliases(build_type)]
        rows = await conn.fetch("""
            SELECT c.component_label,
                   COUNT(*)::int AS evidence_count,
                   AVG(NULLIF(c.quantity, 0))::float AS avg_quantity,
                   PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY NULLIF(c.quantity, 0))::float AS median_quantity,
                   AVG(NULLIF(c.extended_total, 0))::float AS avg_extended_total,
                   ARRAY_AGG(DISTINCT c.vendor) FILTER (WHERE c.vendor IS NOT NULL AND c.vendor != '') AS vendors,
                   ARRAY_AGG(DISTINCT c.description) FILTER (WHERE c.description IS NOT NULL AND c.description != '') AS descriptions
            FROM historical_recipe_components c
            JOIN historical_recipes r ON r.id = c.recipe_id
            WHERE LOWER(r.build_type) = ANY($1::text[])
            GROUP BY c.component_label
            ORDER BY evidence_count DESC, c.component_label
            LIMIT 8
        """, aliases)
        if not rows:
            fallback = [
                {"component_label": label, "evidence_count": 0, "avg_quantity": 1, "median_quantity": 1, "avg_extended_total": None, "vendors": [], "descriptions": []}
                for label in ["Container", "Foliage / Greenery", "Foam", "Top Dressing"]
            ]
            rows = fallback
        recipe_count = await conn.fetchval("SELECT COUNT(*) FROM historical_recipes WHERE LOWER(build_type) = ANY($1::text[])", aliases)
        pricing = await conn.fetchrow("""
            SELECT
                AVG((pricing_summary->'total'->>'value')::numeric)::float AS avg_total,
                MIN((pricing_summary->'total'->>'value')::numeric)::float AS min_total,
                MAX((pricing_summary->'total'->>'value')::numeric)::float AS max_total
            FROM historical_recipes
            WHERE LOWER(build_type) = ANY($1::text[])
              AND pricing_summary ? 'total'
        """, aliases)
        components = []
        for row in rows:
            data = dict(row)
            descriptions = [value for value in list(data.get("descriptions") or []) if value][:6]
            vendors = [value for value in list(data.get("vendors") or []) if value][:6]
            component_label = data.get("component_label") or "Product"
            terms = [component_label, *vendors[:3], *descriptions[:3]]
            components.append({
                "label": component_label,
                "suggested_quantity": round(float(data.get("median_quantity") or data.get("avg_quantity") or 1), 2),
                "average_quantity": round(float(data.get("avg_quantity") or 1), 2),
                "evidence_count": int(data.get("evidence_count") or 0),
                "average_extended_total": data.get("avg_extended_total"),
                "vendors": vendors,
                "examples": descriptions,
                "search_terms": [clean_text(term) for term in terms if clean_text(term)][:8],
            })
        evidence = int(recipe_count or 0)
        return {
            "build_type": build_type,
            "input": body.model_dump(),
            "evidence_count": evidence,
            "confidence": "high" if evidence >= 10 else "medium" if evidence >= 3 else "starter",
            "components": components,
            "cost_range": dict(pricing) if pricing else {"avg_total": None, "min_total": None, "max_total": None},
        }
    finally:
        await conn.close()


@router.get("/pricing-rules")
async def get_pricing_rules(project_id: Optional[int] = None):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
            await refresh_derived_standards(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_pricing_rules_payload(project_id)
        global_row = await conn.fetchrow("SELECT rules FROM recipe_pricing_rules WHERE scope = 'global' AND project_id = 0")
        project_row = None
        if project_id:
            project_row = await conn.fetchrow("SELECT rules FROM recipe_pricing_rules WHERE scope = 'project' AND project_id = $1", project_id)
        source_rows = await conn.fetch("SELECT scope, rules, updated_at FROM recipe_pricing_rules WHERE scope LIKE 'source:%' ORDER BY scope")
        return {
            "global_rules": global_row["rules"] if global_row else {},
            "project_rules": project_row["rules"] if project_row else None,
            "source_rules": [dict(row) for row in source_rows],
        }
    finally:
        await conn.close()


@router.put("/pricing-rules")
async def update_pricing_rules(body: PricingRuleUpdate):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_update_pricing_rules(body)
        scope = "project" if body.project_id else "global"
        await conn.execute("""
            INSERT INTO recipe_pricing_rules (scope, project_id, rules, updated_at)
            VALUES ($1, $2, $3::jsonb, NOW())
            ON CONFLICT (scope, project_id) DO UPDATE SET rules = EXCLUDED.rules, updated_at = NOW()
        """, scope, body.project_id or 0, json_dumps(body.rules))
        return {"ok": True}
    finally:
        await conn.close()


@router.get("/sku-standards")
async def get_sku_standards():
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
            await refresh_derived_standards(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_sku_standards_payload()
        rows = await conn.fetch("SELECT * FROM sku_standards ORDER BY inferred_count DESC, prefix")
        return [dict(row) for row in rows]
    finally:
        await conn.close()


@router.put("/sku-standards")
async def update_sku_standard(body: SkuStandardUpdate):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_update_sku_standard(body)
        prefix = normalize_code(body.prefix)
        await conn.execute("""
            INSERT INTO sku_standards (prefix, label, description, inferred_count, examples, active, updated_at)
            VALUES ($1, $2, $3, 0, '[]'::jsonb, $4, NOW())
            ON CONFLICT (prefix) DO UPDATE SET
                label = EXCLUDED.label,
                description = EXCLUDED.description,
                active = EXCLUDED.active,
                updated_at = NOW()
        """, prefix, body.label, body.description, body.active)
        return {"ok": True}
    finally:
        await conn.close()


@router.get("/recipe-intelligence/visual-references")
async def get_visual_references(item_code: Optional[str] = None, limit: int = 50):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            return local_visual_references_payload(item_code, limit)
        if item_code:
            rows = await conn.fetch("""
                SELECT *
                FROM visual_reference_assets
                WHERE item_code = $1
                ORDER BY file_name
                LIMIT $2
            """, normalize_code(item_code), min(limit, 200))
        else:
            rows = await conn.fetch("SELECT * FROM visual_reference_assets ORDER BY updated_at DESC LIMIT $1", min(limit, 200))
        return [dict(row) for row in rows]
    finally:
        await conn.close()


@router.get("/recipe-intelligence/visual-references/{asset_id}/preview")
async def get_visual_reference_preview(asset_id: int):
    conn = await get_conn()
    try:
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            row = local_visual_reference_by_id(asset_id)
            if not row:
                raise HTTPException(status_code=404, detail="Visual reference not found")
            return preview_file_response(row)
        row = await conn.fetchrow("SELECT * FROM visual_reference_assets WHERE id = $1", asset_id)
        if not row:
            raise HTTPException(status_code=404, detail="Visual reference not found")
        return preview_file_response(dict(row))
    finally:
        await conn.close()


@router.post("/historical/complete")
async def complete_historical_build(body: HistoricalCompleteRequest):
    if body.completion_status != "approved_paid_purchased":
        raise HTTPException(status_code=400, detail="Historical learning only accepts approved, paid, purchased builds")
    conn = await get_conn()
    try:
        sql_history_enabled = True
        try:
            await ensure_schema(conn)
        except asyncpg.InsufficientPrivilegeError:
            sql_history_enabled = False
        arr = await conn.fetchrow("SELECT id, name, client_name, total_cost, total_with_markup FROM arrangements WHERE id = $1", body.arrangement_id)
        if not arr:
            raise HTTPException(status_code=404, detail="Project not found")
        container_filter = "AND ac.id = $2" if body.container_id else ""
        params = [body.arrangement_id] + ([body.container_id] if body.container_id else [])
        item_columns = {
            row["column_name"]
            for row in await conn.fetch("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'container_items'
            """)
        }
        supports_status = "status" in item_columns
        supports_parts = {"part_key", "part_label", "part_order"}.issubset(item_columns)
        native_status = "ci.status" if supports_status else "NULL::text"
        native_part_key = "ci.part_key" if supports_parts else "NULL::text"
        native_part_label = "ci.part_label" if supports_parts else "NULL::text"
        native_part_order = "ci.part_order" if supports_parts else "NULL::integer"
        supports_meta = False
        try:
            supports_meta = bool(await conn.fetchval("SELECT to_regclass('public.container_item_meta') IS NOT NULL"))
            if supports_meta:
                await conn.fetchval("SELECT 1 FROM container_item_meta LIMIT 1")
        except Exception:
            supports_meta = False
        meta_join = (
            "LEFT JOIN container_item_meta cim ON cim.item_id = ci.id"
            if supports_meta
            else "LEFT JOIN (SELECT NULL::integer AS item_id, NULL::text AS status, NULL::text AS part_key, NULL::text AS part_label, NULL::integer AS part_order) cim ON false"
        )
        rows = await conn.fetch(f"""
            SELECT ac.id AS container_id, ac.label, ac.bucket_type, ac.requested_quantity, ac.scope_notes,
                   ci.id AS item_id, ci.quantity,
                   COALESCE(cim.status, {native_status}, 'selected')::text AS status,
                   COALESCE(cim.part_key, {native_part_key})::text AS part_key,
                   COALESCE(cim.part_label, {native_part_label})::text AS part_label,
                   COALESCE(cim.part_order, {native_part_order}, 0)::integer AS part_order,
                   p.id AS product_id, p.name AS product_name, p.supplier_sku, p.current_price,
                   p.category, p.unit,
                   COALESCE(p.photo_url, p.image_urls[1], p.raw_data->>'source_photo_url') AS photo_url,
                   s.name AS supplier_name
            FROM arrangement_containers ac
            LEFT JOIN container_items ci ON ci.container_id = ac.id
            {meta_join}
            LEFT JOIN products p ON p.id = ci.product_id
            LEFT JOIN suppliers s ON s.id = p.supplier_id
            WHERE ac.arrangement_id = $1 {container_filter}
            ORDER BY ac.id, ci.id
        """, *params)
        if body.container_id and not rows:
            raise HTTPException(status_code=404, detail="Scope not found")
        container = dict(rows[0]) if rows else {}
        components = []
        for row in rows:
            if row["item_id"] is None:
                continue
            data = dict(row)
            sidecar = {} if supports_meta else load_item_meta_sidecar(row["item_id"])
            data["status"] = sidecar.get("status") or data.get("status") or "selected"
            data["part_key"] = sidecar.get("part_key") or data.get("part_key")
            data["part_label"] = sidecar.get("part_label") or data.get("part_label")
            data["part_order"] = int((sidecar.get("part_order") if sidecar else data.get("part_order")) or 0)
            if data["status"] != "selected":
                continue
            data["line_total"] = float(data["current_price"] or 0) * float(data["quantity"] or 1)
            components.append(data)
        pricing_snapshot = {
            "project_total_cost": float(arr["total_cost"] or 0),
            "project_total_with_markup": float(arr["total_with_markup"] or 0),
            "selected_component_total": sum(float(row.get("line_total") or 0) for row in components),
        }
        finished_sku = normalize_code(body.finished_sku)
        dimensions = dimensions_from_scope_notes(container.get("scope_notes"))
        build_type = container.get("bucket_type") or arr["name"]
        if sql_history_enabled:
            await conn.execute("""
                INSERT INTO completed_historical_builds (
                    arrangement_id, container_id, finished_sku, build_type, dimensions,
                    component_snapshot, pricing_snapshot, completion_status, notes, completed_at
                )
                VALUES ($1, $2, $3, $4, $5::jsonb, $6::jsonb, $7::jsonb, $8, $9, NOW())
            """,
                body.arrangement_id,
                body.container_id,
                finished_sku,
                build_type,
                json_dumps(dimensions),
                json_dumps(components),
                json_dumps(pricing_snapshot),
                body.completion_status,
                body.notes,
            )
        else:
            store = load_local_store()
            store["completed_builds"].append({
                "arrangement_id": body.arrangement_id,
                "container_id": body.container_id,
                "finished_sku": finished_sku,
                "build_type": build_type,
                "dimensions": dimensions,
                "component_snapshot": components,
                "pricing_snapshot": pricing_snapshot,
                "completion_status": body.completion_status,
                "notes": body.notes,
                "completed_at": datetime.utcnow().isoformat(),
            })
            save_local_store(store)
        return {"ok": True, "components_added": len(components)}
    finally:
        await conn.close()
