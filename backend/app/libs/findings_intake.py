"""Rich intake normalizer for the standardized supplier catalog exports
("THE FINDINGS" / catalog-extraction outputs).

Goal: LOSSLESS capture + rich promotion. Every non-empty source column is
preserved in ``raw_data``; a well-known core is promoted to typed product
columns; all product images are collapsed into a single ordered list.

This module is intentionally supplier-agnostic: it maps by normalized header
name, so the same code handles every supplier's Excel. Unknown columns are
never dropped — they land in ``raw_data``.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from .catalog_importer import normalize_category, normalize_unit, parse_price, safe_int
from .taxonomy import color_families, color_family, product_type_family


def _norm_key(key: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(key).lower()).strip()


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def _num(value: Any) -> Optional[float]:
    s = _clean(value)
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s.replace(",", ""))
    return float(m.group(0)) if m else None


def _pos(v: Optional[float]) -> Optional[float]:
    """Only keep strictly-positive measurements (suppliers use 0 as 'unknown')."""
    return v if (v is not None and v > 0) else None


_DIM_RE = re.compile(r"([LWHDlwhd])\s*[:=]?\s*(\d+(?:\.\d+)?)")
_DIM_MAP = {"l": "length", "w": "width", "h": "height", "d": "diameter"}


def _parse_dim_text(text: Optional[str]) -> dict[str, float]:
    """'L36\" W18\" H15\"' -> {length:36, width:18, height:15}."""
    out: dict[str, float] = {}
    if not text:
        return out
    for letter, val in _DIM_RE.findall(text):
        out.setdefault(_DIM_MAP[letter.lower()], float(val))
    return out


_NON_PRODUCT_RE = re.compile(
    r"\b(restocking fee|freight|shipping|surcharge|handling|deposit|catalog|sample kit|gift card)\b",
    re.I,
)


def _looks_non_product(p: "ProductIntake") -> list[str]:
    """Heuristics for rows that are fees/covers/placeholders, not real products."""
    reasons: list[str] = []
    if p.photo_url and re.search(r"brokenimage|placeholder|noimage|no_image", p.photo_url, re.I):
        reasons.append("placeholder_image")
    if p.name and _NON_PRODUCT_RE.search(p.name):
        reasons.append("non_product_line")
    if p.case_qty and p.case_qty > 100000:
        reasons.append("implausible_case_qty")
    return reasons


# canonical typed field -> ordered list of normalized header aliases
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "supplier_sku": ("supplier sku", "sku", "item no", "itemno", "item number", "item", "base sku"),
    "name": ("product name", "web name", "name", "title", "product title"),
    "description": ("description", "desc", "long description"),
    "current_price": ("price", "current price", "account price", "dealer price", "unit price"),
    "category": ("category", "department"),
    "subcategory": ("subcategory", "sub category"),
    "uom": ("uom", "unit of measure"),
    "moq": ("moq", "min qty", "minimum qty", "minimum order", "order minimum"),
    "case_qty": ("case quantity", "case qty", "case pack", "case"),
    "availability": ("availability", "qty in stock", "units in stock", "stock", "in stock qty"),
    "weight_lb": ("weight lbs", "weight lb", "weight"),
    "height_in": ("height in", "height"),
    "width_in": ("width in", "width"),
    "length_in": ("length in", "length"),
    "diameter_in": ("diameter in", "diameter"),
    "material": ("material", "primary material"),
    "color": ("color", "colour", "primary color"),
    "finish": ("finish", "finish style"),
    "style": ("style", "product type"),
    "country_of_origin": ("country of origin", "country"),
    "currency": ("currency",),
    "supplier_product_id": ("product id", "variation id", "item group id", "klevu id"),
}

# columns whose values are image URLs (besides numbered image_url_N and additional_image_urls)
PRIMARY_IMAGE_KEYS = ("image url", "image", "photo url", "main image", "primary image")
EXTRA_IMAGE_KEYS = ("additional image urls", "additional images", "gallery images", "gallery images json")


@dataclass
class ProductIntake:
    supplier_sku: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    category: str = "other"
    subcategory: Optional[str] = None
    unit: str = "each"
    uom: Optional[str] = None
    current_price: Optional[float] = None
    currency: Optional[str] = None
    moq: Optional[int] = None
    case_qty: Optional[int] = None
    availability: Optional[str] = None
    availability_note: Optional[str] = None
    height_in: Optional[float] = None
    width_in: Optional[float] = None
    length_in: Optional[float] = None
    diameter_in: Optional[float] = None
    weight_lb: Optional[float] = None
    material: Optional[str] = None
    color: Optional[str] = None
    finish: Optional[str] = None
    style: Optional[str] = None
    country_of_origin: Optional[str] = None
    supplier_product_id: Optional[str] = None
    photo_url: Optional[str] = None
    image_urls: list[str] = field(default_factory=list)
    tags: list[str] = field(default_factory=list)
    needs_review: Optional[str] = None
    raw_data: dict[str, Any] = field(default_factory=dict)


def _split_urls(value: Optional[str]) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[;\n,]+", value)
    return [p.strip() for p in parts if p.strip().lower().startswith("http")]


def _collect_images(row: dict[str, Any], norm: dict[str, Any]) -> list[str]:
    urls: list[str] = []

    def add(u: Optional[str]) -> None:
        u = _clean(u)
        if u and u.lower().startswith("http") and u not in urls:
            urls.append(u)

    for k in PRIMARY_IMAGE_KEYS:
        if k in norm:
            add(norm[k])
    for k in EXTRA_IMAGE_KEYS:
        if k in norm:
            for u in _split_urls(_clean(norm[k])):
                add(u)
    # numbered columns: image_url_2 ... image_url_10 (any order)
    numbered = sorted(
        (nk for nk in norm if re.fullmatch(r"image url \d+", nk)),
        key=lambda nk: int(nk.rsplit(" ", 1)[1]),
    )
    for nk in numbered:
        add(norm[nk])
    return urls


def normalize_row(row: dict[str, Any]) -> ProductIntake:
    """Map one source row -> ProductIntake. Lossless: every non-empty source
    column is retained in raw_data."""
    norm = {_norm_key(k): v for k, v in row.items()}

    def pick(field_name: str) -> Optional[str]:
        for alias in FIELD_ALIASES.get(field_name, ()):
            if alias in norm:
                v = _clean(norm[alias])
                if v:
                    return v
        return None

    p = ProductIntake()
    p.supplier_sku = pick("supplier_sku")
    p.name = pick("name")
    p.description = pick("description")
    cat_raw = pick("category")
    p.category = normalize_category(cat_raw)
    p.subcategory = pick("subcategory") or cat_raw
    p.uom = pick("uom")
    p.unit = normalize_unit(p.uom or pick("style"))
    p.current_price = parse_price(pick("current_price") or "")
    p.currency = pick("currency")
    p.moq = safe_int(pick("moq"))
    p.case_qty = safe_int(pick("case_qty"))
    avail = pick("availability")
    p.availability = avail
    if avail and avail.isdigit():
        p.availability_note = f"In stock: {avail}"
    p.weight_lb = _num(pick("weight_lb"))
    # Prefer the numeric columns when they carry a real value; otherwise parse
    # the free-text dimensions column (e.g. 'L36" W18" H15"').
    dims = _parse_dim_text(_clean(norm.get("dimensions in")) or _clean(norm.get("dimensions")))
    p.height_in = _pos(_num(pick("height_in"))) or dims.get("height")
    p.width_in = _pos(_num(pick("width_in"))) or dims.get("width")
    p.length_in = _pos(_num(pick("length_in"))) or dims.get("length")
    p.diameter_in = _pos(_num(pick("diameter_in"))) or dims.get("diameter")
    p.material = pick("material")
    p.color = pick("color")
    p.finish = pick("finish")
    p.style = pick("style")
    p.country_of_origin = pick("country_of_origin")
    p.supplier_product_id = pick("supplier_product_id")

    p.image_urls = _collect_images(row, norm)
    p.photo_url = p.image_urls[0] if p.image_urls else None

    review = _clean(norm.get("needs review")) or ""
    reasons = [r for r in [review] if r] + _looks_non_product(p)
    p.needs_review = "; ".join(dict.fromkeys(reasons)) or None

    # LOSSLESS bin: keep every non-empty original column, plus fields the
    # existing commit path / product view expect. (products has no needs_review
    # column, so the review flag lives here.)
    raw = {str(k): _clean(v) for k, v in row.items() if _clean(v) is not None}
    if p.image_urls:
        raw["image_urls"] = p.image_urls
        raw["source_photo_url"] = p.photo_url
    if p.needs_review:
        raw["needs_review"] = p.needs_review
        raw["needs_review_flag"] = True
    # normalized filter families (clean, grouped dropdown values). Colors are a
    # LIST so combos surface under each color + Multi-color.
    cfs = color_families(p.color)
    tf = product_type_family(p.style)
    if cfs:
        raw["color_families"] = cfs
        raw["color_family"] = cfs[0]
    if tf:
        raw["type_family"] = tf
    p.raw_data = raw
    return p


def parse_findings_xlsx(path: str | Path) -> tuple[list[ProductIntake], dict[str, Any]]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(next(it))]
    rows: list[ProductIntake] = []
    dropped_no_sku = 0
    for values in it:
        row = {headers[i]: (values[i] if i < len(values) else None) for i in range(len(headers))}
        p = normalize_row(row)
        if not p.supplier_sku or not p.name:
            dropped_no_sku += 1
            continue
        rows.append(p)
    wb.close()
    report = {
        "source_headers": headers,
        "rows_in": len(rows) + dropped_no_sku,
        "rows_ok": len(rows),
        "dropped_missing_sku_or_name": dropped_no_sku,
    }
    return rows, report
