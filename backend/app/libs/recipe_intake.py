"""Parse TBDG pricing / production worksheets into historical recipes.

These are hand-built Excel sheets spanning 2020-2025 in several layout families
(identified by cell ``A1``). Each family gets an adapter; everything funnels into
one :class:`ParsedRecipe` shape that maps 1:1 onto ``historical_recipes`` +
``historical_recipe_components``.

Two standing rules drive the design:

1. **Import values, not formulas.** Workbooks are read twice -- ``data_only=True``
   for the computed values that the business actually charged, and
   ``data_only=False`` for the formula text. The values are the record; the
   formulas are kept alongside them in ``ParsedComponent.formulas`` so a
   deviation is *visible* rather than silently "corrected". In the canonical
   fiddle-leaf sheet, for example, the Draggon Wood row multiplies FC by the AR
   factor instead of the landed-cost factor, and the container row applies the
   profit factor twice. Both are imported exactly as the sheet computed them and
   flagged as anomalies.

2. **Lossless capture.** Every non-empty cell of a component row is preserved in
   ``raw_row``; header labels, notes and text-only lines that carry no money go
   into ``raw_header`` so nothing in the sheet is thrown away.

Nothing here hardcodes row numbers: the component block is located by finding the
header row (``PRODUCT SKU`` / ``VENDOR`` / ``DESCRIPTION`` ...) and reading until
the block terminates, because row counts vary between files.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook

from app.libs import taxonomy

PARSER_VERSION = "recipe_intake/1.0"

# ── Format families (value of cell A1) ────────────────────────────────────────

FORMAT_PRODUCTION_2025 = "tbdg_production_2025"
FORMAT_PRICING_2025 = "tbdg_pricing_2025"
FORMAT_CREATIVE_BRANCH = "creative_branch_recipe"
FORMAT_PRICE_SHEET = "cantoni_price_sheet"
FORMAT_UNKNOWN = "unknown"

SUPPORTED_FORMATS = (
    FORMAT_PRODUCTION_2025,
    FORMAT_PRICING_2025,
    FORMAT_CREATIVE_BRANCH,
)

FORMAT_LABELS = {
    FORMAT_PRODUCTION_2025: "TBDG PRODUCTION PRICE WORKSHEET 2025",
    FORMAT_PRICING_2025: "TBDG PRICING WORKSHEET 2025",
    FORMAT_CREATIVE_BRANCH: "Creative Branch Recipe (2020-2023)",
    FORMAT_PRICE_SHEET: "PRICE SHEET CANTONI 2023 (price list, not a recipe)",
    FORMAT_UNKNOWN: "unrecognised layout",
}


class RecipeParseError(Exception):
    """Raised when a workbook cannot be turned into a recipe."""


# ── Parsed shapes ─────────────────────────────────────────────────────────────


@dataclass
class ParsedComponent:
    line_order: int
    component_label: Optional[str] = None
    vendor: Optional[str] = None
    supplier_sku: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = None
    first_cost: Optional[float] = None
    landed_cost: Optional[float] = None
    retail: Optional[float] = None
    extended_total: Optional[float] = None
    formulas: dict[str, Any] = field(default_factory=dict)
    raw_row: dict[str, Any] = field(default_factory=dict)

    @property
    def anomalies(self) -> list[str]:
        return list(self.formulas.get("anomalies") or [])


@dataclass
class ParsedRecipe:
    format_family: str
    item_code: Optional[str] = None
    customer_item_code: Optional[str] = None
    product_family: Optional[str] = None
    build_type: Optional[str] = None
    description: Optional[str] = None
    source_collection: Optional[str] = None
    recipe_year: Optional[int] = None
    dimensions: dict[str, Any] = field(default_factory=dict)
    container_details: dict[str, Any] = field(default_factory=dict)
    pricing_summary: dict[str, Any] = field(default_factory=dict)
    raw_header: dict[str, Any] = field(default_factory=dict)
    visual_reference_count: int = 0
    components: list[ParsedComponent] = field(default_factory=list)

    @property
    def anomaly_count(self) -> int:
        return sum(len(c.anomalies) for c in self.components)


# ── Small helpers ─────────────────────────────────────────────────────────────


def sha256_file(path: str | Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        while block := fh.read(chunk):
            h.update(block)
    return h.hexdigest()


def _norm(value: Any) -> str:
    """Lower-cased, whitespace-collapsed text for label matching."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _number(value: Any) -> Optional[float]:
    """Coerce to float, tolerating '$1,234.50' and '48"'. Non-numeric -> None."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _col_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _close(a: Optional[float], b: Optional[float], tol: float = 0.02) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= max(tol, abs(b) * 0.001)


# ── Sheet grid ────────────────────────────────────────────────────────────────

_MAX_SCAN_COLS = 14
_MAX_SCAN_ROWS = 400


class _Grid:
    """1-indexed value/formula view of a worksheet, bounded to real content.

    openpyxl reports ``max_row`` of 1000+ for these sheets because of stray
    formatting, so the real content bound is found by scanning backwards.
    """

    def __init__(self, path: Path):
        wb_v = load_workbook(path, data_only=True)
        wb_f = load_workbook(path, data_only=False)
        self.sheet_names = list(wb_v.sheetnames)
        ws_v, ws_f = wb_v.active, wb_f.active
        self.title = ws_v.title
        self.image_count = len(getattr(ws_v, "_images", []) or [])

        self.max_col = min(ws_v.max_column or 1, _MAX_SCAN_COLS)
        scan_rows = min(ws_v.max_row or 1, _MAX_SCAN_ROWS)
        self.values: dict[tuple[int, int], Any] = {}
        self.formulas: dict[tuple[int, int], Any] = {}
        last = 0
        for r in range(1, scan_rows + 1):
            for c in range(1, self.max_col + 1):
                v = ws_v.cell(r, c).value
                if v is not None:
                    self.values[(r, c)] = v
                    last = r
                f = ws_f.cell(r, c).value
                if f is not None:
                    self.formulas[(r, c)] = f
        self.max_row = last
        wb_v.close()
        wb_f.close()

    def v(self, row: int, col: int) -> Any:
        return self.values.get((row, col))

    def f(self, row: int, col: int) -> Any:
        return self.formulas.get((row, col))

    def formula(self, row: int, col: int) -> Optional[str]:
        """The formula text for a cell, or None if it holds a literal."""
        raw = self.formulas.get((row, col))
        if isinstance(raw, str) and raw.startswith("="):
            return raw
        return None

    def row_cells(self, row: int) -> dict[str, Any]:
        """Every non-empty cell of a row, keyed by column letter (lossless)."""
        out: dict[str, Any] = {}
        for c in range(1, self.max_col + 1):
            v = self.v(row, c)
            if v is not None:
                out[_col_letter(c)] = v if isinstance(v, (int, float, str)) else str(v)
        return out

    def row_is_blank(self, row: int) -> bool:
        return all(self.v(row, c) is None for c in range(1, self.max_col + 1))

    def find_label(self, patterns: Iterable[str], max_row: Optional[int] = None,
                   max_col: Optional[int] = None, min_row: int = 1) -> Optional[tuple[int, int]]:
        """First cell whose normalised text matches any regex in *patterns*.

        ``min_row`` matters: "PRE-RETAIL TOTAL" and "Retail" are both a component
        column header *and* a totals label in these sheets, so a totals lookup has
        to start below the component block.
        """
        compiled = [re.compile(p) for p in patterns]
        for r in range(max(min_row, 1), (max_row or self.max_row) + 1):
            for c in range(1, (max_col or self.max_col) + 1):
                text = _norm(self.v(r, c))
                if text and any(p.match(text) for p in compiled):
                    return (r, c)
        return None

    def value_right(self, row: int, col: int, span: int = 6) -> Any:
        """First non-empty cell to the right of (row, col) -- the label's value."""
        for c in range(col + 1, min(col + span, self.max_col) + 1):
            v = self.v(row, c)
            if v is not None and _norm(v):
                return v
        return None

    def labelled_value(self, patterns: Iterable[str], max_row: Optional[int] = None,
                       min_row: int = 1) -> Any:
        hit = self.find_label(patterns, max_row=max_row, min_row=min_row)
        if hit is None:
            return None
        value = self.value_right(*hit)
        # Never let one label be read as another label's value.
        return None if _norm(value) in _NOT_A_VALUE or _norm(value).startswith("notes") else value


# ── Component block location ──────────────────────────────────────────────────

# Header text -> canonical component field. Matched against the header row and
# the row beneath it (these sheets split the header across two rows: SKU/VENDOR/
# DESCRIPTION/FC/LC on one, QTY/Sub Total/PRT on the next).
_HEADER_ALIASES: list[tuple[str, set[str]]] = [
    ("supplier_sku", {"product sku", "product", "item number", "item #", "sku", "product #"}),
    ("vendor", {"vendor", "supplier"}),
    ("description", {"description", "desc"}),
    ("first_cost", {"fc", "first cost"}),
    ("landed_cost", {"lc = (fc x 1.2)", "landed cost", "lc", "lc (fc x 1.2)"}),
    ("retail", {"lc x 6", "sub total", "subtotal", "retail"}),
    ("quantity", {"qty", "quantity"}),
    ("extended_total", {"prt", "ext", "extended", "totals", "pre-retail total"}),
]

_ALIAS_LOOKUP = {alias: fieldname for fieldname, aliases in _HEADER_ALIASES for alias in aliases}

# A row containing any of these exact strings closes the component block.
_TERMINATORS = {
    "factors", "total", "totals", "grand total", "notes", "notes :", "notes:",
    "retail", "wholesale", "pre-retail total", "pre retail total",
    "landed cost (lc factor)", "landed cost | 1.2", "arrangement labor | ar",
    "profit factor", "market price", "production price", "retail price",
}


def _header_map(grid: _Grid, row: int) -> dict[str, int]:
    """Canonical field -> column index, merging *row* with the row beneath it."""
    mapping: dict[str, int] = {}
    for r in (row, row + 1):
        for c in range(1, grid.max_col + 1):
            fieldname = _ALIAS_LOOKUP.get(_norm(grid.v(r, c)))
            if fieldname and fieldname not in mapping:
                mapping[fieldname] = c
    # A few sheets blanked out the DESCRIPTION caption (it survives as "     ").
    # The column is still there, wedged between VENDOR and FIRST COST.
    if "description" not in mapping and {"vendor", "first_cost"} <= mapping.keys():
        gap = mapping["first_cost"] - mapping["vendor"]
        if gap == 2:
            mapping["description"] = mapping["vendor"] + 1
    return mapping


def find_component_header(grid: _Grid, search_limit: int = 40) -> tuple[int, int, dict[str, int]]:
    """(header_row, header_end_row, column map). Raises if no block is found.

    Deliberately locates the block by header text rather than a fixed row range,
    because the same family puts it at row 12, 17, 18 or 19 depending on the file.
    The header spans two rows whenever the row beneath it also carries column
    names (``Qty`` / ``Sub Total`` / ``PRT``); ``header_end_row`` says where the
    data actually starts.
    """
    best: Optional[tuple[int, int, dict[str, int]]] = None
    for r in range(1, min(grid.max_row, search_limit) + 1):
        texts = {_norm(grid.v(r, c)) for c in range(1, grid.max_col + 1)}
        if not texts & _ALIAS_LOOKUP.keys():
            continue
        mapping = _header_map(grid, r)
        # a real component header identifies a line item AND a money/qty column
        if "supplier_sku" not in mapping and "description" not in mapping:
            continue
        if not ({"quantity", "extended_total", "first_cost"} & mapping.keys()):
            continue
        below = {_norm(grid.v(r + 1, c)) for c in range(1, grid.max_col + 1)}
        end = r + 1 if below & _ALIAS_LOOKUP.keys() else r
        if best is None or len(mapping) > len(best[2]):
            best = (r, end, mapping)
    if best is None:
        raise RecipeParseError("no component header row found")
    return best


def _is_terminator(grid: _Grid, row: int) -> bool:
    for c in range(1, grid.max_col + 1):
        v = grid.v(row, c)
        if isinstance(v, str) and _norm(v) in _TERMINATORS:
            return True
    return False


# ── Component classification ──────────────────────────────────────────────────

_MECHANIC_WORDS = (
    "foam", "rock", "acrylic", "mechanic", "moss", "glue", "wire", "sand",
    "gravel", "tape", "pick", "stake", "cement", "plaster", "soil",
)
_CONTAINER_WORDS = ("container", "vase", "planter", "pot", "urn", "cachepot", "bowl", "basket")


def classify_component(sku: Optional[str], description: Optional[str]) -> str:
    """Role of a line within the build: container / mechanics / product."""
    text = _norm(f"{sku or ''} {description or ''}")
    if not text:
        return "product"
    if any(w in text for w in _CONTAINER_WORDS):
        return "container"
    if any(w in text for w in _MECHANIC_WORDS):
        return "mechanics"
    return "product"


# ── Build type ────────────────────────────────────────────────────────────────

# Ordered: first hit wins, most specific families first. Keyed off the product /
# job description, which is what the sheets actually name the build.
_BUILD_TYPE_KEYWORDS: list[tuple[str, list[str]]] = [
    ("Wreath", ["wreath"]),
    ("Garland", ["garland", "swag", "roping"]),
    ("Centerpiece", ["centerpiece", "center piece", "tablescape"]),
    ("Topiary", ["topiary"]),
    # Tree-form species. Kept ahead of Planter/Plant & Bush deliberately: the
    # species is a stronger signal than the vessel or the word "bush", which is
    # why "4 Head Yucca Bush" reads as a Tree. `dracaena`/`croton`/`schefflera`/
    # `zamia` were missing entirely, which scattered one species across three
    # build types (dracaena landed in Container Only 5 / Tree 4 / Plant & Bush 1).
    # NOTE: do NOT add `eucalyptus` here — it is far more often a spray or a
    # floral component than a tree, and the tree builds already say "tree".
    ("Tree", ["tree", "ficus", "fiddle leaf", "fiddle-leaf", "palm", "yucca", "olive",
              "birch", "bamboo", "cypress", "citrus", "dracaena", "croton",
              "schefflera", "scheffelera", "schefflara", "zamia"]),
    ("Planter", ["planter", "pot", "potted", "urn", "jardiniere", "cachepot", "trough", "window box"]),
    # Non-tree plant material. `echeveria`/`greenery`/`sedum`/`tillandsia`/
    # `aeonium`/`donkey tail` were missing, so succulent builds fell through to
    # the taxonomy fallback and came back as "Container Only" (or NULL).
    ("Plant & Bush", ["bush", "plant", "succulent", "cactus", "fern", "grass", "ivy",
                      "philo", "monstera", "agave", "aloe", "sansevieria", "pothos",
                      "echeveria", "echeverria", "greenery", "sedum", "tillandsia",
                      "thilandsia", "aeonium", "donkey tail", "lichen",
                      # Matting products. Previously classified only by accident:
                      # bare-substring matching caught "grass" inside "Grassmat".
                      # Word boundaries broke that, so name them explicitly.
                      "grassmat", "grass mat", "moss mat"]),
    ("Floral Arrangement", ["arrangement", "bouquet", "floral", "flower", "hydrangea",
                            "orchid", "rose", "peony", "tulip", "amaryllis", "lily",
                            "anthurium", "protea", "dahlia", "ranunculus", "poinsettia",
                            "magnolia", "cherry blossom", "blossom", "allium"]),
    ("Branch & Stem", ["branch", "stem", "dragonwood", "draggon", "dragon wood", "willow",
                       "manzanita", "pole"]),
    ("Container Only", ["container", "vase", "vessel"]),
]

# Word-boundary matched. Previously these were bare substrings, which needed the
# hand-rolled `"pot "` hack to avoid hitting "pothos" -- and that silently failed
# on a trailing "Cement Pot". \b handles both, and stops keywords matching inside
# unrelated words (e.g. "stem" inside "system").
_BUILD_TYPE_PATTERNS = [
    (
        fam,
        re.compile(
            r"\b(?:"
            + "|".join(re.escape(k) for k in sorted(kws, key=len, reverse=True))
            + r")(?:es|s)?\b"
        ),
    )
    for fam, kws in _BUILD_TYPE_KEYWORDS
]

# taxonomy families that map cleanly onto a build type
_TAXONOMY_TO_BUILD = {
    # product_type_family
    "Tree": "Tree", "Wreath": "Wreath", "Garland": "Garland",
    "Container & Base": "Planter", "Greenery & Foliage": "Plant & Bush",
    "Floral Stems": "Floral Arrangement", "Spray & Picks": "Floral Arrangement",
    "Berries & Pods": "Branch & Stem",
    # category_family
    "Trees": "Tree", "Wreaths & Garland": "Wreath", "Florals": "Floral Arrangement",
    "Greenery & Plants": "Plant & Bush", "Containers & Vases": "Container Only",
    "Botanicals & Fillers": "Branch & Stem",
}


# Mentions of a vessel that say it is NOT part of the build. Matching these as
# "Container Only" inverts their meaning -- it filed four 9-10' Dracaena pom-pom
# trees ("... (Container Not Included)") as containers. Stripped before matching.
_VESSEL_NEGATION = re.compile(
    r"\(?\s*(?:"
    r"(?:no|not|without|excl\.?|excluding)\s+(?:container|vase|vessel|cont\.?)"
    r"|(?:container|vase|vessel|cont\.?)s?\s+(?:not|no)\s+(?:included|incl\.?|inc\.?)"
    r"|(?:container|vase|vessel)s?\s*(?:=|:)?\s*(?:n/?a|tbd|none)"
    r")[^)]*\)?",
    re.I,
)


def derive_build_type(*signals: Optional[str]) -> Optional[str]:
    """What kind of thing was built, from the product name / description.

    Falls back to :mod:`app.libs.taxonomy` so the vocabulary stays consistent
    with the Product Library filters. ``category_family`` is consulted last
    because it never returns "unknown" -- it would otherwise mask the specific
    build keywords above.
    """
    text = " ".join(str(s) for s in signals if s).lower()
    if not text.strip():
        return None
    # "(Container Not Included)" must not read as Container Only.
    text = _VESSEL_NEGATION.sub(" ", text)
    if not text.strip():
        return None
    for family, pattern in _BUILD_TYPE_PATTERNS:
        if pattern.search(text):
            return family
    by_type = _TAXONOMY_TO_BUILD.get(taxonomy.product_type_family(text) or "")
    return by_type or _TAXONOMY_TO_BUILD.get(taxonomy.category_family(text))


# ── Shared component reader ───────────────────────────────────────────────────


def _read_components(
    grid: _Grid,
    header_end: int,
    cols: dict[str, int],
    *,
    landed_factor: Optional[float],
    profit_factor: Optional[float],
    max_rows: int = 60,
) -> tuple[list[ParsedComponent], list[dict[str, Any]], int]:
    """Walk the component block. Returns (components, unpriced lines, last row).

    A row becomes a component only if something was actually committed to the
    build -- a non-zero quantity or extended total. Unit costs alone are not
    enough: the blank ``PRODUCTION_PRICE_WORK_SHEET`` template pre-fills Foam /
    Rocks / Acrylic at $5 each with quantity 0, and those are placeholders, not a
    build. Rows that have text but no money (continuation notes such as
    ``(BUFF ROCKS & ACRYLIC)`` or an unused ``Other Mechanics`` label) are not
    discarded either; they are returned separately and stored on the recipe
    header so the capture stays lossless.

    Some hand-built sheets split one line item over two rows -- vendor and product
    name on the first, dimensions and money on the second. When an unpriced text
    row is immediately followed by a priced row that has no SKU or vendor of its
    own, the two are merged so the supplier attribution is not lost.
    """
    components: list[ParsedComponent] = []
    unpriced: list[dict[str, Any]] = []
    blank_streak = 0
    last_row = header_end
    pending: Optional[dict[str, Any]] = None

    for row in range(header_end + 1, min(grid.max_row, header_end + max_rows) + 1):
        if _is_terminator(grid, row):
            break
        if grid.row_is_blank(row):
            pending = None
            blank_streak += 1
            if blank_streak >= 6:
                break
            continue
        blank_streak = 0
        last_row = row

        raw = grid.row_cells(row)
        get = lambda name: grid.v(row, cols[name]) if name in cols else None  # noqa: E731

        numbers = {
            "quantity": _number(get("quantity")),
            "first_cost": _number(get("first_cost")),
            "landed_cost": _number(get("landed_cost")),
            "retail": _number(get("retail")),
            "extended_total": _number(get("extended_total")),
        }
        sku = _text(get("supplier_sku"))
        vendor = _text(get("vendor"))
        description = _text(get("description"))

        named = bool(sku or vendor or description)
        # "Was anything committed to this build?" -- either money moved, or a
        # named material was used in a non-zero quantity. Unit costs alone are not
        # enough (the blank template pre-fills Foam/Rocks/Acrylic at qty 0), and a
        # bare quantity with no name and no money is leftover spreadsheet noise.
        if "quantity" in cols or "extended_total" in cols:
            committed = bool(numbers["extended_total"]) or (bool(numbers["quantity"]) and named)
        else:
            committed = any(numbers.values())

        if not committed:
            if sku or vendor or description:
                unpriced.append({"row": row, "cells": raw})
                # Only a row carrying supplier attribution can be the top half of
                # a split line item; a bare caption such as "Other Mechanics" is
                # a section label and must not glue itself onto the next row.
                pending = ({"row": row, "sku": sku, "vendor": vendor,
                            "description": description, "cells": raw}
                           if (sku or vendor) else None)
            else:
                pending = None
            continue

        raw_row: dict[str, Any] = {"row": row, "cells": raw}
        merged_from = None
        if pending and pending["row"] == row - 1 and not sku and not vendor:
            merged_from = pending["row"]
            sku = pending["sku"]
            vendor = pending["vendor"]
            if pending["description"]:
                description = (f"{pending['description']} {description}".strip()
                               if description else pending["description"])
            raw_row["merged_row"] = {"row": pending["row"], "cells": pending["cells"]}
            unpriced.pop()  # it is a real component line after all
        pending = None

        comp = ParsedComponent(
            line_order=len(components) + 1,
            component_label=classify_component(sku, description),
            vendor=vendor,
            supplier_sku=sku,
            description=description,
            **numbers,
            raw_row=raw_row,
        )
        comp.formulas = _component_formulas(
            grid, row, cols, comp, landed_factor=landed_factor, profit_factor=profit_factor
        )
        if merged_from is not None:
            comp.formulas["merged_from_row"] = merged_from
        components.append(comp)

    return components, unpriced, last_row


def _component_formulas(
    grid: _Grid,
    row: int,
    cols: dict[str, int],
    comp: ParsedComponent,
    *,
    landed_factor: Optional[float],
    profit_factor: Optional[float],
) -> dict[str, Any]:
    """Formula text + deviation notes for one component row.

    The expected chain is ``FC -> LC (x landed) -> retail (x profit) -> extended
    (x qty)``. Anything that deviates is recorded, never rewritten: the sheets are
    hand-built and what the business charged is the fact worth keeping.

    Deviations are split by intent, because the two mean different things:

    * ``anomalies`` -- the cell holds a *formula* that computes something other
      than the documented chain (the fiddle-leaf Draggon Wood row multiplying by
      the AR factor, the container row applying the profit factor twice, sheets
      quietly using a 1.12 landed factor). These are the ones worth reviewing.
    * ``flat_priced`` -- the cell holds a typed-in number rather than a formula,
      so the designer deliberately set a flat price (mechanics such as foam and
      acrylic are routinely priced this way). Recorded, but not an anomaly.
    """
    out: dict[str, Any] = {}
    for name, col in cols.items():
        formula = grid.formula(row, col)
        if formula:
            out[name] = formula

    anomalies: list[str] = []
    flat: list[str] = []
    fc, lc = comp.first_cost, comp.landed_cost
    retail, qty, ext = comp.retail, comp.quantity, comp.extended_total

    def note(field_name: str, message: str) -> None:
        (anomalies if field_name in out else flat).append(message)

    if fc and lc is not None and landed_factor and not _close(lc, fc * landed_factor):
        note("landed_cost",
             f"landed_cost {lc:g} != first_cost {fc:g} x landed factor {landed_factor:g} "
             f"(expected {fc * landed_factor:g}; implied factor {lc / fc:.4g})")
    if lc and retail is not None and profit_factor and not _close(retail, lc * profit_factor):
        note("retail",
             f"retail {retail:g} != landed_cost {lc:g} x profit factor {profit_factor:g} "
             f"(expected {lc * profit_factor:g}; implied factor {retail / lc:.4g})")
    if retail is not None and qty is not None and ext is not None and not _close(ext, retail * qty):
        note("extended_total",
             f"extended_total {ext:g} != retail {retail:g} x qty {qty:g} "
             f"(expected {retail * qty:g})")

    if anomalies:
        out["anomalies"] = anomalies
    if flat:
        out["flat_priced"] = flat
    return out


def _wholesale_tiers(grid: _Grid, start_row: int, end_row: int) -> list[dict[str, Any]]:
    """Every ``WHOLESALE n | DISCOUNT x%`` row, with its factor and price.

    Only cells at or right of the label are considered: the 2025 production sheet
    keeps its landed/AR/profit factors in column B on the very same rows, and
    those must not be mistaken for tier prices.
    """
    tiers: list[dict[str, Any]] = []
    for r in range(start_row, min(end_row, grid.max_row) + 1):
        label: Optional[str] = None
        label_col = grid.max_col + 1
        factor: Optional[float] = None
        price: Optional[float] = None
        for c in range(1, grid.max_col + 1):
            v = grid.v(r, c)
            text = _norm(v)
            if isinstance(v, str) and ("wholesale" in text or "discount" in text):
                label = f"{label} | {str(v).strip()}" if label else str(v).strip()
                label_col = min(label_col, c)
                continue
            if c < label_col:
                continue
            num = _number(v)
            if num is None:
                continue
            if 0 < num < 1:
                factor = num if factor is None else factor
            elif num >= 1:
                price = num if price is None else price
        if label:
            tiers.append({"label": label, "factor": factor, "price": price})
    return tiers


# ── Adapter: TBDG PRODUCTION PRICE WORKSHEET 2025 ─────────────────────────────


def _parse_production_2025(grid: _Grid) -> ParsedRecipe:
    client = _text(grid.labelled_value([r"client name.*"], max_row=12))
    job = _text(grid.labelled_value([r"job description.*"], max_row=12))

    landed = _number(grid.labelled_value([r"landed cost.*"]))
    ar = _number(grid.labelled_value([r"arrangement labor.*"]))
    profit = _number(grid.labelled_value([r"profit factor.*"]))

    header_row, header_end, cols = find_component_header(grid)
    components, unpriced, block_end = _read_components(
        grid, header_end, cols, landed_factor=landed, profit_factor=profit
    )

    # Totals sit to the right of their own labels below the component block.
    # ``min_row`` keeps the identically-named column headers out of the way.
    pre_retail = _number(grid.labelled_value([r"pre-?retail total"], min_row=block_end + 1))
    retail_hit = grid.find_label([r"x ar.*", r"retail price"], min_row=block_end + 1)
    retail = _number(grid.value_right(*retail_hit)) if retail_hit else None
    if retail is None and retail_hit:  # "RETAIL PRICE" labels the cell above it
        retail = _number(grid.v(retail_hit[0] - 1, retail_hit[1]))

    if pre_retail is None:
        pre_retail = sum(c.extended_total or 0 for c in components) or None

    factor_hit = grid.find_label([r"factors"], min_row=block_end + 1)
    tiers = _wholesale_tiers(grid, factor_hit[0] if factor_hit else block_end + 1, grid.max_row)

    recipe = ParsedRecipe(
        format_family=FORMAT_PRODUCTION_2025,
        description=job,
        recipe_year=2025,
        dimensions=_dimension_block(grid, [r"overall arrangement"], max_row=header_row),
        container_details=_dimension_block(grid, [r"container"], max_row=header_row, material=True),
        pricing_summary={
            "pre_retail_total": pre_retail,
            "retail": retail,
            "component_extended_sum": _round(sum(c.extended_total or 0 for c in components)),
            "factors": {"landed_cost": landed, "arrangement_labor": ar, "profit": profit},
            "wholesale_tiers": tiers,
        },
        raw_header={
            "a1": _text(grid.v(1, 1)),
            "client": client,
            "job_description": job,
            "notes": _text(grid.labelled_value([r"notes.*"], max_row=12)),
            "component_header_row": header_row,
            "component_columns": {k: _col_letter(v) for k, v in cols.items()},
            "unpriced_lines": unpriced,
            "sheets": grid.sheet_names,
        },
        visual_reference_count=grid.image_count,
        components=components,
    )
    _apply_derived(recipe, client=client)
    return recipe


# ── Adapter: TBDG PRICING WORKSHEET 2025 (quote variant) ──────────────────────


def _parse_pricing_2025(grid: _Grid) -> ParsedRecipe:
    client = _text(grid.labelled_value([r"client name.*"], max_row=12))
    job = _text(grid.labelled_value([r"job description.*"], max_row=12))

    landed = _number(grid.labelled_value([r"landed cost.*"]))
    ar = _number(grid.labelled_value([r"arrangement labor.*"]))
    profit = _number(grid.labelled_value([r"profit factor.*"]))

    header_row, header_end, cols = find_component_header(grid)
    components, unpriced, block_end = _read_components(
        grid, header_end, cols, landed_factor=landed, profit_factor=profit
    )

    pre_retail = _number(grid.labelled_value([r"pre-?retail total"], min_row=block_end + 1))
    retail_hit = grid.find_label([r"retail$", r"retail price"], min_row=block_end + 1)
    retail = _number(grid.value_right(*retail_hit)) if retail_hit else None
    tiers = _wholesale_tiers(grid, block_end + 1, grid.max_row)

    recipe = ParsedRecipe(
        format_family=FORMAT_PRICING_2025,
        description=job,
        recipe_year=2025,
        dimensions=_dimension_block(grid, [r"overall arrangement"], max_row=header_row),
        container_details=_dimension_block(grid, [r"container"], max_row=header_row, material=True),
        pricing_summary={
            "pre_retail_total": pre_retail,
            "retail": retail,
            "component_extended_sum": _round(sum(c.extended_total or 0 for c in components)),
            "factors": {"landed_cost": landed, "arrangement_labor": ar, "profit": profit},
            "wholesale_tiers": tiers,
        },
        raw_header={
            "a1": _text(grid.v(1, 1)),
            "client": client,
            "job_description": job,
            "notes": _text(grid.labelled_value([r"notes.*"], max_row=12)),
            "component_header_row": header_row,
            "component_columns": {k: _col_letter(v) for k, v in cols.items()},
            "unpriced_lines": unpriced,
            "sheets": grid.sheet_names,
        },
        visual_reference_count=grid.image_count,
        components=components,
    )
    _apply_derived(recipe, client=client)
    return recipe


# ── Adapter: Creative Branch Recipe 2020-2023 (+ CANTONI variants) ────────────

# Legacy sheets bake the factors into the cell formulas rather than naming them.
_LEGACY_LANDED_FACTOR = 1.2
_LEGACY_PROFIT_FACTOR = 6.0

_OWN_CODE_LABELS = [r"product # ?:?", r"item # ?:?", r"cb item # ?:?", r"creative branch item # ?:?"]
_CUSTOMER_CODE_LABELS = [r"cantoni item # ?:?", r"customer item # ?:?"]


def _parse_creative_branch(grid: _Grid) -> ParsedRecipe:
    header_row, header_end, cols = find_component_header(grid)

    item_code = _text(grid.labelled_value(_OWN_CODE_LABELS, max_row=header_row - 1))
    customer_code = _text(grid.labelled_value(_CUSTOMER_CODE_LABELS, max_row=header_row - 1))
    description = _text(grid.labelled_value([r"description ?:?"], max_row=header_row - 1))

    components, unpriced, block_end = _read_components(
        grid, header_end, cols,
        landed_factor=_LEGACY_LANDED_FACTOR, profit_factor=_LEGACY_PROFIT_FACTOR,
    )

    # The bottom block is three stacked labels: Total / Retail / Wholesale. Each
    # also exists as a column header above, hence the floor at the block end.
    floor = block_end + 1
    total = _number(grid.labelled_value([r"total$"], min_row=floor))
    retail = _number(grid.labelled_value([r"retail$"], min_row=floor))
    wholesale = _number(grid.labelled_value([r"wholesale$"], min_row=floor))

    a1 = _text(grid.v(1, 1)) or ""
    year_match = re.search(r"(20\d{2})", a1)

    recipe = ParsedRecipe(
        format_family=FORMAT_CREATIVE_BRANCH,
        item_code=item_code,
        customer_item_code=customer_code,
        description=description,
        recipe_year=int(year_match.group(1)) if year_match else None,
        dimensions=_dimension_block(grid, [r"overall arrangement"], max_row=header_row),
        container_details=_dimension_block(grid, [r"container"], max_row=header_row, material=True),
        pricing_summary={
            "component_total": total,
            "retail": retail,
            "wholesale": wholesale,
            "component_extended_sum": _round(sum(c.extended_total or 0 for c in components)),
            "factors": {
                "landed_cost": _LEGACY_LANDED_FACTOR,
                "profit": _LEGACY_PROFIT_FACTOR,
                "retail_markup": _round(retail / total) if retail and total else None,
                "wholesale_discount": _round(wholesale / retail) if wholesale and retail else None,
            },
            "wholesale_tiers": [],
        },
        raw_header={
            "a1": a1,
            "weight": _text(grid.labelled_value([r"weight"], max_row=header_row - 1)),
            "notes": _text(grid.labelled_value([r"notes ?:?"], min_row=header_row)),
            "component_header_row": header_row,
            "component_columns": {k: _col_letter(v) for k, v in cols.items()},
            "unpriced_lines": unpriced,
            "sheets": grid.sheet_names,
        },
        visual_reference_count=grid.image_count,
        components=components,
    )
    _apply_derived(recipe)
    return recipe


# ── Shared post-processing ────────────────────────────────────────────────────

_DIM_KEYS = ("length", "width", "height", "depth", "material")

# Text that is a label in its own right and can never be a dimension's value.
# The 2025 production sheet puts the arrangement block in column A and the
# container block in column C, so scanning too far right off "Length" (A7) lands
# on the container's own "Length" label (C7) unless these are rejected.
_NOT_A_VALUE = set(_DIM_KEYS) | {
    "container", "overall arrangement", "weight", "notes", "description",
    "product", "product sku", "vendor", "client name #", "job description",
}


def _dimension_value(grid: _Grid, row: int, col: int) -> Optional[str]:
    """The value for a dimension label, or None if the neighbour is a label."""
    for c in range(col + 1, min(col + 2, grid.max_col) + 1):
        value = _text(grid.v(row, c))
        if value is None:
            continue
        return None if _norm(value) in _NOT_A_VALUE else value
    return None


def _dimension_block(grid: _Grid, anchor_patterns: list[str], *, max_row: int,
                     material: bool = False) -> dict[str, Any]:
    """Read the Length/Width/Height/Depth (+Material) labels under an anchor.

    Both the 2025 and legacy layouts stack these labels vertically under an
    ``Overall Arrangement`` / ``Container`` heading, but in different columns and
    at different rows, so the block is located by its anchor rather than by cell.
    """
    anchor = grid.find_label(anchor_patterns, max_row=max_row)
    if anchor is None:
        return {}
    arow, acol = anchor
    out: dict[str, Any] = {}
    # The heading itself sometimes carries a value (e.g. Container -> "GLASS").
    heading_value = _dimension_value(grid, arow, acol)
    for r in range(arow, min(arow + 8, max_row)):
        key = _norm(grid.v(r, acol))
        if key in _DIM_KEYS and (material or key != "material"):
            value = _dimension_value(grid, r, acol)
            if value is not None and key not in out:
                out[key] = value
    if material and heading_value and "material" not in out:
        out["material"] = heading_value
    return out


def _round(value: Optional[float], places: int = 4) -> Optional[float]:
    return None if value is None else round(value, places)


def _apply_derived(recipe: ParsedRecipe, client: Optional[str] = None) -> None:
    """Fill build_type / product_family from whatever text the sheet gave us."""
    signals = [recipe.description, recipe.item_code]
    signals += [c.description for c in recipe.components[:6]]
    recipe.build_type = derive_build_type(recipe.description) or derive_build_type(*signals)
    recipe.product_family = taxonomy.category_family(recipe.description, recipe.build_type)
    if client:
        recipe.raw_header.setdefault("client", client)


# ── Entry points ──────────────────────────────────────────────────────────────


def sniff_format(path: str | Path) -> tuple[str, Optional[str]]:
    """(format_family, raw A1 text) without a full parse."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        a1 = wb.active["A1"].value
    finally:
        wb.close()
    text = _norm(a1)
    if "production price worksheet" in text:
        return FORMAT_PRODUCTION_2025, _text(a1)
    if "pricing worksheet" in text and "2025" in text:
        return FORMAT_PRICING_2025, _text(a1)
    if "price sheet cantoni" in text:
        return FORMAT_PRICE_SHEET, _text(a1)
    if "creative branch" in text:
        return FORMAT_CREATIVE_BRANCH, _text(a1)
    return FORMAT_UNKNOWN, _text(a1)


_ADAPTERS = {
    FORMAT_PRODUCTION_2025: _parse_production_2025,
    FORMAT_PRICING_2025: _parse_pricing_2025,
    FORMAT_CREATIVE_BRANCH: _parse_creative_branch,
}


def parse_recipe_xlsx(path: str | Path, *, format_family: Optional[str] = None) -> ParsedRecipe:
    """Parse one worksheet. Raises :class:`RecipeParseError` if unusable."""
    path = Path(path)
    if format_family is None:
        format_family, _ = sniff_format(path)
    adapter = _ADAPTERS.get(format_family)
    if adapter is None:
        raise RecipeParseError(f"no adapter for format {format_family!r}")
    grid = _Grid(path)
    recipe = adapter(grid)
    if not recipe.components:
        raise RecipeParseError("no priced components (blank template?)")
    if not recipe.description and not recipe.item_code:
        # Some quote sheets leave JOB DESCRIPTION blank; the file name is the only
        # name the build ever had, so use it rather than storing nothing. Legacy
        # sheets already carry an item code, so they keep a null description.
        recipe.description = re.sub(r"[_\-]+", " ", path.stem).strip()
        recipe.raw_header["description_source"] = "file_name"
        _apply_derived(recipe)
    return recipe


# ── Provenance helpers ────────────────────────────────────────────────────────

_PRICING_DIR = re.compile(r"^pricing(\s+\d+)?$", re.I)


def source_collection(relative_path: str | Path) -> str:
    """Human label for where a recipe came from.

    The corpus is mirrored across ``PRICING``/``PRICING 2..5`` copies, so those
    wrapper directories are dropped -- what is left ("WAYFAIR", "Trees 2022",
    "Recipes Cantoni 2022") is the collection that actually means something.
    """
    parts = [p for p in Path(relative_path).parts[:-1] if not _PRICING_DIR.match(p)]
    return "/".join(parts) if parts else "TBDG Root"


def recipe_year_hint(relative_path: str | Path) -> Optional[int]:
    """A year from the path, used when the sheet header does not carry one."""
    years = re.findall(r"(20[12]\d)", str(relative_path))
    return int(years[-1]) if years else None
