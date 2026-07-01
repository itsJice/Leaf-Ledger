"""Catalog-file parsing helpers for supplier product intake."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional

@dataclass
class CatalogRow:
    supplier_sku: str
    name: str
    description: Optional[str] = None
    upc: Optional[str] = None
    current_price: Optional[float] = None
    unit: str = "each"
    category: str = "other"
    moq: Optional[int] = None
    box_qty: Optional[int] = None
    case_qty: Optional[int] = None
    cubic_ft: Optional[float] = None
    source_page: Optional[int] = None
    source_section: Optional[str] = None
    raw_data: dict[str, Any] = field(default_factory=dict)


SKU_KEYS = ("supplier_sku", "sku", "itemno", "item_no", "item number", "item #", "item")
NAME_KEYS = ("name", "product name", "description", "desc", "item description")
UPC_KEYS = ("upc", "barcode", "gtin")
PRICE_KEYS = ("account price", "customer price", "wholesale price", "price", "listprice", "list price", "msrp")
UNIT_KEYS = ("unit", "uom", "unit of measure")
MOQ_KEYS = ("moq", "min qty", "minimum qty", "minimum order", "minqty")
BOX_KEYS = ("box qty", "boxqty", "box", "box/cs", "box cs")
CASE_KEYS = ("case qty", "caseqty", "case", "case pack")
CATEGORY_KEYS = ("category", "section", "department", "collection")

VALID_UNITS = {
    "stem", "pot", "flat", "bunch", "each", "box", "case", "bag", "roll",
    "yard", "foot", "piece", "set", "pair",
}


def safe_int(raw: Any) -> Optional[int]:
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError, AttributeError):
        return None


def parse_price(raw: str) -> Optional[float]:
    if not raw:
        return None
    cleaned = str(raw).replace(",", "").replace("$", "").strip()
    match = re.search(r"\d+\.?\d*", cleaned)
    return float(match.group(0)) if match else None


def normalize_unit(raw_unit: Optional[str]) -> str:
    if not raw_unit:
        return "each"
    lower = raw_unit.lower().strip()
    if lower in VALID_UNITS:
        return lower
    if "ea" in lower or "each" in lower:
        return "each"
    if "stem" in lower:
        return "stem"
    if "bundle" in lower or "bunch" in lower:
        return "bunch"
    if "case" in lower:
        return "case"
    if "box" in lower:
        return "box"
    return "each"


def normalize_category(raw_category: Optional[str]) -> str:
    if not raw_category:
        return "other"
    lower = raw_category.lower().strip()
    if any(term in lower for term in ("holiday", "christmas", "fall", "season")):
        return "seasonal"
    if any(term in lower for term in ("bouquet", "flower", "floral", "dahlia", "rose", "tulip")):
        return "florals"
    if any(term in lower for term in ("foliage", "greenery", "leaf", "grass")):
        return "greenery"
    if any(term in lower for term in ("container", "vase", "pot", "planter")):
        return "containers"
    if "moss" in lower:
        return "moss"
    if "branch" in lower:
        return "branches"
    if "tree" in lower:
        return "trees"
    return "other"


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text) if text else None


def _key_map(row: dict[str, Any]) -> dict[str, Any]:
    return {re.sub(r"[^a-z0-9]+", " ", str(k).lower()).strip(): v for k, v in row.items()}


def _pick(mapped: dict[str, Any], keys: Iterable[str]) -> Optional[str]:
    normalized = {re.sub(r"[^a-z0-9]+", " ", key.lower()).strip() for key in keys}
    for key, value in mapped.items():
        if key in normalized:
            picked = _clean(value)
            if picked:
                return picked
    return None


def _parse_box_case(value: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    if not value:
        return None, None
    match = re.search(r"(\d+)\s*/\s*(\d+)", value)
    if match:
        return safe_int(match.group(1)), safe_int(match.group(2))
    return safe_int(value), None


def _parse_float(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    cleaned = value.replace(",", "")
    match = re.search(r"\d+(?:\.\d+)?", cleaned)
    return float(match.group(0)) if match else None


def _row_from_mapping(row: dict[str, Any], source_name: str) -> Optional[CatalogRow]:
    mapped = _key_map(row)
    sku = _pick(mapped, SKU_KEYS)
    name = _pick(mapped, NAME_KEYS)
    if not sku or not name:
        return None

    box_qty, case_qty = _parse_box_case(_pick(mapped, BOX_KEYS))
    category_raw = _pick(mapped, CATEGORY_KEYS)
    unit_raw = _pick(mapped, UNIT_KEYS)
    raw = {str(k): v for k, v in row.items() if _clean(v)}
    raw.update({"catalog_source_type": "file", "catalog_source_name": source_name})

    return CatalogRow(
        supplier_sku=sku,
        name=name,
        description=name,
        upc=_pick(mapped, UPC_KEYS),
        current_price=parse_price(_pick(mapped, PRICE_KEYS) or ""),
        unit=normalize_unit(unit_raw),
        category=normalize_category(category_raw),
        moq=safe_int(_pick(mapped, MOQ_KEYS)),
        box_qty=box_qty,
        case_qty=case_qty,
        source_section=category_raw,
        raw_data=raw,
    )


def parse_csv_catalog(data: bytes, filename: str) -> list[CatalogRow]:
    sample = data[:4096].decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    except csv.Error:
        dialect = csv.excel
    stream = io.StringIO(data.decode("utf-8-sig", errors="replace"))
    reader = csv.DictReader(stream, dialect=dialect)
    return [row for source in reader if (row := _row_from_mapping(source, filename))]


def parse_xlsx_catalog(data: bytes, filename: str) -> list[CatalogRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[CatalogRow] = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            continue
        header_names = [str(h or "").strip() for h in headers]
        for values in iterator:
            source = {header_names[i]: values[i] for i in range(min(len(header_names), len(values)))}
            parsed = _row_from_mapping(source, f"{filename} / {sheet.title}")
            if parsed:
                parsed.source_section = parsed.source_section or sheet.title
                parsed.raw_data["catalog_sheet"] = sheet.title
                rows.append(parsed)
    return rows


def _text_between(text: str, start: str, *stops: str) -> Optional[str]:
    start_match = re.search(re.escape(start), text, flags=re.I)
    if not start_match:
        return None
    rest = text[start_match.end() :]
    stop_positions = [
        match.start()
        for stop in stops
        if (match := re.search(re.escape(stop), rest, flags=re.I))
    ]
    value = rest[: min(stop_positions)] if stop_positions else rest
    return _clean(value)


def _parse_allstate_product_block(block: str, page_number: int, section: Optional[str], filename: str) -> Optional[CatalogRow]:
    header = re.search(r"ItemNo:\s*([A-Z0-9./-]+)\s+UPC:\s*([0-9-]+)", block, flags=re.I)
    if not header:
        return None
    sku = header.group(1).strip()
    upc = header.group(2).strip()
    desc = _text_between(block, "Desc:", "Box/Cs:", "ListPrice:")
    box_text = _text_between(block, "Box/Cs:", "Min Qty:", "ListPrice:")
    min_text = _text_between(block, "Min Qty:", "ListPrice:", "CubicFt:")
    price_text = _text_between(block, "ListPrice:", "CubicFt:")
    cubic_text = _text_between(block, "CubicFt:", "Fragile:", "ItemNo:")
    box_qty, case_qty = _parse_box_case(box_text)
    price_match = re.search(r"\$?\s*([0-9][0-9,.]*(?:\.\d+)?)", price_text or "")

    raw = {
        "catalog_source_type": "pdf_catalog",
        "catalog_source_name": filename,
        "catalog_page": page_number,
        "catalog_section": section,
        "UPC": upc,
        "Desc": desc,
        "Box/Cs": box_text,
        "Min Qty": min_text,
        "ListPrice": price_text,
        "CubicFt": cubic_text,
        "detail_status": "pending",
        "image_status": "pending",
        "price_source": "catalog_list_price",
        "category_tags": [section] if section else [],
    }
    return CatalogRow(
        supplier_sku=sku,
        name=desc or sku,
        description=desc,
        upc=upc,
        current_price=float(price_match.group(1).replace(",", "")) if price_match else None,
        unit=normalize_unit(price_text),
        category=normalize_category(section),
        moq=safe_int(min_text.split()[0] if min_text else None),
        box_qty=box_qty,
        case_qty=case_qty,
        cubic_ft=_parse_float(cubic_text),
        source_page=page_number,
        source_section=section,
        raw_data=raw,
    )


def _parse_allstate_column(text: str, page_number: int, section: Optional[str], filename: str) -> list[CatalogRow]:
    blocks = re.split(r"(?=ItemNo:\s*[A-Z0-9./-]+\s+UPC:)", text)
    rows = []
    for block in blocks:
        parsed = _parse_allstate_product_block(block, page_number, section, filename)
        if parsed:
            rows.append(parsed)
    return rows


def parse_allstate_pdf_catalog(path: str | Path, filename: Optional[str] = None) -> list[CatalogRow]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("PDF catalog import requires pdfplumber in the backend environment.") from exc

    file_label = filename or Path(path).name
    rows: list[CatalogRow] = []
    with pdfplumber.open(str(path)) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            header_text = page.crop((0, 0, page.width, min(28, page.height))).extract_text() or ""
            section_match = re.search(r"\(([^)]+)\)", header_text)
            section = section_match.group(1).strip() if section_match else None
            midpoint = page.width / 2
            columns = [
                page.crop((0, 20, midpoint, page.height)),
                page.crop((midpoint, 20, page.width, page.height)),
            ]
            for column in columns:
                text = column.extract_text(x_tolerance=1, y_tolerance=3) or ""
                rows.extend(_parse_allstate_column(text, index, section, file_label))
    deduped: dict[tuple[str, Optional[int]], CatalogRow] = {}
    for row in rows:
        deduped.setdefault((row.supplier_sku.upper(), row.source_page), row)
    return list(deduped.values())


def _group_words_by_line(words: list[dict[str, Any]]) -> list[str]:
    lines: list[list[dict[str, Any]]] = []
    for word in sorted(words, key=lambda item: (float(item["top"]), float(item["x0"]))):
        if not lines or abs(float(lines[-1][0]["top"]) - float(word["top"])) > 3:
            lines.append([word])
        else:
            lines[-1].append(word)
    return [
        _clean(" ".join(str(word["text"]) for word in sorted(line, key=lambda item: float(item["x0"])))) or ""
        for line in lines
    ]


def _american_best_column_bounds(page_width: float, x0: float) -> tuple[float, float]:
    # American Best catalog pages use three product columns with consistent starts.
    if x0 < page_width * 0.30:
        return 0, page_width * 0.31
    if x0 < page_width * 0.62:
        return page_width * 0.31, page_width * 0.62
    return page_width * 0.62, page_width


def _parse_min_case(line: Optional[str]) -> tuple[Optional[int], Optional[int]]:
    if not line:
        return None, None
    min_match = re.search(r"\bMin\.?\s+(\d+)", line, flags=re.I)
    case_match = re.search(r"\bCase\s+(\d+)", line, flags=re.I)
    return (
        safe_int(min_match.group(1)) if min_match else None,
        safe_int(case_match.group(1)) if case_match else None,
    )


def parse_american_best_pdf_catalog(path: str | Path, filename: Optional[str] = None) -> list[CatalogRow]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("PDF catalog import requires pdfplumber in the backend environment.") from exc

    file_label = filename or Path(path).name
    rows: list[CatalogRow] = []
    sku_pattern = re.compile(r"^\d{6,}[A-Z0-9]*[A-Z]$")
    with pdfplumber.open(str(path)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            words = page.extract_words(x_tolerance=1, y_tolerance=3) or []
            sku_words = [
                word for word in words
                if sku_pattern.match(str(word.get("text", "")).strip())
                and float(word.get("top", 0)) > 80
                and "page" not in str(word.get("text", "")).lower()
            ]
            for sku_word in sku_words:
                sku = str(sku_word["text"]).strip()
                x_min, x_max = _american_best_column_bounds(float(page.width), float(sku_word["x0"]))
                next_sku_tops = [
                    float(other["top"])
                    for other in sku_words
                    if other is not sku_word
                    and _american_best_column_bounds(float(page.width), float(other["x0"])) == (x_min, x_max)
                    and float(other["top"]) > float(sku_word["top"]) + 5
                ]
                y_min = float(sku_word["top"]) + 5
                y_max = min(next_sku_tops) if next_sku_tops else float(page.height) - 30
                block_words = [
                    word for word in words
                    if x_min <= float(word["x0"]) < x_max
                    and y_min < float(word["top"]) < y_max
                    and str(word["text"]).strip() != sku
                ]
                lines = [line for line in _group_words_by_line(block_words) if line]
                if not lines:
                    continue
                price_idx = next((idx for idx, line in enumerate(lines) if re.search(r"\$\s*\d", line)), None)
                if price_idx is None:
                    continue
                min_idx = next((idx for idx, line in enumerate(lines) if re.search(r"\bMin\.?\b", line, flags=re.I)), None)
                name = lines[0]
                color = " ".join(lines[1:price_idx]).strip() or None
                price = parse_price(lines[price_idx])
                moq, case_qty = _parse_min_case(lines[min_idx] if min_idx is not None else None)
                if not name or not price:
                    continue
                sku_top = float(sku_word["top"])
                card_images = [
                    image for image in (page.images or [])
                    if float(image.get("width", 0)) > 30
                    and float(image.get("height", 0)) > 30
                    and x_min <= ((float(image["x0"]) + float(image["x1"])) / 2) <= x_max
                    and 80 < float(image.get("top", 0)) < sku_top
                    and float(image.get("bottom", 0)) <= sku_top + 16
                ]
                image_bbox = None
                if card_images:
                    image = max(card_images, key=lambda item: float(item.get("bottom", 0)))
                    image_bbox = [
                        max(0, float(image["x0"])),
                        max(0, float(image["top"])),
                        min(float(page.width), float(image["x1"])),
                        min(float(page.height), float(image["bottom"])),
                    ]
                else:
                    image_bbox = [
                        max(0, x_min + 6),
                        max(80, sku_top - 155),
                        min(float(page.width), x_max - 6),
                        max(80, sku_top - 5),
                    ]
                raw = {
                    "catalog_source_type": "pdf_catalog",
                    "catalog_source_name": file_label,
                    "catalog_page": page_index,
                    "catalog_section": Path(file_label).stem,
                    "Color": color,
                    "Min/Case": lines[min_idx] if min_idx is not None else None,
                    "price_source": "catalog_price",
                    "category_tags": [Path(file_label).stem],
                    "image_status": "pending_pdf_crop",
                    "pdf_image_bbox": image_bbox,
                    "pdf_image_source": "embedded_image" if card_images else "estimated_card_crop",
                }
                rows.append(CatalogRow(
                    supplier_sku=sku,
                    name=name,
                    description=f"{name} - {color}" if color else name,
                    current_price=price,
                    unit="each",
                    category=normalize_category(f"{Path(file_label).stem} {name}"),
                    moq=moq,
                    case_qty=case_qty,
                    source_page=page_index,
                    source_section=Path(file_label).stem,
                    raw_data=raw,
                ))
    deduped: dict[str, CatalogRow] = {}
    for row in rows:
        deduped.setdefault(row.supplier_sku.upper(), row)
    return list(deduped.values())


def parse_catalog_file(data: bytes, filename: str, parser_hint: Optional[str] = None) -> list[CatalogRow]:
    suffix = Path(filename).suffix.lower()
    hint = (parser_hint or "").lower()
    if suffix == ".csv":
        return parse_csv_catalog(data, filename)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx_catalog(data, filename)
    if suffix == ".pdf" or hint == "allstate_pdf":
        temp_path = Path("/tmp") / f"leaf-ledger-catalog-{Path(filename).name}"
        temp_path.write_bytes(data)
        try:
            rows = parse_allstate_pdf_catalog(temp_path, filename)
            return rows if rows else parse_american_best_pdf_catalog(temp_path, filename)
        finally:
            try:
                temp_path.unlink()
            except OSError:
                pass
    raise ValueError("Unsupported catalog file. Upload CSV, XLSX, or a structured Allstate PDF.")
