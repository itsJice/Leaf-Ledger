#!/usr/bin/env python3
"""Import one THE FINDINGS catalog using small, independent transactions.

Why this exists: the Supabase session pooler repeatedly stalls or drops a single
large multi-row transaction (observed on Burton + Burton at 10,488 rows, and
again on Schusters at only 438 rows — so it is duration/pooler related, not a
size threshold). One dropped connection rolls back the entire import.

This commits in CHUNK-sized transactions on a fresh connection each time, with
retries. Upserts are idempotent (ON CONFLICT supplier_id, supplier_sku), so a
partial run is safe to simply re-run.

    set -a && . ./.env.supabase && set +a
    .venv/bin/python scripts/load_findings_chunked.py --file "<path.xlsx>" \
        [--supplier "Name"] [--commit] [--chunk 250]

--supplier defaults to the file's own "supplier" column; the supplier row is
created if it does not exist (same rule as load_all_findings).
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
import time
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(BACKEND / "scripts"))

import asyncpg  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.libs.findings_intake import parse_findings_xlsx  # noqa: E402
from load_findings import build_upsert, dual_price_columns_exist, _values  # noqa: E402
from load_all_findings import MAPPING  # noqa: E402  (filename -> authoritative supplier id)

RETRYABLE = (asyncpg.exceptions.InterfaceError,
             asyncpg.exceptions.ConnectionDoesNotExistError,
             asyncpg.exceptions.TooManyConnectionsError,
             OSError, asyncio.TimeoutError)


def supplier_from_file(path: Path) -> str | None:
    wb = load_workbook(path, read_only=True)
    it = wb.active.iter_rows(values_only=True)
    hdr = [str(h).strip().lower() if h else "" for h in next(it)]
    idx = hdr.index("supplier") if "supplier" in hdr else None
    name = None
    if idx is not None:
        for row in it:
            if idx < len(row) and row[idx]:
                name = str(row[idx]).strip()
                break
    wb.close()
    return name


async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--supplier", default=None)
    ap.add_argument("--commit", action="store_true")
    ap.add_argument("--chunk", type=int, default=250)
    a = ap.parse_args()

    path = Path(a.file)
    name = a.supplier or supplier_from_file(path)
    if not name:
        print("ERROR: no --supplier given and no 'supplier' column in the file")
        return 2

    # Ten of the established catalogs name their supplier differently inside the
    # file than the DB row does ("Winward" vs "Winward Silks", "PMJC Inc" vs
    # "PMJC", "Allstate Floral & Craft" vs "Allstate"...). Matching on the file's
    # name alone would create a SECOND supplier row and split that supplier's
    # catalog in two — which is how the stray empty "Autograph Foliage" row
    # appeared alongside "Autograph Foliages". MAPPING pins filename -> id and
    # wins whenever the file is one of the known catalogs.
    mapped_id = MAPPING.get(path.name) if not a.supplier else None

    rows, rep = parse_findings_xlsx(str(path))
    print(f"file      : {path.name}")
    print(f"supplier  : {name}")
    print(f"rows_ok={rep['rows_ok']} dropped={rep.get('dropped_missing_sku_or_name')} "
          f"unique_sku={len({r.supplier_sku for r in rows if r.supplier_sku})}")

    url = os.environ["DATABASE_URL"]
    c = await asyncpg.connect(url, statement_cache_size=0)
    if mapped_id is not None:
        sid = await c.fetchval("SELECT id FROM suppliers WHERE id=$1", mapped_id)
        db_name = await c.fetchval("SELECT name FROM suppliers WHERE id=$1", mapped_id)
        if sid and (db_name or "").lower() != name.lower():
            print(f"note      : file says {name!r}, using mapped supplier "
                  f"{db_name!r} (id={sid}) — avoids creating a duplicate")
        if db_name:
            name = db_name
    else:
        sid = await c.fetchval(
            "SELECT id FROM suppliers WHERE name ILIKE $1 ORDER BY id LIMIT 1", name)
    ext = await dual_price_columns_exist(c)
    before = await c.fetchval("SELECT COUNT(*) FROM products WHERE supplier_id=$1", sid) if sid else 0
    await c.close()

    if sid is None:
        if not a.commit:
            print("supplier   : NEW (would be created)")
        else:
            c = await asyncpg.connect(url, statement_cache_size=0)
            sid = await c.fetchval(
                "INSERT INTO suppliers (name, credential_status, notes) "
                "VALUES ($1,'n/a',$2) RETURNING id", name, "auto-created by load_findings_chunked")
            await c.close()
            print(f"supplier   : created id={sid}")
    print(f"supplier id={sid} | existing rows={before:,} | dual-price columns={ext}")

    if not a.commit:
        print("\nDRY RUN — re-run with --commit to write.")
        return 0

    sql = build_upsert(ext)
    values = [_values(sid, p, ext) for p in rows]
    n_chunks = (len(values) + a.chunk - 1) // a.chunk
    done = 0
    t0 = time.time()
    for i in range(0, len(values), a.chunk):
        chunk = values[i:i + a.chunk]
        for attempt in range(4):
            cc = None
            try:
                cc = await asyncpg.connect(url, statement_cache_size=0, timeout=30)
                async with cc.transaction():
                    await cc.executemany(sql, chunk)
                break
            except RETRYABLE as e:
                if attempt == 3:
                    print(f"chunk {i // a.chunk + 1}/{n_chunks} FAILED: {type(e).__name__}: {e}")
                    raise
                await asyncio.sleep(2 * (attempt + 1))
            finally:
                if cc is not None:
                    try:
                        await cc.close()
                    except Exception:
                        pass
        done += len(chunk)
        if (i // a.chunk + 1) % 10 == 0 or i + a.chunk >= len(values):
            print(f"  {done:,}/{len(values):,} committed ({i // a.chunk + 1}/{n_chunks} chunks, "
                  f"{time.time() - t0:.0f}s)")

    c = await asyncpg.connect(url, statement_cache_size=0)
    after = await c.fetchval("SELECT COUNT(*) FROM products WHERE supplier_id=$1", sid)
    await c.close()
    print(f"\nCOMMITTED. supplier {sid} ({name}): {before:,} -> {after:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
