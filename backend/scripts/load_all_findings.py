#!/usr/bin/env python3
"""Batch-load every catalog in THE FINDINGS into products.

Dry-run by default; pass --commit to write. Reuses the single-supplier importer
(findings_intake) + upsert SQL (load_findings).

Supplier resolution is future-proof — drop a new catalog xlsx into THE FINDINGS
and it imports with no code change:
  1. If the filename is in MAPPING, use that supplier id (authoritative for the
     known files, so slightly-different names never create duplicates).
  2. Otherwise read the "supplier" column from the file, match an existing
     supplier by name (case-insensitive), and CREATE one if none exists.

    python scripts/load_all_findings.py [--commit]
"""
from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(BACKEND / "scripts"))

import dotenv

dotenv.load_dotenv(BACKEND / ".env")
dotenv.load_dotenv(BACKEND / ".env.dev", override=True)

import asyncpg  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.libs.findings_intake import parse_findings_xlsx  # noqa: E402
from load_findings import build_upsert, dual_price_columns_exist, _values  # noqa: E402

FINDINGS = Path("/Users/justice/Documents/From Selenium To Leaf & Ledger/THE FINDINGS")

# Authoritative filename -> supplier_id for the established catalogs (avoids
# duplicate suppliers from minor name spelling differences). New files not listed
# here are auto-resolved by their "supplier" column.
MAPPING = {
    "AccentDecor_Catalog_2026.xlsx": 15, "AmazingGreen_Catalog_2026.xlsx": 7,
    "AmericanBest_Catalog_2026.xlsx": 4, "AutographFoliages_Catalog_2026.xlsx": 5,
    "Craftex_Catalog_2026.xlsx": 8, "HRCasabella_Catalog_2026.xlsx": 22,
    "JacksonPottery_Catalog_2026.xlsx": 23, "PMJC_Catalog_2026.xlsx": 21,
    "RockWarehouse_Catalog_2026.xlsx": 25, "SelectArtificials_Catalog_2026.xlsx": 2,
    "SuperMoss_Catalog_2026.xlsx": 14, "Vickerman_Catalog_2026.xlsx": 9,
    "Winward_Catalog_2026.xlsx": 6, "allstate_products.xlsx": 1,
    "athome_selected_categories.xlsx": 18, "dfw_vases_products.xlsx": 19,
    "forestline_products.xlsx": 11, "jayscotts_products.xlsx": 20,
    "regency_products.xlsx": 3, "schusters_products.xlsx": 10,
    "second_flor_products.xlsx": 13, "unlimited_container_products.xlsx": 16,
    "wgv_products.xlsx": 17,
}

# xlsx in THE FINDINGS that are not product catalogs
SKIP_SUBSTR = ("categorybreakdown", "pricing_coverage", "readme")


def _is_catalog(path: Path) -> bool:
    if any(s in path.name.lower() for s in SKIP_SUBSTR):
        return False
    try:
        wb = load_workbook(path, read_only=True)
        hdr = [str(h).strip().lower() if h else "" for h in next(wb.active.iter_rows(values_only=True))]
        wb.close()
        return "supplier" in hdr and any(h in ("sku", "supplier sku", "item") for h in hdr)
    except Exception:
        return False


def _supplier_name(path: Path) -> str | None:
    """The 'supplier' cell from the first data row."""
    wb = load_workbook(path, read_only=True)
    it = wb.active.iter_rows(values_only=True)
    hdr = [str(h).strip().lower() if h else "" for h in next(it)]
    idx = hdr.index("supplier") if "supplier" in hdr else None
    name = None
    if idx is not None:
        for row in it:
            if idx < len(row) and row[idx]:
                name = str(row[idx]).strip()
                break
    wb.close()
    return name


async def resolve_supplier(c, path: Path, commit: bool) -> tuple[int | None, str]:
    """Return (supplier_id, status). Creates the supplier on --commit if new."""
    if path.name in MAPPING:
        sid = MAPPING[path.name]
        name = await c.fetchval("SELECT name FROM suppliers WHERE id=$1", sid)
        return (sid, name or "??? (mapped id missing)")
    name = _supplier_name(path)
    if not name:
        return (None, "NO SUPPLIER COLUMN")
    row = await c.fetchrow("SELECT id, name FROM suppliers WHERE name ILIKE $1 ORDER BY id LIMIT 1", name)
    if row:
        return (row["id"], row["name"])
    if commit:
        sid = await c.fetchval(
            "INSERT INTO suppliers (name, credential_status, notes) VALUES ($1,'n/a',$2) RETURNING id",
            name, "auto-created by load_all_findings")
        return (sid, f"{name} (CREATED id={sid})")
    return (-1, f"{name} (NEW — would create)")


async def main() -> int:
    commit = "--commit" in sys.argv
    files = sorted(p for p in FINDINGS.glob("*.xlsx") if _is_catalog(p))
    c = await asyncpg.connect(os.environ["DATABASE_URL"])
    ext = await dual_price_columns_exist(c)
    sql = build_upsert(ext)
    total = 0
    print(f"dual-price columns present: {ext} "
          f"(retail/margin stored {'as columns' if ext else 'in raw_data only'})\n")
    print(f"{'sid':>4}  {'supplier':30s} {'file':40s} {'rows':>7}")
    try:
        for path in files:
            sid, status = await resolve_supplier(c, path, commit)
            rows, rep = parse_findings_xlsx(str(path))
            if sid is None:
                print(f"{'--':>4}  {status[:30]:30s} {path.name[:40]:40s}  SKIP")
                continue
            if commit and sid > 0:
                values = [_values(sid, p, ext) for p in rows]
                async with c.transaction():
                    for i in range(0, len(values), 1000):
                        await c.executemany(sql, values[i:i + 1000])
            total += len(rows)
            print(f"{sid:>4}  {status[:30]:30s} {path.name[:40]:40s} {rep['rows_ok']:>7}")
        print(f"\n{'COMMITTED' if commit else 'DRY RUN'} — {len(files)} files, total rows: {total}")
        if commit:
            gt = await c.fetchval("SELECT COUNT(*) FROM products")
            print(f"products in DB now: {gt}")
            import subprocess
            norm = BACKEND.parent / "catalog-extraction" / "scripts" / "backfill_norm_to_db.py"
            if norm.exists():
                print("\n→ auto-normalizing newly-imported products…")
                subprocess.run([sys.executable, str(norm), "--new-only"], check=False)
        return 0
    finally:
        await c.close()


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
